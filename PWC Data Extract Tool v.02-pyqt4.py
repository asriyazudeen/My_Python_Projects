# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'PWC_UI.ui'
#
# Created by: Riyazudeen Abdul Subhan
#

from PyQt4 import QtCore, QtGui
import sys
import cx_Oracle
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Alignment
import os
import smtplib

#import xlsxwriter 

#dest_loc='/Users/abdulr06/Documents/Python Scripts/'

np.seterr(divide='ignore', invalid='ignore')

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)

class Ui_MainWindow(QtGui.QWidget):

    def __init__(self):
        QtGui.QWidget.__init__(self)
        self.setupUi(self)
    
    
    
    
    def setupUi(self, MainWindow):
        MainWindow.setObjectName(_fromUtf8("MainWindow"))
        MainWindow.resize(1161, 567)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(_fromUtf8("Image.png")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        self.centralWidget = QtGui.QWidget(MainWindow)
        self.centralWidget.setObjectName(_fromUtf8("centralWidget"))
        self.groupBox = QtGui.QGroupBox(self.centralWidget)
        self.groupBox.setGeometry(QtCore.QRect(40, 270, 461, 201))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.groupBox.setFont(font)
        self.groupBox.setObjectName(_fromUtf8("groupBox"))
        self.checkBox = QtGui.QCheckBox(self.groupBox)
        self.checkBox.setGeometry(QtCore.QRect(10, 40, 81, 20))
        self.checkBox.setObjectName(_fromUtf8("checkBox"))
        self.checkBox_2 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_2.setGeometry(QtCore.QRect(10, 70, 81, 20))
        self.checkBox_2.setObjectName(_fromUtf8("checkBox_2"))
        self.checkBox_3 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_3.setGeometry(QtCore.QRect(10, 100, 81, 20))
        self.checkBox_3.setObjectName(_fromUtf8("checkBox_3"))
        self.checkBox_4 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_4.setGeometry(QtCore.QRect(10, 130, 81, 20))
        self.checkBox_4.setObjectName(_fromUtf8("checkBox_4"))
        self.checkBox_5 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_5.setGeometry(QtCore.QRect(130, 40, 81, 20))
        self.checkBox_5.setObjectName(_fromUtf8("checkBox_5"))
        self.checkBox_6 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_6.setGeometry(QtCore.QRect(130, 70, 81, 20))
        self.checkBox_6.setObjectName(_fromUtf8("checkBox_6"))
        self.checkBox_7 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_7.setGeometry(QtCore.QRect(130, 100, 81, 20))
        self.checkBox_7.setObjectName(_fromUtf8("checkBox_7"))
        self.checkBox_8 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_8.setGeometry(QtCore.QRect(130, 130, 81, 20))
        self.checkBox_8.setObjectName(_fromUtf8("checkBox_8"))
        self.checkBox_9 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_9.setGeometry(QtCore.QRect(10, 160, 81, 20))
        self.checkBox_9.setObjectName(_fromUtf8("checkBox_9"))
        self.checkBox_10 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_10.setGeometry(QtCore.QRect(130, 160, 81, 20))
        self.checkBox_10.setObjectName(_fromUtf8("checkBox_10"))
        self.checkBox_11 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_11.setGeometry(QtCore.QRect(260, 70, 211, 20))
        self.checkBox_11.setObjectName(_fromUtf8("checkBox_11"))
        self.checkBox_12 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_12.setGeometry(QtCore.QRect(260, 100, 191, 20))
        self.checkBox_12.setObjectName(_fromUtf8("checkBox_12"))
        self.checkBox_23 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_23.setGeometry(QtCore.QRect(260, 40, 81, 20))
        self.checkBox_23.setObjectName(_fromUtf8("checkBox_23"))
        self.checkBox_24 = QtGui.QCheckBox(self.groupBox)
        self.checkBox_24.setGeometry(QtCore.QRect(260, 130, 191, 20))
        self.checkBox_24.setObjectName(_fromUtf8("checkBox_24"))
        self.groupBox.setEnabled(False)
        self.lineEdit = QtGui.QLineEdit(self.centralWidget)
        self.lineEdit.setEnabled(False)
        self.lineEdit.setGeometry(QtCore.QRect(370, 180, 271, 22))
#        self.lineEdit.setClearButtonEnabled(True)
        self.lineEdit.setObjectName(_fromUtf8("lineEdit"))
        self.label = QtGui.QLabel(self.centralWidget)
        self.label.setEnabled(False)
        self.label.setGeometry(QtCore.QRect(120, 180, 211, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName(_fromUtf8("label"))
        self.pushButton = QtGui.QPushButton(self.centralWidget)
        self.pushButton.setEnabled(False)
        self.pushButton.setGeometry(QtCore.QRect(680, 177, 93, 31))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName(_fromUtf8("pushButton"))
        self.pushButton_2 = QtGui.QPushButton(self.centralWidget)
        self.pushButton_2.setEnabled(False)
        self.pushButton_2.setGeometry(QtCore.QRect(810, 177, 93, 31))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName(_fromUtf8("pushButton_2"))
        self.groupBox_2 = QtGui.QGroupBox(self.centralWidget)
        self.groupBox_2.setGeometry(QtCore.QRect(630, 270, 491, 201))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.groupBox_2.setFont(font)
        self.groupBox_2.setObjectName(_fromUtf8("groupBox_2"))
        self.checkBox_13 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_13.setGeometry(QtCore.QRect(20, 40, 81, 20))
        self.checkBox_13.setObjectName(_fromUtf8("checkBox_13"))
        self.checkBox_14 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_14.setGeometry(QtCore.QRect(20, 70, 81, 20))
        self.checkBox_14.setObjectName(_fromUtf8("checkBox_14"))
        self.checkBox_15 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_15.setGeometry(QtCore.QRect(20, 100, 81, 20))
        self.checkBox_15.setObjectName(_fromUtf8("checkBox_15"))
        self.checkBox_16 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_16.setGeometry(QtCore.QRect(20, 130, 221, 20))
        self.checkBox_16.setObjectName(_fromUtf8("checkBox_16"))
        self.checkBox_17 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_17.setGeometry(QtCore.QRect(20, 160, 221, 20))
        self.checkBox_17.setObjectName(_fromUtf8("checkBox_17"))
        self.checkBox_18 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_18.setGeometry(QtCore.QRect(260, 40, 221, 20))
        self.checkBox_18.setObjectName(_fromUtf8("checkBox_18"))
        self.checkBox_19 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_19.setGeometry(QtCore.QRect(260, 70, 221, 20))
        self.checkBox_19.setObjectName(_fromUtf8("checkBox_19"))
        self.checkBox_20 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_20.setGeometry(QtCore.QRect(260, 100, 221, 20))
        self.checkBox_20.setObjectName(_fromUtf8("checkBox_20"))
        self.checkBox_21 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_21.setGeometry(QtCore.QRect(260, 130, 221, 20))
        self.checkBox_21.setObjectName(_fromUtf8("checkBox_21"))
        self.checkBox_22 = QtGui.QCheckBox(self.groupBox_2)
        self.checkBox_22.setGeometry(QtCore.QRect(260, 160, 331, 20))
        self.checkBox_22.setObjectName(_fromUtf8("checkBox_22"))
        self.groupBox_2.setEnabled(False)
        self.lineEdit_2 = QtGui.QLineEdit(self.centralWidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(150, 500, 281, 22))
        self.lineEdit_2.setObjectName(_fromUtf8("lineEdit_2"))
        self.lineEdit_2.setEnabled(False)
        self.label_2 = QtGui.QLabel(self.centralWidget)
        self.label_2.setGeometry(QtCore.QRect(40, 502, 131, 16))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName(_fromUtf8("label_2"))
        self.label_2.setEnabled(False)
        self.groupBox_3 = QtGui.QGroupBox(self.centralWidget)
        #changed value here
        self.groupBox_3.setGeometry(QtCore.QRect(20, 40, 1131, 81))
        self.groupBox_3.setObjectName(_fromUtf8("groupBox_3"))
        self.label_3 = QtGui.QLabel(self.groupBox_3)
        self.label_3.setGeometry(QtCore.QRect(10, 40, 71, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName(_fromUtf8("label_3"))      
        self.label_4 = QtGui.QLabel(self.groupBox_3)
        self.label_4.setGeometry(QtCore.QRect(260, 40, 161, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName(_fromUtf8("label_4"))
        self.label_5 = QtGui.QLabel(self.groupBox_3)
        self.label_5.setGeometry(QtCore.QRect(600, 40, 91, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setObjectName(_fromUtf8("label_5"))
        self.lineEdit_3 = QtGui.QLineEdit(self.groupBox_3)
        self.lineEdit_3.setGeometry(QtCore.QRect(90, 40, 151, 22))
        self.lineEdit_3.setObjectName(_fromUtf8("lineEdit_3"))
        self.lineEdit_4 = QtGui.QLineEdit(self.groupBox_3)
        self.lineEdit_4.setGeometry(QtCore.QRect(430, 40, 151, 22))
        self.lineEdit_4.setObjectName(_fromUtf8("lineEdit_4"))
        self.lineEdit_5 = QtGui.QLineEdit(self.groupBox_3)
        self.lineEdit_5.setGeometry(QtCore.QRect(700, 40, 151, 22))
        self.lineEdit_5.setObjectName(_fromUtf8("lineEdit_5"))
        self.lineEdit_5.setEchoMode(QtGui.QLineEdit.Password)
        self.pushButton_3 = QtGui.QPushButton(self.groupBox_3)
        self.pushButton_3.setGeometry(QtCore.QRect(900, 37, 93, 31))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName(_fromUtf8("pushButton_3"))
        self.pushButton_4 = QtGui.QPushButton(self.groupBox_3)
        self.pushButton_4.setGeometry(QtCore.QRect(1020, 37, 93, 31))
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setObjectName(_fromUtf8("pushButton_4"))
        self.label_6 = QtGui.QLabel(self.centralWidget)
        self.label_6.setGeometry(QtCore.QRect(430, 10, 321, 21))
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Arial"))
        font.setPointSize(14)
        font.setItalic(True)
        self.label_6.setFont(font)
        self.label_6.setObjectName(_fromUtf8("label_6"))
        self.label_7 = QtGui.QLabel(self.centralWidget)
        self.label_7.setGeometry(QtCore.QRect(490, 500, 101, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setObjectName(_fromUtf8("label_7"))
        self.label_7.setEnabled(False)
        self.lineEdit_6 = QtGui.QLineEdit(self.centralWidget)
        self.lineEdit_6.setGeometry(QtCore.QRect(600, 500, 521, 22))
        self.lineEdit_6.setObjectName(_fromUtf8("lineEdit_6"))
        self.lineEdit_6.setEnabled(False)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(_translate("MainWindow", "EDM PWC Automation Tool v.02", None))
        self.groupBox.setTitle(_translate("MainWindow", "Loans", None))
        self.checkBox.setText(_translate("MainWindow", "ALS", None))
        self.checkBox_2.setText(_translate("MainWindow", "AFS", None))
        self.checkBox_3.setText(_translate("MainWindow", "DF", None))
        self.checkBox_4.setText(_translate("MainWindow", "FD", None))
        self.checkBox_5.setText(_translate("MainWindow", "IMOD", None))
        self.checkBox_6.setText(_translate("MainWindow", "LS", None))
        self.checkBox_7.setText(_translate("MainWindow", "LV", None))
        self.checkBox_8.setText(_translate("MainWindow", "ML", None))
        self.checkBox_9.setText(_translate("MainWindow", "MO", None))
        self.checkBox_10.setText(_translate("MainWindow", "MS", None))
        self.checkBox_11.setText(_translate("MainWindow", "533975", None))
        self.checkBox_12.setText(_translate("MainWindow", "534070", None))
        self.checkBox_23.setText(_translate("MainWindow", "TSYS", None))
        self.checkBox_24.setText(_translate("MainWindow", "533901", None))
        self.lineEdit.setText(_translate("MainWindow", "YYYYDDMM", None))
#        self.lineEdit.setText(_translate("MainWindow", "20170929", None))
        self.label.setText(_translate("MainWindow", "Enter the Cycle Date\'s", None))
        self.pushButton.setText(_translate("MainWindow", "Execute", None))
        self.pushButton_2.setText(_translate("MainWindow", "Exit", None))
        self.groupBox_2.setTitle(_translate("MainWindow", "Deposits", None))
        self.checkBox_13.setText(_translate("MainWindow", "ST", None))
        self.checkBox_14.setText(_translate("MainWindow", "IM", None))
        self.checkBox_15.setText(_translate("MainWindow", "LG", None))
        self.checkBox_16.setText(_translate("MainWindow", "550139", None))
        self.checkBox_17.setText(_translate("MainWindow", "550141", None))
        self.checkBox_18.setText(_translate("MainWindow", "550557", None))
        self.checkBox_19.setText(_translate("MainWindow", "550569 ", None))
        self.checkBox_20.setText(_translate("MainWindow", "552872", None))
        self.checkBox_21.setText(_translate("MainWindow", "552877", None))
        self.checkBox_22.setText(_translate("MainWindow", "550129 and 550710", None))
        self.lineEdit_2.setText(_translate("MainWindow", "/PWC/Data Extract/", None))
        self.label_2.setText(_translate("MainWindow", "File Location", None))
        self.groupBox_3.setTitle(_translate("MainWindow", "Login", None))
        self.label_3.setText(_translate("MainWindow", "User ID", None))
        self.label_4.setText(_translate("MainWindow", "TNS Service Name", None))
        self.label_5.setText(_translate("MainWindow", "Password", None))
        self.pushButton_3.setText(_translate("MainWindow", "Connect", None))
        self.pushButton_4.setText(_translate("MainWindow", "Clear", None))
        self.label_6.setText(_translate("MainWindow", "PWC Data Extraction Tool v.02", None))
        self.label_7.setText(_translate("MainWindow", "Send Email", None))
        self.lineEdit_6.setText(_translate("MainWindow", "xxxxxxxx@BankoftheWest.com", None))
        
        self.pushButton_3.clicked.connect(self.Oracle) 
        self.pushButton_4.clicked.connect(self.Clear) 
        
        
        
        
    def Clear(self)    :
        
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
        self.lineEdit_5.clear()
        
        
    def Oracle(self):
        """Connect to the database"""
        username=self.lineEdit_3.text()
        servicename=self.lineEdit_4.text()
        password=self.lineEdit_5.text()
        
      
        
        print('Connecting please wait....')
        
        try:
            self.db=cx_Oracle.connect(username,password,servicename)
            self.cursor=self.db.cursor() 
        except cx_Oracle.DatabaseError as e:
            print(e)
            error, = e.args
            if error.code == 1017:
                QtGui.QMessageBox.critical(self, 'Login Failed!',
                                            "Please check your credentials")
                
            else:                
#                QtGui.QMessageBox.critical(self, 'Login Failed!',
#                                            "Database connection errror: ".format(e))
                QtGui.QMessageBox.critical(self, 'DB Error!','Database connection error')                             
            raise
               
        
        self.label.setEnabled(True)
        self.lineEdit.setEnabled(True)
        self.pushButton.setEnabled(True)
        self.pushButton_2.setEnabled(True)
        self.groupBox_3.setEnabled(False)
        self.groupBox_2.setEnabled(True)
        self.groupBox.setEnabled(True)
        self.label_2.setEnabled(True)      
        self.label_7.setEnabled(True)
        self.lineEdit_2.setEnabled(True)
        self.lineEdit_6.setEnabled(True)        
        print("Connected Sucessfully")     
        
        self.pushButton.clicked.connect(self.path)        
        self.pushButton_2.clicked.connect(QtCore.QCoreApplication.instance().quit)    

    def disconnect(self):
        """Disconnect from the Database"""
        try:
            self.cursor.close()
            self.db.close()
        except cx_Oracle.DatabaseError:
            pass

    def path(self):
        
        path=self.lineEdit_2.text()
        
        if path.endswith('/'):        
            if os.path.exists(path):
                self.check()
            else:
                QtGui.QMessageBox.warning(self, 'Warning!',"Plese enter the correct file path")
        else:
             QtGui.QMessageBox.warning(self, 'Warning!',"File path should end with / ")
             
    def Email(self)             :
        smtpObj=smtplib.SMTP('SMTPINT1.bankofthewest.com') 
        mail=self.lineEdit_6.text()
        smtpObj.sendmail('PWC Extraction Status',mail,'Subject: PWC Extraction Status\nDear User,\n\nPWC data extraction is completed successfully, please check the output folder for the results. \n\nRegards,\nPWC Admin')
        smtpObj.quit()
    
    def Email_Error(self,e, SRC)             :
      
        smtpObj=smtplib.SMTP('SMTPINT1.bankofthewest.com') 
        part1='Subject: PWC Failure '     
        message= part1 + SRC+ e
        mail=self.lineEdit_6.text()
        smtpObj.sendmail('PWC Extraction failed',mail, message)
        smtpObj.quit()         
        
    def check(self):
        if  self.checkBox.isChecked() == False and self.checkBox_2.isChecked() == False and self.checkBox_3.isChecked() == False and self.checkBox_4.isChecked() == False and self.checkBox_5.isChecked() == False and self.checkBox_6.isChecked() == False and self.checkBox_7.isChecked() == False and self.checkBox_8.isChecked() == False and self.checkBox_9.isChecked() == False and self.checkBox_10.isChecked() == False and self.checkBox_11.isChecked() == False and self.checkBox_12.isChecked() == False and self.checkBox_13.isChecked() == False and self.checkBox_14.isChecked() == False and self.checkBox_15.isChecked() == False and self.checkBox_16.isChecked() == False and self.checkBox_17.isChecked() == False and self.checkBox_18.isChecked() == False and self.checkBox_19.isChecked() == False and self.checkBox_20.isChecked() == False and self.checkBox_21.isChecked() == False and self.checkBox_22.isChecked() == False and self.checkBox_23.isChecked() == False and self.checkBox_24.isChecked() == False:
            QtGui.QMessageBox.warning(self, 'Warning!',"Please select atleast one CheckBox")
            return
        else:
            self.SQL()
             
    def execution(self,dt1, YM1, SRC, SQL):

             SQL=SQL.replace(':dt1_cycle_dt_num',dt1)             
             SQL=SQL.replace(':dt1_year_month',YM1)  

             try:             
                 df_SQL = pd.io.sql.read_sql(SQL, self.db)  
                  
             except cx_Oracle.DatabaseError as e:     
   
                 self.Email_Error(str(e), SRC)  
                 return                  
                 
             cx_Oracle.connect.close
             return df_SQL      

    def SQL(self):      
        
        if len(self.lineEdit.text()) < 8:        
            QtGui.QMessageBox.warning(self, 'Warning!',"Please enter a valid DATE and click Execute")
            return  
            
        dt=self.lineEdit.text()
        dt=dt.split(',')   
        
        i=0          
        while i < len(dt):
            dt1=dt[i].strip()
            YM1=str(dt1[0:6])            
            print('Execution Begins for the Cycle Date: '+ dt1)         



            SQL_GL="SELECT A.YEAR_MONTH,        A.COMPANY_NUM,        TO_NUMBER (A.GL_ACCT_NUM) AS GL_ACCT_NUM,        SUM (A.MTD_GL_VAL) AS MTD_GL_VAL   FROM GLM.FACT_GL_MONTHLY_SUMMARY A  WHERE     A.YEAR_MONTH IN (SELECT YEAR_MONTH2                               FROM edr.period                              WHERE     year_month2 <= :dt1_year_month                                    AND year_NM =                                           SUBSTR ( :dt1_year_month, 1, 4)                             GROUP BY year_month2)        AND A.GL_MEASURE_KEY = 100        AND A.SCENARIO_GROUP = 'ACTUAL_GL' GROUP BY A.YEAR_MONTH, A.GL_ACCT_NUM, A.COMPANY_NUM"
            SRC='GL_SUMM'
            print("STEP1: Extracting Data from GL Summary table")
            GL_SUMM_fee=self.execution(dt1,YM1,SRC,SQL_GL)
            GL_SUMM=GL_SUMM_fee[GL_SUMM_fee['YEAR_MONTH'] == int(YM1)]
#            print(GL_SUMM.head())
            
#            GL_SUMM=self.execution(dt1,YM1,SRC,SQL_GL)       
#            GL_SUMM_fee=self.GL_SUMMARY_fee(YM1)
            
            if self.checkBox.isChecked() == True:                
                SRC='ALS'           
                SQL="WITH ALS_INSTR      AS (SELECT DISTINCT                A.SRC_SYS_CD,                 A.YEAR_MONTH,                 C.CUST_NUM,                 C.GU_FULL_NAME AS CUSTOMER_NAME,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 A.LAST_PAYMENT_DT,                 D.STATUS_DESC AS LOAN_STATUS,                 D.ACCRUAL_STATUS_DESC AS LOAN_ACCRUAL_STATUS,                 E.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD,                 A.ORIGINAL_INTEREST_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_UNPAID_BAL,                 A.INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 A.LAST_PAYMENT_AMT,                 E.INSTR_TYPE_CD_DESC,                 A.PRINCIPAL_BAL_GL_ACCT_NUM,                 F.GL_INTEREST_INCOME,                 F.GL_ACCR_INTEREST_RECEIVABLE,                 E.INSTR_PROD_CD_DESC,                 A.GL_NA_CD,                 A.TOTAL_PAST_DUE_DAYS            FROM EDR.FACT_LOAN_INSTR_MONTHLY A                 LEFT JOIN EDR.CUST_INSTR_RELATION B                    ON     A.SRC_INSTR_ID = B.ACCT_NUM_DERIVED                       AND B.SRC_SYS_CD = 'RM'                       AND A.SRC_SYS_CD = B.APPLCN_CD_DERIVED                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                              B.BEGIN_EFF_DTTM                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                              NVL (B.END_EFF_DTTM, '31-DEC-9999')                       AND A.RELATIONSHIP_MANAGER_NUM = B.CUST_NUM                       AND B.DELETE_INDICATOR = 'N'                 LEFT JOIN EDR.DIM_CUSTOMER C                    ON     B.CUST_NUM = C.CUST_NUM                       AND C.SRC_SYS_CD = 'RM'                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                              C.BEGIN_EFF_DTTM                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                              NVL (C.END_EFF_DTTM, '31-DEC-9999')                       AND C.DELETE_INDICATOR = 'N'                       AND A.RELATIONSHIP_MANAGER_NUM = C.CUST_NUM                 LEFT OUTER JOIN EDR.DIM_STATUS D                    ON     A.DIM_STATUS_KEY = D.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY E                    ON     A.DIM_PRODUCT_KEY = E.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = E.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_ALS_GL_MAP F                    ON     A.GL_MAP_KEY = F.DIM_ALS_GL_MAP_KEY                       AND A.PRINCIPAL_BAL_GL_ACCT_NUM = F.GL_PRINCIPAL                       AND A.GL_INVESTOR_CD = F.GL_INVESTOR_CD                       AND A.GL_NA_CD = F.GL_NA_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'ALS'                 AND E.CURRENT_FLG = 'Y'                 AND E.LEVEL_NUM IS NULL                 AND D.CURRENT_FLG = 'Y'                 AND F.CURRENT_FLG = 'Y'                 AND (A.INTEREST_INCOME <> 0 OR A.PRINCIPAL_UNPAID_BAL <> 0)                     ) SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        SUM (A.PRINCIPAL_UNPAID_BAL) AS TOTAL_ENDING_BAL,        SUM (A.INTEREST_INCOME) AS TOTAL_INT_INCOME,        A.PRINCIPAL_BAL_GL_ACCT_NUM AS GL_ACCT_NUM,        A.GL_INTEREST_INCOME   FROM ALS_INSTR A GROUP BY A.SRC_SYS_CD,          A.YEAR_MONTH,          A.PRINCIPAL_BAL_GL_ACCT_NUM,          A.GL_INTEREST_INCOME"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1)  
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)  
                Data_INT=1
                Data_PAR=1
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM) 
                
            if self.checkBox_2.isChecked() == True:
                
                SRC='AFS'              
                SQL="WITH AF_INSTR AS ( SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.OBLIGOR,        A.OBLIGATION,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        CASE WHEN A.AF_COLLATERAL_CD IN ('000','055')           THEN 'N'           ELSE 'Y'        END COLLATERAL,        A.ACCRUAL_METHOD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        A.PRINCIPAL_BAL_GL_ACCT_NUM,        B.ACCRUAL_STATUS_DESC,        C.INSTR_TYPE_CD_DESC,        C.INSTR_PROD_CD_DESC,        C.INSTR_INT_TYPE_CD_DESC,        D.GL_INTEREST_INCOME,        D.GL_ACCR_INTEREST_RECEIVABLE,        B.STATUS_CD,        B.STATUS_DESC,        TO_NUMBER(NVL(E.COST_CENTER,'0')) AS COST_CENTER,        TO_NUMBER(NVL(E.COMPANY,'0')) AS COMPANY,        A.AF_FED_CLASS   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN        EDR.DIM_STATUS B     ON A.DIM_STATUS_KEY = B.DIM_STATUS_KEY    AND A.SRC_SYS_CD = B.SRC_SYS_CD    AND B.CURRENT_FLG = 'Y'    AND B.STATUS_CD = 'Y'   LEFT OUTER JOIN        EDR.DIM_PRODUCT_HIERARCHY C     ON A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD    AND C.CURRENT_FLG = 'Y'    AND C.HIERARCHY_TYPE = 'DEFAULT'   LEFT OUTER JOIN        EDR.DIM_AF_GL_MAP D     ON A.DIM_GL_MAP_KEY = D.DIM_AF_GL_MAP_KEY    AND A.PRINCIPAL_BAL_GL_ACCT_NUM = D.GL_PRINCIPAL    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_ORG E     ON A.DIM_ORG_KEY = E.DIM_ORG_KEY    AND E.CURRENT_FLG = 'Y'  WHERE A.SRC_SYS_CD = 'AF'    AND A.YEAR_MONTH = :dt1_year_month    AND A.PRINCIPAL_ENDING_BAL <> 0    ORDER BY A.SRC_INSTR_ID ), AF_CL999 AS ( SELECT CYCLE_DT_NUM,        LPAD(CL999_OBLIGOR,10,'0') || '-' || LPAD(CL999_OBLIGATION,5,'0') AS SRC_INSTR_ID,        CL999_OBLIGOR,        CL999_OBLIGATION,        CL999_SHORT_NAME AS CUSTOMER_NAME,        CL999_EFFECTIVE_DATE,        CL999_MATURITY_DATE,        CL999_ORIGINAL_BALANCE,        CL999_CURRENT_BALANCE,        CL999_BANK_BALANCE,        CL999_INT_PAID_TO_DATE AS INT_PD_TO_DATE,        CL999_PRIN_BILL_DATE AS LAST_STMT_DATE,        CL999_PAYMENT_AMOUNT AS LOAN_PAYMENT_AMOUNT,        CL999_COB_PASTDUE_DAYS AS PASTDUE_DAYS   FROM DATASTORE_AFS.CL999  WHERE CYCLE_DT_NUM = :dt1_cycle_dt_num    AND CL999_CURRENT_BALANCE > 0  ORDER BY SRC_INSTR_ID ) SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.OBLIGOR,        A.OBLIGATION,        A.SRC_INSTR_ID,        B.CUSTOMER_NAME,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        B.INT_PD_TO_DATE,        A.COLLATERAL,        A.ACCRUAL_STATUS_DESC,        A.INSTR_TYPE_CD_DESC,        A.ACCRUAL_METHOD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL AS TOTAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        B.LAST_STMT_DATE,        B.LOAN_PAYMENT_AMOUNT,        A.INSTR_INT_TYPE_CD_DESC,        A.PRINCIPAL_BAL_GL_ACCT_NUM AS GL_ACCT_NUM,        A.GL_INTEREST_INCOME,        A.GL_ACCR_INTEREST_RECEIVABLE,        A.INSTR_PROD_CD_DESC,        B.PASTDUE_DAYS,        A.STATUS_CD,        A.STATUS_DESC,        A.COST_CENTER,        A.COMPANY,        A.AF_FED_CLASS   FROM AF_INSTR A   LEFT OUTER JOIN        AF_CL999 B     ON A.SRC_INSTR_ID = B.SRC_INSTR_ID "
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1)  
                Data_BAL=self.execution(dt1,YM1,SRC,SQL) 
                
                """Following is AFS INTR SQL"""                    
                SQL="WITH AF_INSTR      AS (SELECT A.YEAR_MONTH,                 A.COMPANY_NUM,                 A.SRC_SYS_CD,                 A.OBLIGOR,                 A.OBLIGATION,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 CASE                    WHEN A.AF_COLLATERAL_CD IN ('000', '055') THEN 'N'                    ELSE 'Y'                 END                    COLLATERAL,                 A.ACCRUAL_METHOD,                 A.ORIGINAL_INT_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL,                 A.INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 A.PRINCIPAL_BAL_GL_ACCT_NUM,                 B.ACCRUAL_STATUS_DESC,                 C.INSTR_TYPE_CD_DESC,                 C.INSTR_PROD_CD_DESC,                 C.INSTR_INT_TYPE_CD_DESC,                 D.GL_INTEREST_INCOME,                 D.GL_ACCR_INTEREST_RECEIVABLE,                 B.STATUS_CD,                 B.STATUS_DESC,                 TO_NUMBER (NVL (E.COST_CENTER, '0')) AS COST_CENTER,                 TO_NUMBER (NVL (E.COMPANY, '0')) AS COMPANY            FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A                 LEFT OUTER JOIN EDR.DIM_STATUS B                    ON     A.DIM_STATUS_KEY = B.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = B.SRC_SYS_CD                       AND B.CURRENT_FLG = 'Y'                       AND B.STATUS_CD = 'Y'                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY C                    ON     A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                       AND C.CURRENT_FLG = 'Y'                       AND C.HIERARCHY_TYPE = 'DEFAULT'                 LEFT OUTER JOIN EDR.DIM_AF_GL_MAP D                    ON     A.DIM_GL_MAP_KEY = D.DIM_AF_GL_MAP_KEY                       AND A.PRINCIPAL_BAL_GL_ACCT_NUM = D.GL_PRINCIPAL                       AND D.CURRENT_FLG = 'Y'                 LEFT OUTER JOIN EDR.DIM_ORG E                    ON A.DIM_ORG_KEY = E.DIM_ORG_KEY AND E.CURRENT_FLG = 'Y'           WHERE A.SRC_SYS_CD = 'AF' AND A.YEAR_MONTH = :dt1_year_month          ORDER BY A.SRC_INSTR_ID),      AF_CL999      AS (SELECT CYCLE_DT_NUM,                    LPAD (CL999_OBLIGOR, 10, '0')                 || '-'                 || LPAD (CL999_OBLIGATION, 5, '0')                    AS SRC_INSTR_ID,                 CL999_OBLIGOR,                 CL999_OBLIGATION,                 CL999_SHORT_NAME AS CUSTOMER_NAME,                 CL999_EFFECTIVE_DATE,                 CL999_MATURITY_DATE,                 CL999_ORIGINAL_BALANCE,                 CL999_CURRENT_BALANCE,                 CL999_BANK_BALANCE,                 CL999_INT_PAID_TO_DATE AS INT_PD_TO_DATE,                 CL999_PRIN_BILL_DATE AS LAST_STMT_DATE,                 CL999_PAYMENT_AMOUNT AS LOAN_PAYMENT_AMOUNT,                 CL999_COB_PASTDUE_DAYS AS PASTDUE_DAYS            FROM DATASTORE_AFS.CL999           WHERE CYCLE_DT_NUM = :dt1_cycle_dt_num          ORDER BY SRC_INSTR_ID),      PARTSOLD_TO_INSTR_RELTN      AS (SELECT LI.SRC_INSTR_ID,                 SR.SRC_INSTR_ID PARTSOLD_TO_INSTR_ID,                 LI.PRINCIPAL_BAL_GL_ACCT_NUM AS GL_ACCT_NUM            FROM EDR.FACT_COM_LOAN_INSTR_DAILY LI,                 EDR.FACT_PART_SYND_REFERENCE_DAILY SR           WHERE     LI.CYCLE_DT_NUM = :dt1_cycle_dt_num                 AND SR.CYCLE_DT_NUM = :dt1_cycle_dt_num                 AND LI.SRC_SYS_CD = 'AF'                 AND LI.SRC_INSTR_ID = SR.REF_SRC_INSTR_ID                 AND SUBSTR (SR.PROCESS_TYPE, 2, 1) = 3),      PARTSOLD_TO_INSTR      AS (SELECT SUBSTR (A.CYCLE_DT_NUM, 1, 6) AS YEAR_MONTH,                 A.SRC_SYS_CD,                 A.SRC_INSTR_ID,                 A.PRINCIPAL_BAL_GL_ACCT_NUM AS PARTSOLD_TO_GL_ACCT_NUM,                 A.PRINCIPAL_ENDING_BAL,                 B.GL_INTEREST_INCOME,                 B.GL_INTEREST_EXP_PARTICPATED,                 A.AF_INT_AMT_MTD            FROM EDR.FACT_COM_LOAN_INSTR_DAILY A, EDR.DIM_AF_GL_MAP B           WHERE     A.CYCLE_DT_NUM = :dt1_cycle_dt_num                 AND A.SRC_SYS_CD = 'AF'                 AND A.SRC_INSTR_ID IN (SELECT SR.SRC_INSTR_ID                                          FROM EDR.FACT_COM_LOAN_INSTR_DAILY LI,                                               EDR.FACT_PART_SYND_REFERENCE_DAILY SR                                         WHERE     LI.CYCLE_DT_NUM =                                                      :dt1_cycle_dt_num                                               AND SR.CYCLE_DT_NUM =                                                      :dt1_cycle_dt_num                                               AND LI.SRC_INSTR_ID =                                                      SR.REF_SRC_INSTR_ID                                               AND SUBSTR (SR.PROCESS_TYPE,                                                           2,                                                           1) = 3)                 AND A.AF_INT_AMT_MTD <> 0                 AND A.DIM_GL_MAP_KEY = B.DIM_AF_GL_MAP_KEY),      PARTSOLD_TO_DETAIL      AS (SELECT B.YEAR_MONTH,                 B.SRC_SYS_CD,                 A.SRC_INSTR_ID,                 A.PARTSOLD_TO_INSTR_ID,                 A.GL_ACCT_NUM,                 B.PARTSOLD_TO_GL_ACCT_NUM,                 B.PRINCIPAL_ENDING_BAL AS PARTSOLD_TO_ENDING_BAL,                 B.GL_INTEREST_INCOME AS PARTSOLD_TO_GL_INTEREST_INCOME,                 B.GL_INTEREST_EXP_PARTICPATED AS PARTSOLD_TO_GL_INT_EXP_PART,                 B.AF_INT_AMT_MTD AS PARTSOLD_TO_INTEREST_INCOME            FROM PARTSOLD_TO_INSTR_RELTN A                 LEFT JOIN PARTSOLD_TO_INSTR B                    ON A.PARTSOLD_TO_INSTR_ID = B.SRC_INSTR_ID           WHERE B.AF_INT_AMT_MTD <> 0),      PARTSOLD_TO_SUM      AS (SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.SRC_INSTR_ID,                 A.GL_ACCT_NUM,                 A.PARTSOLD_TO_GL_INTEREST_INCOME,                 A.PARTSOLD_TO_GL_INT_EXP_PART,                 NVL (SUM (A.PARTSOLD_TO_INTEREST_INCOME), 0)                    AS PARTSOLD_TO_TOTAL_INT_INCOME            FROM PARTSOLD_TO_DETAIL A          GROUP BY A.YEAR_MONTH,                   A.SRC_SYS_CD,                   A.SRC_INSTR_ID,                   A.GL_ACCT_NUM,                   A.PARTSOLD_TO_GL_INTEREST_INCOME,                   A.PARTSOLD_TO_GL_INT_EXP_PART          HAVING SUM (A.PARTSOLD_TO_INTEREST_INCOME) <> 0          ORDER BY A.SRC_INSTR_ID,                   A.GL_ACCT_NUM,                   A.PARTSOLD_TO_GL_INT_EXP_PART) SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.OBLIGATION,        A.SRC_INSTR_ID,        B.CUSTOMER_NAME,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        B.INT_PD_TO_DATE,        A.COLLATERAL,        A.ACCRUAL_STATUS_DESC,        A.INSTR_TYPE_CD_DESC,        A.ACCRUAL_METHOD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        B.LAST_STMT_DATE,        B.LOAN_PAYMENT_AMOUNT,        A.INSTR_INT_TYPE_CD_DESC,        A.PRINCIPAL_BAL_GL_ACCT_NUM,        A.GL_INTEREST_INCOME,        A.GL_ACCR_INTEREST_RECEIVABLE,        A.INSTR_PROD_CD_DESC,        B.PASTDUE_DAYS,        C.PARTSOLD_TO_GL_INT_EXP_PART,        NVL (C.PARTSOLD_TO_TOTAL_INT_INCOME, 0),        A.INTEREST_INCOME - NVL (C.PARTSOLD_TO_TOTAL_INT_INCOME, 0)           AS TOTAL_INT_INCOME,        A.STATUS_CD,        A.STATUS_DESC,        A.COST_CENTER,        A.COMPANY   FROM AF_INSTR A        LEFT OUTER JOIN AF_CL999 B ON A.SRC_INSTR_ID = B.SRC_INSTR_ID        LEFT OUTER JOIN PARTSOLD_TO_SUM C ON A.SRC_INSTR_ID = C.SRC_INSTR_ID  WHERE (A.INTEREST_INCOME <> 0 OR C.PARTSOLD_TO_TOTAL_INT_INCOME IS NOT NULL)"
                print(SRC +' Intrest Income Query execution is in progress for ' +  dt1) 
                Data_INT=self.execution(dt1,YM1,SRC,SQL)  
                
                """Following is AFS Part Sold QuerySQL"""      
                SQL="WITH AF_INSTR AS ( SELECT A.YEAR_MONTH,        A.COMPANY_NUM,        A.SRC_SYS_CD,        A.OBLIGOR,        A.OBLIGATION,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        CASE WHEN A.AF_COLLATERAL_CD IN ('000','055')           THEN 'N'           ELSE 'Y'        END COLLATERAL,        A.ACCRUAL_METHOD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        A.PRINCIPAL_BAL_GL_ACCT_NUM,        B.ACCRUAL_STATUS_DESC,        C.INSTR_TYPE_CD_DESC,        C.INSTR_PROD_CD_DESC,        C.INSTR_INT_TYPE_CD_DESC,        D.GL_INTEREST_INCOME,        D.GL_ACCR_INTEREST_RECEIVABLE,        B.STATUS_CD,        B.STATUS_DESC,        TO_NUMBER(NVL(E.COST_CENTER,'0')) AS COST_CENTER,        TO_NUMBER(NVL(E.COMPANY,'0')) AS COMPANY   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN        EDR.DIM_STATUS B     ON A.DIM_STATUS_KEY = B.DIM_STATUS_KEY    AND A.SRC_SYS_CD = B.SRC_SYS_CD    AND B.CURRENT_FLG = 'Y'    AND B.STATUS_CD = 'Y'   LEFT OUTER JOIN        EDR.DIM_PRODUCT_HIERARCHY C     ON A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD    AND C.CURRENT_FLG = 'Y'    AND C.HIERARCHY_TYPE = 'DEFAULT'   LEFT OUTER JOIN        EDR.DIM_AF_GL_MAP D     ON A.DIM_GL_MAP_KEY = D.DIM_AF_GL_MAP_KEY    AND A.PRINCIPAL_BAL_GL_ACCT_NUM = D.GL_PRINCIPAL    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_ORG E     ON A.DIM_ORG_KEY = E.DIM_ORG_KEY    AND E.CURRENT_FLG = 'Y'  WHERE A.SRC_SYS_CD = 'AF'    AND A.YEAR_MONTH = :dt1_year_month  ORDER BY A.SRC_INSTR_ID ), AF_CL999 AS ( SELECT CYCLE_DT_NUM,        LPAD(CL999_OBLIGOR,10,'0') || '-' || LPAD(CL999_OBLIGATION,5,'0') AS SRC_INSTR_ID,        CL999_OBLIGOR,        CL999_OBLIGATION,        CL999_SHORT_NAME AS CUSTOMER_NAME,        CL999_EFFECTIVE_DATE,        CL999_MATURITY_DATE,        CL999_ORIGINAL_BALANCE,        CL999_CURRENT_BALANCE,        CL999_BANK_BALANCE,        CL999_INT_PAID_TO_DATE AS INT_PD_TO_DATE,        CL999_PRIN_BILL_DATE AS LAST_STMT_DATE,        CL999_PAYMENT_AMOUNT AS LOAN_PAYMENT_AMOUNT,        CL999_COB_PASTDUE_DAYS AS PASTDUE_DAYS   FROM DATASTORE_AFS.CL999  WHERE CYCLE_DT_NUM = :dt1_cycle_dt_num  ORDER BY SRC_INSTR_ID ), PARTSOLD_TO_INSTR_RELTN AS (SELECT LI.SRC_INSTR_ID,        SR.SRC_INSTR_ID PARTSOLD_TO_INSTR_ID,        LI.PRINCIPAL_BAL_GL_ACCT_NUM AS GL_ACCT_NUM   FROM EDR.FACT_COM_LOAN_INSTR_DAILY LI,        EDR.FACT_PART_SYND_REFERENCE_DAILY SR  WHERE LI.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND SR.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND LI.SRC_SYS_CD = 'AF'    AND LI.SRC_INSTR_ID = SR.REF_SRC_INSTR_ID    AND SUBSTR(SR.PROCESS_TYPE,2,1) = 3 ), PARTSOLD_TO_INSTR AS ( SELECT SUBSTR(A.CYCLE_DT_NUM,1,6) AS YEAR_MONTH,        A.SRC_SYS_CD,        A.SRC_INSTR_ID,        A.PRINCIPAL_BAL_GL_ACCT_NUM AS PARTSOLD_TO_GL_ACCT_NUM,        A.PRINCIPAL_ENDING_BAL,        B.GL_INTEREST_INCOME,        B.GL_INTEREST_EXP_PARTICPATED,        A.AF_INT_AMT_MTD   FROM EDR.FACT_COM_LOAN_INSTR_DAILY A,        EDR.DIM_AF_GL_MAP B  WHERE A.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND A.SRC_SYS_CD = 'AF'    AND A.SRC_INSTR_ID IN         (SELECT SR.SRC_INSTR_ID            FROM EDR.FACT_COM_LOAN_INSTR_DAILY LI,                 EDR.FACT_PART_SYND_REFERENCE_DAILY SR           WHERE LI.CYCLE_DT_NUM = :dt1_cycle_dt_num             AND SR.CYCLE_DT_NUM = :dt1_cycle_dt_num             AND LI.SRC_INSTR_ID = SR.REF_SRC_INSTR_ID             AND SUBSTR(SR.PROCESS_TYPE,2,1) = 3)    AND A.AF_INT_AMT_MTD <> 0    AND A.DIM_GL_MAP_KEY = B.DIM_AF_GL_MAP_KEY  ),  PARTSOLD_TO_DETAIL AS  (  SELECT B.YEAR_MONTH,         B.SRC_SYS_CD,         A.SRC_INSTR_ID,         A.PARTSOLD_TO_INSTR_ID,         A.GL_ACCT_NUM,         B.PARTSOLD_TO_GL_ACCT_NUM,         B.PRINCIPAL_ENDING_BAL AS PARTSOLD_TO_ENDING_BAL,         B.GL_INTEREST_INCOME AS PARTSOLD_TO_GL_INTEREST_INCOME,         B.GL_INTEREST_EXP_PARTICPATED AS PARTSOLD_TO_GL_INT_EXP_PART,         B.AF_INT_AMT_MTD AS PARTSOLD_TO_INTEREST_INCOME    FROM PARTSOLD_TO_INSTR_RELTN A    LEFT JOIN PARTSOLD_TO_INSTR B      ON A.PARTSOLD_TO_INSTR_ID = B.SRC_INSTR_ID   WHERE B.AF_INT_AMT_MTD <> 0   ),   PARTSOLD_TO_SUM AS   (   SELECT A.YEAR_MONTH,          A.SRC_SYS_CD,          A.SRC_INSTR_ID,          A.GL_ACCT_NUM,          A.PARTSOLD_TO_GL_INTEREST_INCOME,          A.PARTSOLD_TO_GL_INT_EXP_PART,          NVL(SUM(A.PARTSOLD_TO_INTEREST_INCOME),0) AS PARTSOLD_TO_TOTAL_INT_INCOME     FROM PARTSOLD_TO_DETAIL A    GROUP BY A.YEAR_MONTH,             A.SRC_SYS_CD,             A.SRC_INSTR_ID,             A.GL_ACCT_NUM,             A.PARTSOLD_TO_GL_INTEREST_INCOME,             A.PARTSOLD_TO_GL_INT_EXP_PART    HAVING SUM(A.PARTSOLD_TO_INTEREST_INCOME) <> 0    ORDER BY A.SRC_INSTR_ID,             A.GL_ACCT_NUM,             A.PARTSOLD_TO_GL_INT_EXP_PART )    SELECT *  FROM PARTSOLD_TO_DETAIL "
                print(SRC +' Part Sold Query execution is in progress for ' +  dt1) 
                Data_PAR=self.execution(dt1,YM1,SRC,SQL)  
                
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                   
                self.Excel(SRC,YM1)               
                
            if self.checkBox_3.isChecked() == True:
                
                SRC='DF'         
                
                """DF Query """
                SQL="WITH DF_INSTR AS ( SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.OBLIGOR,        A.OBLIGATION AS LOAN_NBR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        B.STATUS_DESC AS LOAN_STATUS,        B.ACCRUAL_STATUS_CD,        C.INSTR_INT_TYPE_CD,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        C.INSTR_TYPE_CD_DESC,        D.GL_PRINCIPAL,        D.GL_INTEREST_INCOME,        D.GL_ACCRUED_INTEREST_RECEIVABLE,        C.INSTR_PROD_CD   FROM EDR.FACT_COM_LOAN_FACIL_MTHLY A   LEFT OUTER JOIN EDR.DIM_STATUS B     ON A.DIM_STATUS_KEY = B.DIM_STATUS_KEY    AND A.SRC_SYS_CD = B.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY C     ON A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_DF_GL_MAP D     ON A.DIM_GL_MAP_KEY = D.DIM_DF_GL_MAP_KEY  WHERE A.SRC_SYS_CD = 'DF'    AND A.YEAR_MONTH = :dt1_year_month    AND B.STATUS_CD = '0'     AND B.CURRENT_FLG = 'Y'    AND TRIM(C.HIERARCHY_TYPE) = 'DEFAULT'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'  ORDER BY A.OBLIGOR, A.OBLIGATION, A.SRC_INSTR_ID  ),  DF_LOAN AS  (  SELECT C.CYCLE_DT_NUM,         C.CUST_NBR AS OBLIGOR,         C.LOAN_NBR,         C.CLTR_CD,         C.CUST_CLTR_TYP_DESC,         UPPER(C.CUST_NAME) AS CUSTOMER_NAME,         C.LAST_PMT_AMT,         C.LAST_PMT_DT    FROM DATASTORE_DATASCAN4.LOAN C   WHERE C.CYCLE_DT_NUM = :dt1_cycle_dt_num   ORDER BY C.CUST_NBR, C.LOAN_NBR  ), DF_PRIN_DAYS_PAST_DUE AS ( SELECT DISTINCT L.CUST_NBR AS OBLIGOR,        P.DT_VALUE-MAX(P.PAST_DUE_DT) AS PRIN_DAYS_PAST_DUE   FROM DATASTORE_DATASCAN4.LOAN L,        DATASTORE_DATASCAN4.PAYMENT_SCHEDULE P  WHERE L.CYCLE_DT_NUM=:dt1_cycle_dt_num    AND L.CYCLE_DT_NUM=P.CYCLE_DT_NUM    AND L.LOAN_NBR=P.LOAN_NBR    AND P.BILLED_IND='YES'    AND P.DT_VALUE>P.PAST_DUE_DT    AND NVL(P.PAID_AMT,0)<(NVL(P.ORIG_AMT,0)-NVL(P.DO_NOT_BILL_AMT,0)- NVL(P.WAIVE_AMT,0))  GROUP BY L.CUST_NBR,P.DT_VALUE,L.LOAN_NBR  ORDER BY L.CUST_NBR ), DF_INTR_DAYS_PAST_DUE AS ( SELECT DISTINCT L.CUST_NBR AS OBLIGOR,        MAX (C.NBR_DAYS_OLD_BILL_UNPAID_INTR) AS INTR_DAYS_PAST_DUE   FROM DATASTORE_DATASCAN4.LOAN L,        DATASTORE_DATASCAN4.CUSTOMER_COLLATERAL_TYPE C  WHERE L.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND L.CYCLE_DT_NUM = C.CYCLE_DT_NUM    AND L.CUST_NBR = C.CUST_NBR    AND C.NBR_DAYS_OLD_BILL_UNPAID_INTR>0  GROUP BY L.CUST_NBR, L.CUST_NAME  ORDER BY L.CUST_NBR ), DF_DAYS_PAST_DUE AS ( SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.SRC_INSTR_ID,        TO_CHAR(TO_NUMBER(SUBSTR(A.SRC_INSTR_ID,1,10))) AS OBLIGOR,        CASE WHEN A.FULLY_UNDRAWN_IND = 'Y'             THEN 0             WHEN NVL(B.PRIN_DAYS_PAST_DUE,0) > NVL(C.INTR_DAYS_PAST_DUE,0)             THEN B.PRIN_DAYS_PAST_DUE             WHEN NVL(C.INTR_DAYS_PAST_DUE,0) > NVL(B.PRIN_DAYS_PAST_DUE,0)             THEN C.INTR_DAYS_PAST_DUE             WHEN NVL(B.PRIN_DAYS_PAST_DUE,0) = NVL(C.INTR_DAYS_PAST_DUE,0)             THEN NVL(B.PRIN_DAYS_PAST_DUE,0)        ELSE 0        END AS PAST_DUE_DAYS   FROM EDR.FACT_COM_LOAN_FACIL_MTHLY_SUPP A   LEFT JOIN DF_PRIN_DAYS_PAST_DUE B     ON TO_CHAR(TO_NUMBER(SUBSTR(A.SRC_INSTR_ID,1,10))) = B.OBLIGOR   LEFT JOIN DF_INTR_DAYS_PAST_DUE C     ON TO_CHAR(TO_NUMBER(SUBSTR(A.SRC_INSTR_ID,1,10))) = C.OBLIGOR  WHERE YEAR_MONTH = :dt1_year_month    AND SRC_SYS_CD = 'DF'   ORDER BY SRC_INSTR_ID ) SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.OBLIGOR,        B.CUSTOMER_NAME,        B.LOAN_NBR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        B.LAST_PMT_DT,        'Y' AS COLLATERAL,        A.LOAN_STATUS,        A.ACCRUAL_STATUS_CD,        A.INSTR_INT_TYPE_CD,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        B.LAST_PMT_AMT,        A.INSTR_TYPE_CD_DESC,        A.GL_PRINCIPAL,        A.GL_INTEREST_INCOME,        A.GL_ACCRUED_INTEREST_RECEIVABLE,        A.INSTR_PROD_CD,        C.PAST_DUE_DAYS   FROM DF_INSTR A   LEFT OUTER JOIN DF_LOAN B     ON TRIM(A.OBLIGOR) = TRIM(B.OBLIGOR)    AND TRIM(A.LOAN_NBR) = TRIM(B.LOAN_NBR)   LEFT OUTER JOIN DF_DAYS_PAST_DUE C     ON TRIM(A.OBLIGOR) = TRIM(C.OBLIGOR)  ORDER BY A.OBLIGOR, A.LOAN_NBR"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1)  
                Data_BAL=self.execution(dt1,YM1,SRC,SQL) 
                
                """ DF Intrest Income """
                SQL="SELECT SUBSTR(A.CYCLE_DT_NUM,1,6) AS YEAR_MONTH,        'DF' AS SRC_SYS_CD,        A.CUST_NBR AS OBLIGOR,        A.LOAN_NBR,        A.CUST_NAME,        A.COST_CTR_NBR,        A.CUSTOMER_STATUS,        A.CUST_TYPE,        A.FIN_CO_NBR AS COMPANY_NUM,        A.FINANCE_COMPANY,        A.CURR_PRINC_AMT,        510180 AS GL_INTEREST_INCOME,        A.BILLED_AMOUNT AS TOTAL_INT_INCOME,        A.PRODUCT_TYPE   FROM DATASTORE_DATASCAN4.BILLING_DETAIL A  WHERE A.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND A.BILLING_DTL_TYP = 'INTR_REC'  ORDER BY A.CUST_NBR, A.LOAN_NBR"
                print(SRC +' Intrest Income Query execution is in progress for ' +  dt1)
                Data_INT=self.execution(dt1,YM1,SRC,SQL)     
                
                Data_PAR=1 #dummy value                
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)   
                self.Excel(SRC,YM1)                     
                
            if self.checkBox_4.isChecked() == True:

                SRC='FD'           
                """ FD Query"""
                SQL="WITH FD_INSTR      AS (SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.SRC_INSTR_ID,                 A.INSTR_OPEN_DT,                 B.STATUS_DESC AS LOAN_STATUS,                 B.ACCRUAL_STATUS_DESC AS LOAN_ACCRUAL_STATUS,                 C.INSTR_INT_TYPE_CD,                 A.PRI_CASH_INT_RT,                 A.PRI_PUR_INT_RT,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_UNPAID_BAL,                 A.LAST_BILLING_CYCLE_DT AS LAST_STMT_DT,                 E.PRINCIPAL_ENDING_BAL AS STMT_PRINCIPAL_ENDING_BAL,                 E.PRINCIPAL_UNPAID_BAL AS STMT_PRINCIPAL_UNPAID_BAL,                 E.PREV_PRINCIPAL_ENDING_BAL AS STMT_PREV_PRINCIPAL_ENDING_BAL,                 A.MTD_INTEREST_INCOME,                 A.ACCRUED_INTEREST_RECEIVABLE,                 A.LAST_PAYMENT_AMT,                 C.INSTR_TYPE_CD_DESC,                 A.PRINCIPAL_BAL_GL_ACCT_NUM AS GL_ACCT_NUM,                 D.GL_INTEREST_INCOME,                 D.GL_ACCR_INTEREST_RECEIVABLE,                 C.INSTR_PROD_CD_DESC,                 A.CUSTOMER_GROUP_DESC,                 A.ACTUAL_PAST_DUE_DAYS            FROM (SELECT *                    FROM EDR.FACT_CREDITCARD_INSTR_MONTHLY A                   WHERE     A.YEAR_MONTH = :dt1_year_month                         AND A.SRC_SYS_CD = 'FD'                         AND A.PRINCIPAL_UNPAID_BAL <> 0) A                 LEFT OUTER JOIN (SELECT *                                    FROM EDR.DIM_STATUS B                                   WHERE B.CURRENT_FLG = 'Y') B                    ON     A.DIM_STATUS_KEY = B.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = B.SRC_SYS_CD                 LEFT OUTER JOIN                 (SELECT *                    FROM EDR.DIM_PRODUCT_HIERARCHY C                   WHERE C.LEVEL_NUM IS NULL AND C.CURRENT_FLG = 'Y') C                    ON     A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_FD_GL_MAP D                    ON     A.DIM_GL_MAP_KEY = D.DIM_FD_GL_MAP_KEY                       AND A.PRINCIPAL_BAL_GL_ACCT_NUM =                              D.GL_PRINCIPAL_BALANCE                       AND D.CURRENT_FLG = 'Y'                 LEFT OUTER JOIN EDR.FACT_CREDITCARD_INSTR_CYCLE E                    ON     A.SRC_INSTR_ID = E.SRC_INSTR_ID                       AND A.LAST_BILLING_CYCLE_DT = E.LAST_BILLING_CYCLE_DT) SELECT *   FROM FD_INSTR"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1)
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)        
                
                """FD INT. INCOME"""
                SQL="SELECT to_number(SUBSTR (A.CYCLE_DT_NUM, 1, 6)) AS YEAR_MONTH,        A.SRC_SYS_CD,        A.CYCLE_DT_NUM,        A.SRC_INSTR_ID,        A.PRINCIPAL_BAL_GL_ACCT_NUM,        C.GL_INTEREST_INCOME,        B.STATUS_DESC,        A.LAST_FCHG_ASSESSED_AMT  as TOTAL_INT_INCOME  FROM EDR.FACT_CREDITCARD_INSTR_DAILY A,        EDR.DIM_STATUS B,        EDR.DIM_FD_GL_MAP C  WHERE     A.CYCLE_DT_NUM BETWEEN (select min(cycle_dt_num2) from edr.period where year_month2=:dt1_year_month and cycle_dt_num2 is not null) AND :dt1_cycle_dt_num        AND A.SRC_SYS_CD = 'FD'        AND A.LAST_FCHG_ASSESSED_AMT <> 0        AND A.DIM_STATUS_KEY = B.DIM_STATUS_KEY        AND A.DIM_GL_MAP_KEY = C.DIM_FD_GL_MAP_KEY"
                print(SRC +' Intrest Income Query execution is in progress for ' +  dt1)
                Data_INT=self.execution(dt1,YM1,SRC,SQL)  
                                 
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM) 
                self.Excel(SRC,YM1)    
                
            if self.checkBox_5.isChecked() == True:
                SRC='IMOD'                
                
                """IMOD Query """
                SQL="SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.SRC_INSTR_ID,        NULL AS EFFECTIVE_START_DT,        NULL AS CURR_MATURITY_DT,        NULL AS ORIGINAL_BAL,        A.YTD_INTEREST_AMT AS INT_PAID_TO_DATE,        NULL AS COLLATERAL,        F.STATUS_CD AS LOAN_STATUS,        F.STATUS_DESC AS LOAN_STATUS_DESC,        F.ACCRUAL_STATUS_CD,        F.ACCRUAL_STATUS_DESC,        NULL AS ORIGINAL_INT_RATE,        A.ANNUAL_RATE AS CURR_INTEREST_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.MONTH_END_CAL_AVG_PRIN_END_BAL AS AVERAGE_BAL,        A.INTEREST_BAL AS TOTAL_INT_INCOME,        A.ACCRUED_INTEREST_RECEIVABLE,        A.LAST_STATEMENT_DT,        A.LAST_PAYMENT_AMT,        B.IMOD_GL_PRINCIPAL AS GL_ACCT_NUM,        B.IMOD_GL_INTEREST_INCOME as GL_INTEREST_INCOME,        B.IMOD_GL_ACCR_INT_RECEIVABLE,        A.PAST_DUE_DAYS,        D.INSTR_TYPE_CD AS SRC_ACCT_TYPE,        D.INSTR_TYPE_CD_DESC AS SRC_ACCT_TYPE_DESC,        D.INSTR_PROD_CD AS PRODUCT_CD,        D.INSTR_PROD_CD_DESC AS PRODUCT_DESCRIPTION,        A.FIXED_PAYMENT_INDICATOR   FROM EDR.FACT_OVERDRAFT_MONTHLY A        LEFT JOIN EDR.DIM_IM_GL_MAP B           ON A.DIM_GL_MAP_KEY = B.DIM_IM_GL_MAP_KEY AND B.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D           ON     A.SRC_SYS_CD = D.SRC_SYS_CD              AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY              AND D.HIERARCHY_TYPE = 'RBG'              AND D.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_STATUS F           ON     A.SRC_SYS_CD = F.SRC_SYS_CD              AND A.DIM_STATUS_KEY = F.DIM_STATUS_KEY              AND F.CURRENT_FLG = 'Y'  WHERE     A.YEAR_MONTH = :dt1_year_month        AND A.SRC_SYS_CD = 'IMOD'        AND (A.PRINCIPAL_ENDING_BAL <> 0 OR A.INTEREST_BAL <> 0)"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)     
                
                Data_INT= 1 #passing dummy variables                 
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                 
                self.Excel(SRC,YM1)         
                
                
            if self.checkBox_6.isChecked() == True:           
                SRC='LS'
                
                """LS Query"""
                SQL="WITH LS_INSTR AS ( SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        C.STATUS_DESC,        C.ACCRUAL_STATUS_DESC,        D.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        'CS_LP_PRIN' AS BALANCE_TYPE,        A.LS_CS_LP_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.LS_CS_LP_INT_INCOME AS INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        D.INSTR_TYPE_CD_DESC,        B.GL_CS_LP_PRINCIPAL AS GL_PRINCIPAL,        B.GL_CS_LP_INTEREST_INCOME AS GL_INTEREST_INCOME,        B.GL_ACCR_INTEREST_RECEIVABLE,        D.INSTR_PROD_CD_DESC,        A.LS_LEASE_TYPE AS LEASE_TYPE   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY   LEFT OUTER JOIN EDR.DIM_STATUS C     ON A.DIM_STATUS_KEY = C.DIM_STATUS_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'LS'    AND A.LS_CS_LP_PRIN_ENDING_BAL <> 0    AND SUBSTR(A.LS_LESSOR_TYPE,2,2) <> '99'    AND B.CURRENT_FLG = 'Y'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'    AND D.HIERARCHY_TYPE = 'DEFAULT'    AND A.COMPANY_NUM = 2   UNION ALL     SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        C.STATUS_DESC,        C.ACCRUAL_STATUS_DESC,        D.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        'RA_PRIN' AS BALANCE_TYPE,        A.LS_RA_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.LS_RA_INT_INCOME AS INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        D.INSTR_TYPE_CD_DESC,        B.GL_RA_PRINCIPAL AS GL_PRINCIPAL,        B.GL_RA_INTEREST_INCOME AS GL_INTEREST_INCOME,        B.GL_ACCR_INTEREST_RECEIVABLE,        D.INSTR_PROD_CD_DESC,        A.LS_LEASE_TYPE AS LEASE_TYPE   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY   LEFT OUTER JOIN EDR.DIM_STATUS C     ON A.DIM_STATUS_KEY = C.DIM_STATUS_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'LS'    AND A.LS_RA_PRIN_ENDING_BAL <> 0    AND SUBSTR(A.LS_LESSOR_TYPE,2,2) <> '99'    AND B.CURRENT_FLG = 'Y'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'    AND D.HIERARCHY_TYPE = 'DEFAULT'    AND A.COMPANY_NUM = 2  UNION ALL  SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        C.STATUS_DESC,        C.ACCRUAL_STATUS_DESC,        D.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        'TL_PRIN' AS BALANCE_TYPE,        A.LS_TL_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.LS_TL_INT_INCOME AS INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        D.INSTR_TYPE_CD_DESC,        B.GL_TL_PRINCIPAL AS GL_PRINCIPAL,        B.GL_TL_INTEREST_INCOME AS GL_INTEREST_INCOME,        B.GL_ACCR_INTEREST_RECEIVABLE,        D.INSTR_PROD_CD_DESC,        A.LS_LEASE_TYPE AS LEASE_TYPE   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY   LEFT OUTER JOIN EDR.DIM_STATUS C     ON A.DIM_STATUS_KEY = C.DIM_STATUS_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'LS'    AND A.LS_TL_PRIN_ENDING_BAL <> 0    AND SUBSTR(A.LS_LESSOR_TYPE,2,2) <> '99'    AND B.CURRENT_FLG = 'Y'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'    AND D.HIERARCHY_TYPE = 'DEFAULT'    AND A.COMPANY_NUM = 2     UNION ALL SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        C.STATUS_DESC,        C.ACCRUAL_STATUS_DESC,        D.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        'MUNI_PRIN' AS BALANCE_TYPE,        A.LS_MUNI_PRINCIPAL_ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.LS_MUNI_INTEREST_INCOME AS INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        D.INSTR_TYPE_CD_DESC,        B.GL_MUNI_PRINCIPAL AS GL_PRINCIPAL,        B.GL_MUNI_INTEREST_INCOME AS GL_INTEREST_INCOME,        B.GL_MUNI_ACCR_INT_RECEIVABLE AS GL_ACCR_INTEREST_RECEIVABLE,        D.INSTR_PROD_CD_DESC,        A.LS_LEASE_TYPE AS LEASE_TYPE   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY   LEFT OUTER JOIN EDR.DIM_STATUS C     ON A.DIM_STATUS_KEY = C.DIM_STATUS_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'LS'    AND A.LS_MUNI_PRINCIPAL_ENDING_BAL <> 0    AND SUBSTR(A.LS_LESSOR_TYPE,2,2) <> '99'    AND B.CURRENT_FLG = 'Y'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'    AND D.HIERARCHY_TYPE = 'DEFAULT'    AND A.COMPANY_NUM = 2     UNION ALL SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        C.STATUS_DESC,        C.ACCRUAL_STATUS_DESC,        D.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        'RL_FIXED_PRIN' AS BALANCE_TYPE,        A.LS_RL_FIXED_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.LS_RL_FIXED_INT_INCOME AS INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        D.INSTR_TYPE_CD_DESC,        B.GL_RL_FIXED_PRINCIPAL AS GL_PRINCIPAL,        B.GL_RL_FIXED_INTEREST_INCOME AS GL_INTEREST_INCOME,        B.GL_ACCR_INTEREST_RECEIVABLE,        D.INSTR_PROD_CD_DESC,        A.LS_LEASE_TYPE AS LEASE_TYPE   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY   LEFT OUTER JOIN EDR.DIM_STATUS C     ON A.DIM_STATUS_KEY = C.DIM_STATUS_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'LS'    AND A.LS_RL_FIXED_PRIN_ENDING_BAL <> 0    AND SUBSTR(A.LS_LESSOR_TYPE,2,2) <> '99'    AND B.CURRENT_FLG = 'Y'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'    AND D.HIERARCHY_TYPE = 'DEFAULT'    AND A.COMPANY_NUM = 2     UNION ALL   SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        C.STATUS_DESC,        C.ACCRUAL_STATUS_DESC,        D.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        'RL_VAR_PRIN' AS BALANCE_TYPE,        A.LS_RL_VARIABLE_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.LS_RL_VARIABLE_INT_INCOME AS INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        D.INSTR_TYPE_CD_DESC,        B.GL_RL_VARIABLE_PRINCIPAL AS GL_PRINCIPAL,        B.GL_RL_VARIABLE_INTEREST_INCOME AS GL_INTEREST_INCOME,        B.GL_ACCR_INTEREST_RECEIVABLE,        D.INSTR_PROD_CD_DESC,        A.LS_LEASE_TYPE AS LEASE_TYPE   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY   LEFT OUTER JOIN EDR.DIM_STATUS C     ON A.DIM_STATUS_KEY = C.DIM_STATUS_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'LS'    AND A.LS_RL_VARIABLE_PRIN_ENDING_BAL <> 0    AND SUBSTR(A.LS_LESSOR_TYPE,2,2) <> '99'    AND B.CURRENT_FLG = 'Y'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'    AND D.HIERARCHY_TYPE = 'DEFAULT'    AND A.COMPANY_NUM = 2     UNION ALL   SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        C.STATUS_DESC,        C.ACCRUAL_STATUS_DESC,        D.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        'NON_ACCRUAL_PRIN' AS BALANCE_TYPE,        A.LS_NON_ACCRUAL_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        D.INSTR_TYPE_CD_DESC,        B.GL_NON_ACCRUAL_PRINCIPAL,        null AS GL_INTEREST_INCOME,        B.GL_ACCR_INTEREST_RECEIVABLE,        D.INSTR_PROD_CD_DESC,        A.LS_LEASE_TYPE AS LEASE_TYPE   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY   LEFT OUTER JOIN EDR.DIM_STATUS C     ON A.DIM_STATUS_KEY = C.DIM_STATUS_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'LS'    AND A.LS_NON_ACCRUAL_PRIN_ENDING_BAL <> 0    AND SUBSTR(A.LS_LESSOR_TYPE,2,2) <> '99'    AND B.CURRENT_FLG = 'Y'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'    AND D.HIERARCHY_TYPE = 'DEFAULT'    AND A.COMPANY_NUM = 2     UNION ALL  SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.OBLIGOR,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        'Y' AS COLLATERAL,        C.STATUS_DESC,        C.ACCRUAL_STATUS_DESC,        D.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        'LL_PRIN' AS BALANCE_TYPE,        A.LS_LL_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.LS_LL_INT_INCOME AS INTEREST_INCOME ,        A.ACCR_INTEREST_RECEIVABLE,        D.INSTR_TYPE_CD_DESC,        B.GL_LL_PRINCIPAL AS GL_PRINCIPAL,        B.GL_LL_INTEREST_INCOME AS GL_INTEREST_INCOME,        B.GL_ACCR_INTEREST_RECEIVABLE,        D.INSTR_PROD_CD_DESC,        A.LS_LEASE_TYPE AS LEASE_TYPE   FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A   LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY   LEFT OUTER JOIN EDR.DIM_STATUS C     ON A.DIM_STATUS_KEY = C.DIM_STATUS_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'LS'    AND A.LS_LL_PRIN_ENDING_BAL <> 0    AND SUBSTR(A.LS_LESSOR_TYPE,2,2) <> '99'    AND B.CURRENT_FLG = 'Y'    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'    AND D.HIERARCHY_TYPE = 'DEFAULT'    AND A.COMPANY_NUM = 2   ),  LS_MASTER AS  (  SELECT A.ACCOUNT_NUMBER,        LPAD(SUBSTR(A.ACCOUNT_NUMBER,4,7),10,0) ||'-'||        LPAD(SUBSTR(A.ACCOUNT_NUMBER,11,3),5,0) ||'-'||        LPAD(SUBSTR(A.ACCOUNT_NUMBER,0,3),3,0) AS SRC_INSTR_ID,        A.END_BAL,        A.LOAN_PAYMENT,        A.LOAN_RATE,        A.LOAN_TYPE,        A.ORIGINAL_BALANCE,        A.LEASE_TYPE,        A.CUST_NAME,        A.DELIN_STATUS   FROM DATASTORE_INFOLEASE.STG_LS_MASTER A  WHERE A.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND A.END_BAL <> 0  ORDER BY SRC_INSTR_ID, A.ACCOUNT_NUMBER, A.CUST_NAME  )  SELECT A.YEAR_MONTH,         A.SRC_SYS_CD,         A.COMPANY_NUM,         A.OBLIGOR,         B.CUST_NAME,         A.SRC_INSTR_ID,         A.EFFECTIVE_START_DT,         A.CURR_MATURITY_DT,         A.ORIGINAL_BAL,         A.COLLATERAL,         A.STATUS_DESC,         A.ACCRUAL_STATUS_DESC,         A.INSTR_INT_TYPE_CD_DESC,         A.ACCRUAL_METHOD_CD,         A.ORIGINAL_INT_RATE,         A.CURR_INT_RATE,         A.PREV_PRINCIPAL_ENDING_BAL,         A.BALANCE_TYPE,         A.PRINCIPAL_ENDING_BAL,         A.INTEREST_INCOME,         A.ACCR_INTEREST_RECEIVABLE,         A.INSTR_TYPE_CD_DESC,         A.GL_PRINCIPAL,         A.GL_INTEREST_INCOME,         A.GL_ACCR_INTEREST_RECEIVABLE,         B.LOAN_PAYMENT,         A.INSTR_PROD_CD_DESC,         A.LEASE_TYPE,         B.DELIN_STATUS    FROM LS_INSTR A    LEFT OUTER JOIN LS_MASTER B      ON A.SRC_INSTR_ID = B.SRC_INSTR_ID     AND A.LEASE_TYPE = B.LEASE_TYPE   ORDER BY A.SRC_INSTR_ID"                
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)        
                 
                """ LS INT. INCOME""" 
                SQL="WITH LS_INSTR      AS (SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.COST_CTR_NUM,                 A.OBLIGOR,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 'Y' AS COLLATERAL,                 C.STATUS_DESC,                 C.ACCRUAL_STATUS_DESC,                 D.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD_CD,                 A.ORIGINAL_INT_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 'CS_LP_PRIN' AS BALANCE_TYPE,                 A.LS_CS_LP_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,                 ABS (A.LS_CS_LP_INT_INCOME) AS INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.INSTR_TYPE_CD_DESC,                 B.GL_CS_LP_PRINCIPAL AS GL_PRINCIPAL,                 B.GL_CS_LP_INTEREST_INCOME AS GL_INTEREST_INCOME,                 B.GL_ACCR_INTEREST_RECEIVABLE,                 D.INSTR_PROD_CD_DESC,                 A.LS_LEASE_TYPE AS LEASE_TYPE            FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A                 LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B                    ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY                 LEFT OUTER JOIN EDR.DIM_STATUS C                    ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D                    ON     A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'LS'                 AND A.LS_CS_LP_INT_INCOME <> 0                 AND A.LS_LESSOR_TYPE NOT IN ('199', '299', '399')                 AND B.CURRENT_FLG = 'Y'                 AND C.CURRENT_FLG = 'Y'                 AND D.CURRENT_FLG = 'Y'                 AND D.HIERARCHY_TYPE = 'DEFAULT'                 AND A.COMPANY_NUM = 2          UNION ALL          SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.COST_CTR_NUM,                 A.OBLIGOR,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 'Y' AS COLLATERAL,                 C.STATUS_DESC,                 C.ACCRUAL_STATUS_DESC,                 D.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD_CD,                 A.ORIGINAL_INT_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 'RA_PRIN' AS BALANCE_TYPE,                 A.LS_RA_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,                 A.LS_RA_INT_INCOME AS INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.INSTR_TYPE_CD_DESC,                 B.GL_RA_PRINCIPAL AS GL_PRINCIPAL,                 B.GL_RA_INTEREST_INCOME AS GL_INTEREST_INCOME,                 B.GL_ACCR_INTEREST_RECEIVABLE,                 D.INSTR_PROD_CD_DESC,                 A.LS_LEASE_TYPE AS LEASE_TYPE            FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A                 LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B                    ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY                 LEFT OUTER JOIN EDR.DIM_STATUS C                    ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D                    ON     A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'LS'                 AND A.LS_RA_INT_INCOME <> 0                 AND A.LS_LESSOR_TYPE NOT IN ('199', '299', '399')                 AND B.CURRENT_FLG = 'Y'                 AND C.CURRENT_FLG = 'Y'                 AND D.CURRENT_FLG = 'Y'                 AND D.HIERARCHY_TYPE = 'DEFAULT'                 AND A.COMPANY_NUM = 2          UNION ALL          SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.COST_CTR_NUM,                 A.OBLIGOR,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 'Y' AS COLLATERAL,                 C.STATUS_DESC,                 C.ACCRUAL_STATUS_DESC,                 D.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD_CD,                 A.ORIGINAL_INT_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 'TL_PRIN' AS BALANCE_TYPE,                 A.LS_TL_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,                 A.LS_TL_INT_INCOME AS INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.INSTR_TYPE_CD_DESC,                 B.GL_TL_PRINCIPAL AS GL_PRINCIPAL,                 B.GL_TL_INTEREST_INCOME AS GL_INTEREST_INCOME,                 B.GL_ACCR_INTEREST_RECEIVABLE,                 D.INSTR_PROD_CD_DESC,                 A.LS_LEASE_TYPE AS LEASE_TYPE            FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A                 LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B                    ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY                 LEFT OUTER JOIN EDR.DIM_STATUS C                    ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D                    ON     A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'LS'                 AND A.LS_TL_INT_INCOME <> 0                 AND A.LS_LESSOR_TYPE NOT IN ('199', '299', '399')                 AND B.CURRENT_FLG = 'Y'                 AND C.CURRENT_FLG = 'Y'                 AND D.CURRENT_FLG = 'Y'                 AND D.HIERARCHY_TYPE = 'DEFAULT'                 AND A.COMPANY_NUM = 2          UNION ALL                                             /* MUNI_PRIN */          SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.COST_CTR_NUM,                 A.OBLIGOR,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 'Y' AS COLLATERAL,                 C.STATUS_DESC,                 C.ACCRUAL_STATUS_DESC,                 D.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD_CD,                 A.ORIGINAL_INT_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 'MUNI_PRIN' AS BALANCE_TYPE,                 A.LS_MUNI_PRINCIPAL_ENDING_BAL AS PRINCIPAL_ENDING_BAL,                 A.LS_MUNI_INTEREST_INCOME AS INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.INSTR_TYPE_CD_DESC,                 B.GL_MUNI_PRINCIPAL AS GL_PRINCIPAL,                 B.GL_MUNI_INTEREST_INCOME AS GL_INTEREST_INCOME,                 B.GL_MUNI_ACCR_INT_RECEIVABLE AS GL_ACCR_INTEREST_RECEIVABLE,                 D.INSTR_PROD_CD_DESC,                 A.LS_LEASE_TYPE AS LEASE_TYPE            FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A                 LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B                    ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY                 LEFT OUTER JOIN EDR.DIM_STATUS C                    ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D                    ON     A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'LS'                 AND A.LS_MUNI_INTEREST_INCOME <> 0                 AND A.LS_LESSOR_TYPE NOT IN ('199', '299', '399')                 AND B.CURRENT_FLG = 'Y'                 AND C.CURRENT_FLG = 'Y'                 AND D.CURRENT_FLG = 'Y'                 AND D.HIERARCHY_TYPE = 'DEFAULT'                 AND A.COMPANY_NUM = 2          UNION ALL                                          /*RL_FIXED_PRIN */          SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.COST_CTR_NUM,                 A.OBLIGOR,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 'Y' AS COLLATERAL,                 C.STATUS_DESC,                 C.ACCRUAL_STATUS_DESC,                 D.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD_CD,                 A.ORIGINAL_INT_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 'RL_FIXED_PRIN' AS BALANCE_TYPE,                 A.LS_RL_FIXED_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,                 A.LS_RL_FIXED_INT_INCOME AS INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.INSTR_TYPE_CD_DESC,                 B.GL_RL_FIXED_PRINCIPAL AS GL_PRINCIPAL,                 B.GL_RL_FIXED_INTEREST_INCOME AS GL_INTEREST_INCOME,                 B.GL_ACCR_INTEREST_RECEIVABLE,                 D.INSTR_PROD_CD_DESC,                 A.LS_LEASE_TYPE AS LEASE_TYPE            FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A                 LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B                    ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY                 LEFT OUTER JOIN EDR.DIM_STATUS C                    ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D                    ON     A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'LS'                 AND A.LS_RL_FIXED_INT_INCOME <> 0                 AND A.LS_LESSOR_TYPE NOT IN ('199', '299', '399')                 AND B.CURRENT_FLG = 'Y'                 AND C.CURRENT_FLG = 'Y'                 AND D.CURRENT_FLG = 'Y'                 AND D.HIERARCHY_TYPE = 'DEFAULT'                 AND A.COMPANY_NUM = 2          UNION ALL                                           /* RL_VAR_PRIN */          SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.COST_CTR_NUM,                 A.OBLIGOR,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 'Y' AS COLLATERAL,                 C.STATUS_DESC,                 C.ACCRUAL_STATUS_DESC,                 D.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD_CD,                 A.ORIGINAL_INT_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 'RL_VAR_PRIN' AS BALANCE_TYPE,                 A.LS_RL_VARIABLE_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,                 A.LS_RL_VARIABLE_INT_INCOME AS INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.INSTR_TYPE_CD_DESC,                 B.GL_RL_VARIABLE_PRINCIPAL AS GL_PRINCIPAL,                 B.GL_RL_VARIABLE_INTEREST_INCOME AS GL_INTEREST_INCOME,                 B.GL_ACCR_INTEREST_RECEIVABLE,                 D.INSTR_PROD_CD_DESC,                 A.LS_LEASE_TYPE AS LEASE_TYPE            FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A                 LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B                    ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY                 LEFT OUTER JOIN EDR.DIM_STATUS C                    ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D                    ON     A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'LS'                 AND A.LS_RL_VARIABLE_INT_INCOME <> 0                 AND A.LS_LESSOR_TYPE NOT IN ('199', '299', '399')                 AND B.CURRENT_FLG = 'Y'                 AND C.CURRENT_FLG = 'Y'                 AND D.CURRENT_FLG = 'Y'                 AND D.HIERARCHY_TYPE = 'DEFAULT'                 AND A.COMPANY_NUM = 2          UNION ALL          SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.COST_CTR_NUM,                 A.OBLIGOR,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 'Y' AS COLLATERAL,                 C.STATUS_DESC,                 C.ACCRUAL_STATUS_DESC,                 D.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD_CD,                 A.ORIGINAL_INT_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 'LL_PRIN' AS BALANCE_TYPE,                 A.LS_LL_PRIN_ENDING_BAL AS PRINCIPAL_ENDING_BAL,                 A.LS_LL_INT_INCOME AS INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.INSTR_TYPE_CD_DESC,                 B.GL_LL_PRINCIPAL AS GL_PRINCIPAL,                 B.GL_LL_INTEREST_INCOME AS GL_INTEREST_INCOME,                 B.GL_ACCR_INTEREST_RECEIVABLE,                 D.INSTR_PROD_CD_DESC,                 A.LS_LEASE_TYPE AS LEASE_TYPE            FROM EDR.FACT_COM_LOAN_INSTR_MTHLY A                 LEFT OUTER JOIN EDR.DIM_LS_GL_MAP B                    ON A.DIM_GL_MAP_KEY = B.DIM_LS_GL_MAP_KEY                 LEFT OUTER JOIN EDR.DIM_STATUS C                    ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D                    ON     A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.LS_LEASE_TYPE = D.INSTR_TYPE_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'LS'                 AND A.LS_LL_INT_INCOME <> 0                 AND A.LS_LESSOR_TYPE NOT IN ('199', '299', '399')                 AND B.CURRENT_FLG = 'Y'                 AND C.CURRENT_FLG = 'Y'                 AND D.CURRENT_FLG = 'Y'                 AND D.HIERARCHY_TYPE = 'DEFAULT'                 AND A.COMPANY_NUM = 2),      LS_MASTER      AS (SELECT A.ACCOUNT_NUMBER,                    LPAD (SUBSTR (A.ACCOUNT_NUMBER, 4, 7), 10, 0)                 || '-'                 || LPAD (SUBSTR (A.ACCOUNT_NUMBER, 11, 3), 5, 0)                 || '-'                 || LPAD (SUBSTR (A.ACCOUNT_NUMBER, 0, 3), 3, 0)                    AS SRC_INSTR_ID,                 A.END_BAL,                 A.LOAN_PAYMENT,                 A.LOAN_RATE,                 A.LOAN_TYPE,                 A.ORIGINAL_BALANCE,                 A.LEASE_TYPE,                 A.CUST_NAME,                 A.DELIN_STATUS            FROM DATASTORE_INFOLEASE.STG_LS_MASTER A           WHERE A.CYCLE_DT_NUM = :dt1_cycle_dt_num          ORDER BY SRC_INSTR_ID, A.ACCOUNT_NUMBER, A.CUST_NAME) SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.COST_CTR_NUM,        A.OBLIGOR,        B.CUST_NAME,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        A.COLLATERAL,        A.STATUS_DESC,        A.ACCRUAL_STATUS_DESC,        A.INSTR_INT_TYPE_CD_DESC,        A.ACCRUAL_METHOD_CD,        A.ORIGINAL_INT_RATE,        A.CURR_INT_RATE,        A.PREV_PRINCIPAL_ENDING_BAL,        A.BALANCE_TYPE,        A.PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME as TOTAL_INT_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        A.INSTR_TYPE_CD_DESC,        A.GL_PRINCIPAL,        A.GL_INTEREST_INCOME,        A.GL_ACCR_INTEREST_RECEIVABLE,        B.LOAN_PAYMENT,        A.INSTR_PROD_CD_DESC,        A.LEASE_TYPE,        B.DELIN_STATUS   FROM LS_INSTR A        LEFT OUTER JOIN LS_MASTER B           ON A.SRC_INSTR_ID = B.SRC_INSTR_ID AND A.LEASE_TYPE = B.LEASE_TYPE ORDER BY A.SRC_INSTR_ID"
                print(SRC +' Intrest Income Query execution is in progress for ' +  dt1)
                Data_INT=self.execution(dt1,YM1,SRC,SQL)   
                
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)    
                self.Excel(SRC,YM1)   
                
            if self.checkBox_7.isChecked() == True:           
                SRC='LV'       
                SQL="WITH LV_INSTR      AS (SELECT SUBSTR (TO_CHAR (A.CYCLE_DT_NUM), 1, 6) AS YEAR_MONTH,                 'LV' AS SRC_SYS_CD,                    MF12_LSE_LEVEL1                 || LPAD (A.MF12_LSE_LEVEL2, 4, '0')                 || LPAD (A.MF12_LSE_NBR, 5, '0')                    AS SRC_INSTR_ID,                    '20'                 || LPAD (MF12_LSE_DATE_START_YY, 2, 0)                 || TO_CHAR (LPAD (MF12_LSE_DATE_START_MM, 2, 0))                 || TO_CHAR (LPAD (MF12_LSE_DATE_START_DD, 2, 0))                    AS EFFECTIVE_START_DT,                    '20'                 || LPAD (MF12_LSE_DATE_MATURE_YY, 2, 0)                 || LPAD (MF12_LSE_DATE_MATURE_MM, 2, 0)                 || LPAD (MF12_LSE_DATE_MATURE_DD, 2, 0)                    AS CURR_MATURITY_DT,                 MF12_VEH_COST AS ORIGINAL_BAL,                 'Y' AS COLLATERAL,                 CASE                    WHEN TRIM (MF12_LSE_FLAG_INC_ACCRUAL) IS NULL THEN 'N'                    ELSE MF12_LSE_FLAG_INC_ACCRUAL                 END                    ACCRUAL_STATUS,                 MF12_LSE_CODE_PMT_TYPE AS PAYMENT_METHOD_CD,                 MF12_LSE_RATE AS CURR_INT_RATE,                 MF12_LSE_REC_BAL_TOT AS BANK_BAL,                 'INDIRECT AUTO LEASES' AS INSTR_TYPE_CD,                 '03' AS INSTR_PROD_CD            FROM DATASTORE_LV.LVSP912_CONTRACT A           WHERE A.CYCLE_DT_NUM = :dt1_cycle_dt_num          ORDER BY    MF12_LSE_LEVEL1                   || LPAD (A.MF12_LSE_LEVEL2, 4, '0')                   || LPAD (A.MF12_LSE_NBR, 5, '0')),      LV_COLLATERAL      AS (SELECT DISTINCT                 A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.SRC_INSTR_ID,                 CASE WHEN B.COLLATERAL_TYPE = 'AUT' THEN 'Y' ELSE 'N' END                    AS COLLATERAL            FROM EDR.FACT_LOAN_INSTR_MONTHLY A                 LEFT OUTER JOIN EDR.FACT_LOAN_COLLATERAL B                    ON     A.SRC_SYS_CD = B.SRC_SYS_CD                       AND A.YEAR_MONTH = B.YEAR_MONTH                       AND A.DIM_INSTRUMENT_KEY = B.DIM_INSTRUMENT_KEY           WHERE A.SRC_SYS_CD = 'LV' AND A.YEAR_MONTH = :dt1_year_month          ORDER BY A.SRC_INSTR_ID),      LV_CUST      AS (SELECT A.CYCLE_DT_NUM,                    A.MF11_NA_LEVEL1                 || LPAD (A.MF11_NA_LEVEL2, 4, '0')                 || LPAD (A.MF11_NA_NBR, 5, '0')                    AS SRC_INSTR_ID,                 LPAD (A.MF11_NA_NBR, 5, '0') AS CUST_NUM,                 A.MF11_NA_NAME AS CUST_NAME            FROM DATASTORE_LV.LVSP911_CUST_NAME_ADD A           WHERE CYCLE_DT_NUM = :dt1_cycle_dt_num          ORDER BY    A.MF11_NA_LEVEL1                   || LPAD (A.MF11_NA_LEVEL2, 4, '0')                   || LPAD (A.MF11_NA_NBR, 5, '0')),      LV_INTEREST_INCOME_RECV      AS (SELECT A.CYCLE_DT_NUM,                    A.MF12_LSE_LEVEL1                 || LPAD (A.MF12_LSE_LEVEL2, 4, '0')                 || LPAD (A.MF12_LSE_NBR, 5, '0')                    AS SRC_INSTR_ID,                 LPAD (A.MF12_LSE_NBR, 5, '0') AS CUST_NUM,                 127175 AS GL_UNEARNED_INCOME,                 A.MF12_LSE_REC_UNEARNED AS RECEIVABLES_UNEARNED,                 531186 AS GL_AUTO_LEASE_INCOME,                 A.MF12_LSE_INC_EARNED_MTD AS INCOME_EARNED,                   A.MF12_LSE_REC_DUE_30                 + A.MF12_LSE_REC_DUE_60                 + A.MF12_LSE_REC_DUE_90                 + A.MF12_LSE_REC_DUE_120                 + A.MF12_LSE_REC_DUE_150                    AS TOTAL_PAST_DUE_DAYS            FROM DATASTORE_LV.LVSP912_CONTRACT A           WHERE A.CYCLE_DT_NUM = (SELECT CYCLE_DT_NUM2                                         FROM (SELECT YEAR_MONTH2,                                                      CYCLE_DT_NUM2,                                                      WEEKDAY_FLG,                                                      RANK () OVER (ORDER BY CYCLE_DT_NUM2 DESC) RNK                                                 FROM edr.period                                                WHERE year_month2 = :dt1_year_month AND WEEKDAY_FLG = 'Y')                                        WHERE RNK = 2)          ORDER BY    A.MF12_LSE_LEVEL1                   || LPAD (A.MF12_LSE_LEVEL2, 4, '0')                   || LPAD (A.MF12_LSE_NBR, 5, '0')) SELECT to_number(A.YEAR_MONTH) as YEAR_MONTH,        A.SRC_SYS_CD,        D.CUST_NUM,        C.CUST_NAME,        A.SRC_INSTR_ID,        TO_DATE (A.EFFECTIVE_START_DT, 'YYYYMMDD') AS EFFECTIVE_START_DT,        TO_DATE (A.CURR_MATURITY_DT, 'YYYYMMDD') AS CURR_MATURITY_DT,        A.ORIGINAL_BAL,        A.COLLATERAL,        CASE           WHEN A.ACCRUAL_STATUS = 'N' THEN 'In accrual status'           WHEN A.ACCRUAL_STATUS = 'Y' THEN 'Reinstated'        END           AS ACCRUAL_STATUS_DESC,        'LEASE' AS INTEREST_INDICATOR,        A.PAYMENT_METHOD_CD,        A.CURR_INT_RATE,        A.BANK_BAL AS TOTAL_ENDING_BAL,        A.INSTR_TYPE_CD AS LOAN_TYPE,        127167 AS GL_ACCT_NUM,        A.INSTR_PROD_CD,        D.GL_UNEARNED_INCOME,        D.RECEIVABLES_UNEARNED,        D.GL_AUTO_LEASE_INCOME,        D.INCOME_EARNED,        D.TOTAL_PAST_DUE_DAYS   FROM LV_INSTR A        LEFT OUTER JOIN LV_CUST C ON A.SRC_INSTR_ID = C.SRC_INSTR_ID        LEFT OUTER JOIN LV_INTEREST_INCOME_RECV D           ON A.SRC_INSTR_ID = D.SRC_INSTR_ID ORDER BY A.SRC_INSTR_ID "
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                 
                Data_INT=1 #passing dummy variables      
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                 
                self.Excel(SRC,YM1)          
                
            if self.checkBox_8.isChecked() == True:               
                SRC='ML'      
                """ML Query"""
                SQL="WITH ML_INSTR AS ( SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        C.CUST_NUM,        C.GU_FULL_NAME AS CUSTOMER_NAME,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        A.INTEREST_PAID_TO_DT,        'Y' AS COLLATERAL,        D.STATUS_DESC AS LOAN_STATUS,        D.ACCRUAL_STATUS_DESC LOAN_ACCRUAL_STATUS,        E.INSTR_INT_TYPE_CD_DESC AS INTEREST_INDICATOR,        A.ACCRUAL_METHOD,        A.ORIGINAL_RT,        A.CURR_INT_RT,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        CASE WHEN A.INVESTOR_CD BETWEEN 80000 AND 83000             THEN A.PRINCIPAL_UNPAID_BAL - A.CHARGE_OFF_PRINCIPAL_BAL             ELSE A.PRINCIPAL_UNPAID_BAL        END AS PRINCIPAL_UNPAID_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        A.MTD_BRF_PRIN_PAYMENT_AMT AS PRINCIPAL_PAYMENT,        (A.MTD_BRF_PRIN_PAYMENT_AMT + A.INTEREST_INCOME) AS LOAN_PAYMENT,        E.INSTR_TYPE_CD_DESC AS LOAN_TYPE,        A.PRINCIPAL_BAL_GL_ACCT_NUM,        F.GL_INTEREST_INCOME,        F.GL_ACCR_INTEREST_RECEIVABLE,        E.INSTR_PROD_CD_DESC AS PRODUCT_CD_DESC,        A.TOTAL_PAST_DUE_DAYS,        A.INVESTOR_CD,        A.ESCROW_BAL,        A.CHARGE_OFF_PRINCIPAL_BAL   FROM EDR.FACT_MORTGAGE_INSTR_MONTHLY A   LEFT OUTER JOIN EDR.CUST_INSTR_RELATION B     ON TRIM(A.SRC_INSTR_ID) = TRIM(B.ACCT_NUM_DERIVED)    AND B.SRC_SYS_CD = 'RM'    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= B.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(B.END_EFF_DTTM, '31-DEC-9999')    AND A.SRC_SYS_CD = B.APPLCN_CD_DERIVED    AND B.DELETE_INDICATOR = 'N'    AND A.RELATIONSHIP_MANAGER_NUM = B.CUST_NUM   LEFT OUTER JOIN EDR.DIM_CUSTOMER C    ON B.CUST_NUM = C.CUST_NUM    AND C.SRC_SYS_CD = 'RM'    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= C.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(C.END_EFF_DTTM, '31-DEC-9999')    AND C.DELETE_INDICATOR = 'N'    AND A.RELATIONSHIP_MANAGER_NUM = C.CUST_NUM   LEFT OUTER JOIN EDR.DIM_STATUS D     ON A.DIM_STATUS_KEY = D.DIM_STATUS_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY E     ON A.DIM_PRODUCT_KEY = E.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = E.SRC_SYS_CD    AND E.CURRENT_FLG = 'Y'    AND E.LEVEL_NUM IS NULL   LEFT OUTER JOIN EDR.DIM_ML_GL_MAP F     ON A.DIM_GL_MAP_KEY = F.DIM_ML_GL_MAP_KEY    AND A.INVESTOR_CD = TO_NUMBER(F.GL_INVESTOR_CD)    AND F.CURRENT_FLG = 'Y'  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'ML'    AND (A.PRINCIPAL_UNPAID_BAL <> 0 OR A.PRINCIPAL_ENDING_BAL <> 0)    AND A.PRINCIPAL_BAL_GL_ACCT_NUM <> -1  ORDER BY C.CUST_NUM,           A.SRC_INSTR_ID ) SELECT DISTINCT        A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.CUST_NUM,        A.CUSTOMER_NAME,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        A.INTEREST_PAID_TO_DT,        A.COLLATERAL,        A.LOAN_STATUS,        A.LOAN_ACCRUAL_STATUS,        A.INTEREST_INDICATOR,        A.ACCRUAL_METHOD,        A.ORIGINAL_RT,        A.CURR_INT_RT,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL AS PRINCIPAL_END_BAL,        A.PRINCIPAL_UNPAID_BAL AS TOTAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        A.PRINCIPAL_PAYMENT,        A.LOAN_PAYMENT,        A.LOAN_TYPE,        A.PRINCIPAL_BAL_GL_ACCT_NUM AS GL_ACCT_NUM,        A.GL_INTEREST_INCOME,        A.GL_ACCR_INTEREST_RECEIVABLE,        A.PRODUCT_CD_DESC,        A.TOTAL_PAST_DUE_DAYS,        A.INVESTOR_CD,        A.ESCROW_BAL,        A.CHARGE_OFF_PRINCIPAL_BAL   FROM ML_INSTR A "
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)               
                
                """ML Int. Income"""
                SQL="SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        C.CUST_NUM,        C.GU_FULL_NAME AS CUSTOMER_NAME,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        A.INTEREST_PAID_TO_DT,        'Y' AS COLLATERAL,        D.STATUS_DESC AS LOAN_STATUS,        D.ACCRUAL_STATUS_DESC LOAN_ACCRUAL_STATUS,        E.INSTR_INT_TYPE_CD_DESC AS INTEREST_INDICATOR,        A.ACCRUAL_METHOD,        A.ORIGINAL_RT,        A.CURR_INT_RT,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        CASE WHEN A.INVESTOR_CD BETWEEN 80000 AND 83000             THEN A.PRINCIPAL_UNPAID_BAL - A.CHARGE_OFF_PRINCIPAL_BAL             ELSE A.PRINCIPAL_UNPAID_BAL        END AS PRINCIPAL_UNPAID_BAL,        A.INTEREST_INCOME AS TOTAL_INT_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        A.MTD_BRF_PRIN_PAYMENT_AMT AS PRINCIPAL_PAYMENT,        (A.MTD_BRF_PRIN_PAYMENT_AMT + A.INTEREST_INCOME) AS LOAN_PAYMENT,        E.INSTR_TYPE_CD_DESC AS LOAN_TYPE,        A.PRINCIPAL_BAL_GL_ACCT_NUM,        CASE WHEN A.INVESTOR_CD = 52000             THEN 531101             ELSE TO_NUMBER(F.GL_INTEREST_INCOME)                   END AS GL_INTEREST_INCOME,        F.GL_ACCR_INTEREST_RECEIVABLE,        E.INSTR_PROD_CD_DESC AS PRODUCT_CD_DESC,        A.TOTAL_PAST_DUE_DAYS,        A.INVESTOR_CD,        A.ESCROW_BAL,        A.CHARGE_OFF_PRINCIPAL_BAL   FROM EDR.FACT_MORTGAGE_INSTR_MONTHLY A   LEFT OUTER JOIN EDR.CUST_INSTR_RELATION B     ON TRIM(A.SRC_INSTR_ID) = TRIM(B.ACCT_NUM_DERIVED)    AND B.SRC_SYS_CD = 'RM'    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= B.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(B.END_EFF_DTTM, '31-DEC-9999')    AND A.SRC_SYS_CD = B.APPLCN_CD_DERIVED    AND B.DELETE_INDICATOR = 'N'    AND A.RELATIONSHIP_MANAGER_NUM = B.CUST_NUM   LEFT OUTER JOIN EDR.DIM_CUSTOMER C    ON B.CUST_NUM = C.CUST_NUM    AND C.SRC_SYS_CD = 'RM'    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= C.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(C.END_EFF_DTTM, '31-DEC-9999')    AND C.DELETE_INDICATOR = 'N'    AND A.RELATIONSHIP_MANAGER_NUM = C.CUST_NUM   LEFT OUTER JOIN EDR.DIM_STATUS D     ON A.DIM_STATUS_KEY = D.DIM_STATUS_KEY    AND A.SRC_SYS_CD = D.SRC_SYS_CD    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY E     ON A.DIM_PRODUCT_KEY = E.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = E.SRC_SYS_CD    AND E.CURRENT_FLG = 'Y'    AND E.LEVEL_NUM IS NULL   LEFT OUTER JOIN EDR.DIM_ML_GL_MAP F     ON A.DIM_GL_MAP_KEY = F.DIM_ML_GL_MAP_KEY    AND A.INVESTOR_CD = TO_NUMBER(F.GL_INVESTOR_CD)    AND F.GL_INTEREST_INCOME IS NOT NULL    AND F.CURRENT_FLG = 'Y'  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'ML'    AND A.PRINCIPAL_BAL_GL_ACCT_NUM <> -1    AND A.INTEREST_INCOME != 0  ORDER BY C.CUST_NUM,           A.SRC_INSTR_ID "
                print(SRC +' Intrest Income Query execution is in progress for ' +  dt1)
                Data_INT=self.execution(dt1,YM1,SRC,SQL)    
                
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)  
                self.Excel(SRC,YM1)        
                
            if self.checkBox_9.isChecked() == True:            
                SRC='MO'      
                
                """MO QUERY"""
                SQL="WITH MO_INSTR      AS (SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 A.INTEREST_PAID_TO_DT,                 'Y' AS COLLATERAL,                 B.STATUS_DESC AS LOAN_STATUS,                 B.ACCRUAL_STATUS_DESC AS LOAN_ACCRUAL_STATUS,                 C.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD,                 A.ORIGINAL_RT,                 A.CURR_INT_RT,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_UNPAID_BAL,                 A.INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.LOAN_PAYMENT_AMOUNT,                 C.INSTR_TYPE_CD_DESC,                 A.PRINCIPAL_BAL_GL_ACCT_NUM,                 E.GL_INTEREST_INCOME,                 E.GL_ACCR_INTEREST_RECEIVABLE,                 C.INSTR_PROD_CD_DESC,                 A.RELATIONSHIP_MANAGER_NUM,                 F.PAST_DUE_CODE,                 R.BEGIN_EFF_DTTM,                 R.END_EFF_DTTM,                 K.CUST_NUM,                 K.GU_FULL_NAME AS CUSTOMER_NAME            FROM EDR.FACT_MORTGAGE_INSTR_MONTHLY A                 LEFT OUTER JOIN EDR.DIM_STATUS B                    ON     A.DIM_STATUS_KEY = B.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = B.SRC_SYS_CD                       AND B.CURRENT_FLG = 'Y'                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY C                    ON     A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                       AND C.CURRENT_FLG = 'Y' AND C.LEVEL_NUM IS NULL                 LEFT OUTER JOIN EDR.FACT_MORTGAGE_INSTR_MTHLY_SUPP D                    ON     A.YEAR_MONTH = D.YEAR_MONTH                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.SRC_INSTR_ID = D.SRC_INSTR_ID                 LEFT OUTER JOIN EDR.DIM_MO_GL_MAP E                    ON A.DIM_GL_MAP_KEY = E.DIM_MO_GL_MAP_KEY                    AND E.CURRENT_FLG = 'Y'                 LEFT OUTER JOIN DATASTORE_SBO.LOANMAST_MONTHEND F                    ON     A.CYCLE_DT_NUM = F.CYCLE_DT_NUM                       AND A.SRC_INSTR_ID = F.LOAN_NUMBER                 LEFT OUTER JOIN EDR.CUST_INSTR_RELATION R                     ON     A.SRC_INSTR_ID = R.ACCT_NUM_DERIVED                        AND R.SRC_SYS_CD = 'MO'                        AND R.DELETE_INDICATOR = 'N'                        AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                                R.BEGIN_EFF_DTTM                        AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                                NVL (R.END_EFF_DTTM, '31-DEC-9999')                 LEFT OUTER JOIN EDR.DIM_CUSTOMER K                     ON     R.CUST_NUM = K.CUST_NUM                        AND K.SRC_SYS_CD = 'MO'                        AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                                K.BEGIN_EFF_DTTM                        AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                                NVL (K.END_EFF_DTTM, '31-DEC-9999')                        AND K.DELETE_INDICATOR = 'N'                WHERE     A.YEAR_MONTH = :dt1_year_month                         AND A.SRC_SYS_CD = 'MO'                         AND A.PRINCIPAL_BAL_GL_ACCT_NUM IS NOT NULL                         AND A.PRINCIPAL_UNPAID_BAL <> 0)                             SELECT DISTINCT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.CUST_NUM,                 A.CUSTOMER_NAME,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 A.INTEREST_PAID_TO_DT,                 A.COLLATERAL,                 A.LOAN_STATUS,                 A.LOAN_ACCRUAL_STATUS,                 A.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD,                 A.ORIGINAL_RT,                 A.CURR_INT_RT,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL AS PRINC_ENDING_BAL,                 A.PRINCIPAL_UNPAID_BAL AS TOTAL_ENDING_BAL,                 A.INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 A.LOAN_PAYMENT_AMOUNT,                 A.INSTR_TYPE_CD_DESC,                 A.PRINCIPAL_BAL_GL_ACCT_NUM AS GL_ACCT_NUM,                 A.GL_INTEREST_INCOME,                 A.GL_ACCR_INTEREST_RECEIVABLE,                 A.INSTR_PROD_CD_DESC,                 A.PAST_DUE_CODE   FROM MO_INSTR A   WHERE SUBSTR (A.CUST_NUM, LENGTH (A.CUST_NUM), 1) = '1'"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)    
                
                """MO INT. INCOME"""
                SQL="WITH MO_INSTR      AS (SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 A.INTEREST_PAID_TO_DT,                 'Y' AS COLLATERAL,                 B.STATUS_DESC AS LOAN_STATUS,                 B.ACCRUAL_STATUS_DESC AS LOAN_ACCRUAL_STATUS,                 C.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD,                 A.ORIGINAL_RT,                 A.CURR_INT_RT,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_UNPAID_BAL,                 A.INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.LOAN_PAYMENT_AMOUNT,                 C.INSTR_TYPE_CD_DESC,                 A.PRINCIPAL_BAL_GL_ACCT_NUM,                 E.GL_INTEREST_INCOME,                 E.GL_ACCR_INTEREST_RECEIVABLE,                 C.INSTR_PROD_CD_DESC,                 A.RELATIONSHIP_MANAGER_NUM,                 F.PAST_DUE_CODE,                 R.BEGIN_EFF_DTTM,                 R.END_EFF_DTTM,                 K.CUST_NUM,                 K.GU_FULL_NAME AS CUSTOMER_NAME            FROM EDR.FACT_MORTGAGE_INSTR_MONTHLY A                 LEFT OUTER JOIN EDR.DIM_STATUS B                    ON     A.DIM_STATUS_KEY = B.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = B.SRC_SYS_CD                       AND B.CURRENT_FLG = 'Y'                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY C                    ON     A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                       AND C.CURRENT_FLG = 'Y' AND C.LEVEL_NUM IS NULL                 LEFT OUTER JOIN EDR.FACT_MORTGAGE_INSTR_MTHLY_SUPP D                    ON     A.YEAR_MONTH = D.YEAR_MONTH                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                       AND A.SRC_INSTR_ID = D.SRC_INSTR_ID                 LEFT OUTER JOIN EDR.DIM_MO_GL_MAP E                    ON A.DIM_GL_MAP_KEY = E.DIM_MO_GL_MAP_KEY                    AND E.CURRENT_FLG = 'Y'                 LEFT OUTER JOIN DATASTORE_SBO.LOANMAST_MONTHEND F                    ON     A.CYCLE_DT_NUM = F.CYCLE_DT_NUM                       AND A.SRC_INSTR_ID = F.LOAN_NUMBER                 LEFT OUTER JOIN EDR.CUST_INSTR_RELATION R                     ON     A.SRC_INSTR_ID = R.ACCT_NUM_DERIVED                        AND R.SRC_SYS_CD = 'MO'                        AND R.DELETE_INDICATOR = 'N'                        AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                                R.BEGIN_EFF_DTTM                        AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                                NVL (R.END_EFF_DTTM, '31-DEC-9999')                 LEFT OUTER JOIN EDR.DIM_CUSTOMER K                     ON     R.CUST_NUM = K.CUST_NUM                        AND K.SRC_SYS_CD = 'MO'                        AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                                K.BEGIN_EFF_DTTM                        AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                                NVL (K.END_EFF_DTTM, '31-DEC-9999')                        AND K.DELETE_INDICATOR = 'N'                WHERE     A.YEAR_MONTH = :dt1_year_month                         AND A.SRC_SYS_CD = 'MO'                         AND A.PRINCIPAL_BAL_GL_ACCT_NUM IS NOT NULL                         AND A.INTEREST_INCOME <> 0)                             SELECT DISTINCT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.CUST_NUM,                 A.CUSTOMER_NAME,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 A.INTEREST_PAID_TO_DT,                 A.COLLATERAL,                 A.LOAN_STATUS,                 A.LOAN_ACCRUAL_STATUS,                 A.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD,                 A.ORIGINAL_RT,                 A.CURR_INT_RT,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_UNPAID_BAL,                 A.INTEREST_INCOME AS TOTAL_INT_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 A.LOAN_PAYMENT_AMOUNT,                 A.INSTR_TYPE_CD_DESC,                 A.PRINCIPAL_BAL_GL_ACCT_NUM,                 A.GL_INTEREST_INCOME,                 A.GL_ACCR_INTEREST_RECEIVABLE,                 A.INSTR_PROD_CD_DESC,                 A.PAST_DUE_CODE   FROM MO_INSTR A   WHERE SUBSTR (A.CUST_NUM, LENGTH (A.CUST_NUM), 1) = '1'"
                print(SRC +' Intrest Income Query execution is in progress for ' +  dt1)
                Data_INT=self.execution(dt1,YM1,SRC,SQL)                   
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                 
                self.Excel(SRC,YM1)                     
                
            if self.checkBox_10.isChecked() == True:             
                SRC='MS'       
                
                """MS Query"""
                SQL="WITH MS_INSTR      AS (SELECT A.YEAR_MONTH,                 A.SRC_SYS_CD,                 A.COMPANY_NUM,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 A.INTEREST_PAID_TO_DT,                 'Y' AS COLLATERAL,                 B.STATUS_DESC AS LOAN_STATUS,                 B.ACCRUAL_STATUS_DESC AS LOAN_ACCRUAL_STATUS,                 C.INSTR_INT_TYPE_CD,                 A.ACCRUAL_METHOD,                 A.ORIGINAL_RT,                 A.CURR_INT_RT,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL,                 A.INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 D.TOT_PAYMT AS LOAN_PAYMENT_AMOUNT,                 C.INSTR_TYPE_CD_DESC,                 A.PRINCIPAL_BAL_GL_ACCT_NUM,                 E.GL_INTEREST_INCOME,                 E.GL_ACCR_INTEREST_RECEIVABLE,                 C.INSTR_PROD_CD_DESC,                 A.RELATIONSHIP_MANAGER_NUM AS CUST_NUM,                 A.TOTAL_PAST_DUE_DAYS,                 A.PAYMENTS_PAST_DUE_NUM,                 A.PAST_DUE_CD,                 A.PAST_DUE_CD_DESC,                 A.LAST_PAYMENT_EFFECTIVE_DT,                 A.CYCLE_DT_NUM,                 A.NEXT_DUE_DT,                 CASE                    WHEN (TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') - A.NEXT_DUE_DT) <=                            0                    THEN                       0                    ELSE                       ROUND (                          (  TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD')                           - A.NEXT_DUE_DT),                          0)                 END                    AS DAYS_PAST_DUE,                 A.CHARGE_OFF_PRINCIPAL_BAL,                 K.GU_FULL_NAME AS CUSTOMER_NAME            FROM EDR.FACT_MORTGAGE_INSTR_MONTHLY A                 LEFT OUTER JOIN EDR.DIM_STATUS B                    ON     A.DIM_STATUS_KEY = B.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = B.SRC_SYS_CD                       AND B.STATUS_CD = 'AC'                       AND B.CURRENT_FLG = 'Y'                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY C                    ON     A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = C.SRC_SYS_CD                       AND UPPER (TRIM (C.LEVEL3_CD_DESC)) = 'LOAN'                       AND C.HIERARCHY_TYPE = 'RBG_DDR'                       AND C.CURRENT_FLG = 'Y'                 LEFT OUTER JOIN STG_CPI.MS_MASTER D                    ON     A.CYCLE_DT_NUM = D.CYCLE_DT_NUM                       AND A.SRC_INSTR_ID = D.LOAN_NO                 LEFT OUTER JOIN EDR.DIM_MS_GL_MAP E                    ON A.DIM_GL_MAP_KEY = E.DIM_MS_GL_MAP_KEY                 LEFT OUTER JOIN EDR.CUST_INSTR_RELATION R                    ON     A.SRC_INSTR_ID = R.ACCT_NUM_DERIVED                       AND R.SRC_SYS_CD = 'RM'                       AND R.DELETE_INDICATOR = 'N'                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                              R.BEGIN_EFF_DTTM                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                              NVL (R.END_EFF_DTTM, '31-DEC-9999')                 LEFT OUTER JOIN EDR.DIM_CUSTOMER K                    ON     R.CUST_NUM = K.CUST_NUM                       AND K.SRC_SYS_CD = 'RM'                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                              K.BEGIN_EFF_DTTM                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                              NVL (K.END_EFF_DTTM, '31-DEC-9999')                       AND K.DELETE_INDICATOR = 'N'           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'MS'                 AND A.PRINCIPAL_BAL_GL_ACCT_NUM IS NOT NULL                 AND A.PRINCIPAL_ENDING_BAL <> 0) SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.CUST_NUM,        A.CUSTOMER_NAME,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        A.INTEREST_PAID_TO_DT,        A.COLLATERAL,        A.LOAN_STATUS,        A.LOAN_ACCRUAL_STATUS,        A.INSTR_INT_TYPE_CD,        A.ACCRUAL_METHOD,        A.ORIGINAL_RT,        A.CURR_INT_RT,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        A.LOAN_PAYMENT_AMOUNT,        A.INSTR_TYPE_CD_DESC,        A.PRINCIPAL_BAL_GL_ACCT_NUM as GL_ACCT_NUM,        A.GL_INTEREST_INCOME,        A.GL_ACCR_INTEREST_RECEIVABLE,        A.INSTR_PROD_CD_DESC,        A.DAYS_PAST_DUE,        A.CHARGE_OFF_PRINCIPAL_BAL,        A.PRINCIPAL_ENDING_BAL + NVL (CHARGE_OFF_PRINCIPAL_BAL, 0)           AS FINAL_ENDING_BAL   FROM MS_INSTR A"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)        
                
                
                """MS Int.Income Query"""
                SQL="SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.SRC_INSTR_ID,        A.EFFECTIVE_START_DT,        A.CURR_MATURITY_DT,        A.ORIGINAL_BAL,        A.INTEREST_PAID_TO_DT,        'Y' AS COLLATERAL,        B.STATUS_DESC AS LOAN_STATUS,        B.ACCRUAL_STATUS_DESC AS LOAN_ACCRUAL_STATUS,        C.INSTR_INT_TYPE_CD,        A.ACCRUAL_METHOD,        A.ORIGINAL_RT,        A.CURR_INT_RT,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.INTEREST_INCOME as TOTAL_INT_INCOME,        A.ACCR_INTEREST_RECEIVABLE,        D.TOT_PAYMT AS LOAN_PAYMENT_AMOUNT,        C.INSTR_TYPE_CD_DESC,        A.PRINCIPAL_BAL_GL_ACCT_NUM,        E.GL_INTEREST_INCOME,        E.GL_ACCR_INTEREST_RECEIVABLE,        C.INSTR_PROD_CD_DESC,        A.RELATIONSHIP_MANAGER_NUM AS CUST_NUM,        A.TOTAL_PAST_DUE_DAYS,        A.PAYMENTS_PAST_DUE_NUM,        A.PAST_DUE_CD,        A.PAST_DUE_CD_DESC,        A.LAST_PAYMENT_EFFECTIVE_DT,        A.CYCLE_DT_NUM,        A.NEXT_DUE_DT,        CASE           WHEN (TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') - A.NEXT_DUE_DT) <= 0           THEN              0           ELSE              ROUND ( (TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') - A.NEXT_DUE_DT),                     0)        END           AS DAYS_PAST_DUE   FROM (Select * from EDR.FACT_MORTGAGE_INSTR_MONTHLY A WHERE A.YEAR_MONTH = :dt1_year_month AND A.SRC_SYS_CD = 'MS' AND A.INTEREST_INCOME <> 0) A        LEFT OUTER JOIN (Select * from EDR.DIM_STATUS B WHERE B.CURRENT_FLG = 'Y' AND B.STATUS_CD = 'AC' ) B           ON     A.DIM_STATUS_KEY = B.DIM_STATUS_KEY              AND A.SRC_SYS_CD = B.SRC_SYS_CD        LEFT OUTER JOIN (Select * from EDR.DIM_PRODUCT_HIERARCHY C where C.CURRENT_FLG = 'Y' and UPPER (TRIM (C.LEVEL3_CD_DESC)) = 'LOAN') C           ON     A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY              AND A.SRC_SYS_CD = C.SRC_SYS_CD        LEFT OUTER JOIN STG_CPI.MS_MASTER D           ON A.CYCLE_DT_NUM = D.CYCLE_DT_NUM AND A.SRC_INSTR_ID = D.LOAN_NO        LEFT OUTER JOIN (select * from EDR.DIM_MS_GL_MAP E WHERE E.GL_INTEREST_INCOME IS NOT NULL and  E.CURRENT_FLG = 'Y') E           ON A.DIM_GL_MAP_KEY = E.DIM_MS_GL_MAP_KEY"
                print(SRC +' Intrest Income Query execution is in progress for ' +  dt1)
                Data_INT=self.execution(dt1,YM1,SRC,SQL)       
                
                
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                 
                self.Excel(SRC,YM1)                 
                
            if self.checkBox_11.isChecked() == True:          
                SRC='LoanFee-533975'                
                GL_SUMM=GL_SUMM_fee
                
                SQL="SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.SRC_INSTR_ID,        B.GL_SYNDICATION_FEE_2 AS GL_ACCT_NUM,        A.AF_SYNDICATION_FEE_2 AS TOTAL_ENDING_BAL   FROM EDR.FACT_COM_LOAN_FACIL_MTHLY A        LEFT OUTER JOIN EDR.DIM_AF_GL_MAP B           ON A.DIM_GL_MAP_KEY = B.DIM_AF_GL_MAP_KEY AND B.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_STATUS C           ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY              AND A.SRC_SYS_CD = C.SRC_SYS_CD              AND C.CURRENT_FLG = 'Y'  WHERE     A.YEAR_MONTH in (select YEAR_MONTH2 FROM edr.period where year_month2 <= :dt1_year_month AND year_NM = SUBSTR ( :dt1_year_month, 1, 4) group by year_month2 )         AND A.SRC_SYS_CD = 'AF'        AND C.STATUS_CD = 'Y'        AND A.AF_SYNDICATION_FEE_2 <> 0        AND B.GL_SYNDICATION_FEE_2 = 533975        order by 1"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)               
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                 
                self.Excel(SRC,YM1) 
                
            if self.checkBox_12.isChecked() == True:               
                SRC='LoanFee-534070'                
                GL_SUMM=GL_SUMM_fee
                
                SQL="SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.SRC_INSTR_ID,        B.GL_SYNDICATION_FEE_3 AS GL_ACCT_NUM,        A.AF_SYNDICATION_FEE_3 AS TOTAL_ENDING_BAL   FROM EDR.FACT_COM_LOAN_FACIL_MTHLY A        LEFT OUTER JOIN EDR.DIM_AF_GL_MAP B           ON A.DIM_GL_MAP_KEY = B.DIM_AF_GL_MAP_KEY AND B.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_STATUS C           ON     A.DIM_STATUS_KEY = C.DIM_STATUS_KEY              AND A.SRC_SYS_CD = C.SRC_SYS_CD              AND C.CURRENT_FLG = 'Y'  WHERE     A.YEAR_MONTH IN (SELECT YEAR_MONTH2                               FROM edr.period                              WHERE     year_month2 <= :dt1_year_month                                    AND year_NM =                                           SUBSTR ( :dt1_year_month, 1, 4)                             GROUP BY year_month2)        AND A.SRC_SYS_CD = 'AF'        AND C.STATUS_CD = 'Y'        AND A.AF_SYNDICATION_FEE_3 <> 0        AND B.GL_SYNDICATION_FEE_3 = 534070 ORDER BY A.YEAR_MONTH"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)               
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                 
                self.Excel(SRC,YM1)                     
                
            if self.checkBox_13.isChecked() == True:            
                SRC='ST'                
                SQL="SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        SUM(A.INTEREST_EXPENSE) AS TOTAL_INT_INCOME,        B.GL_INTEREST_EXPENSE AS GL_INTEREST_INCOME,        SUM(A.PRINCIPAL_ENDING_BAL) as PRINCIPAL_ENDING_BAL,        B.GL_PRINCIPAL AS GL_ACCT_NUM   FROM EDR.FACT_DEPOSITS_MONTHLY A        LEFT OUTER JOIN EDR.DIM_ST_GL_MAP B           ON A.DIM_GL_MAP_KEY = B.DIM_ST_GL_MAP_KEY AND B.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_ORG_HIERARCHY C           ON A.DIM_ORG_KEY = C.DIM_ORG_KEY AND C.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D           ON     A.SRC_SYS_CD = D.SRC_SYS_CD              AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY              AND D.HIERARCHY_TYPE = 'RBG_DDR'              AND D.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_DEPOSIT_PROFILE E           ON     A.DIM_INSTRUMENT_KEY = E.DIM_INSTRUMENT_KEY              AND E.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_CUSTOMER F           ON     A.SRC_SYS_CD = F.SRC_SYS_CD              AND A.SRC_INSTR_ID =                     LTRIM (SUBSTR (F.CUST_NUM, 1, LENGTH (F.CUST_NUM) - 2),                            '0')              AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >= F.BEGIN_EFF_DTTM              AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                     NVL (F.END_EFF_DTTM, '31-DEC-9999')              AND F.DELETE_INDICATOR = 'N'              AND SUBSTR (F.CUST_NUM, LENGTH (F.CUST_NUM), 1) = '1'  WHERE     A.SRC_SYS_CD = 'ST'        AND A.YEAR_MONTH = :dt1_year_month        AND C.COMPANY = '0002'        AND (A.PRINCIPAL_ENDING_BAL <> 0 OR A.INTEREST_EXPENSE <> 0) Group by     A.SRC_SYS_CD,     A.YEAR_MONTH,     B.GL_INTEREST_EXPENSE,     B.GL_PRINCIPAL"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                 
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)  
                
            if self.checkBox_14.isChecked() == True:        
                SRC='IM'         
                
                SQL="SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        SUM (A.INTEREST_EXPENSE) AS TOTAL_INT_INCOME,        to_number(nvl(B.IM_GL_INTEREST_EXPENSE,-1)) AS GL_INTEREST_INCOME,        SUM (A.PRINCIPAL_ENDING_BAL) AS PRINCIPAL_ENDING_BAL,        to_number(nvl(CASE           WHEN A.PRINCIPAL_ENDING_BAL >= 0 THEN B.IM_GL_PRINCIPAL           ELSE B.IM_GL_OVERDRAFT        END,-1))           AS GL_ACCT_NUM   FROM EDR.FACT_DEPOSITS_MONTHLY A        LEFT OUTER JOIN EDR.DIM_IM_GL_MAP B           ON A.DIM_GL_MAP_KEY = B.DIM_IM_GL_MAP_KEY AND B.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_ORG_HIERARCHY C           ON A.DIM_ORG_KEY = C.DIM_ORG_KEY AND C.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D           ON     A.SRC_SYS_CD = D.SRC_SYS_CD              AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY              AND D.HIERARCHY_TYPE = 'RBG_DDR'              AND D.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_DEPOSIT_PROFILE E           ON     A.DIM_INSTRUMENT_KEY = E.DIM_INSTRUMENT_KEY              AND E.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_CUSTOMER F           ON     A.SRC_SYS_CD = F.SRC_SYS_CD              AND A.SRC_INSTR_ID =                     LTRIM (SUBSTR (F.CUST_NUM, 1, LENGTH (F.CUST_NUM) - 2),                            '0')              AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >= F.BEGIN_EFF_DTTM              AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                     NVL (F.END_EFF_DTTM, '31-DEC-9999')              AND F.DELETE_INDICATOR = 'N'              AND SUBSTR (F.CUST_NUM, LENGTH (F.CUST_NUM), 1) = '1'  WHERE     A.SRC_SYS_CD = 'IM'        AND A.YEAR_MONTH = :dt1_year_month        AND C.COMPANY = '0002'        AND (A.PRINCIPAL_ENDING_BAL <> 0 OR A.INTEREST_EXPENSE <> 0) GROUP BY A.SRC_SYS_CD,          A.YEAR_MONTH,          B.IM_GL_INTEREST_EXPENSE,          CASE             WHEN A.PRINCIPAL_ENDING_BAL >= 0 THEN B.IM_GL_PRINCIPAL             ELSE B.IM_GL_OVERDRAFT          END"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)           
                
            if self.checkBox_15.isChecked() == True:           
                SRC='LG'     
                
                SQL="WITH LG_PM_INSTR AS (  SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        C.COMPANY,        A.SRC_INSTR_ID,        F.GU_FULL_NAME AS CUSTOMER_NAME,        A.INTEREST_RATE,        A.INTEREST_EXPENSE,        A.PRINCIPAL_ENDING_BAL AS ENDING_BAL,        B.GL_PRINCIPAL_ACCT_NUM AS GL_ACCT_NUM,        A.MTD_AVG_BAL AS AVERAGE_BAL,        D.INSTR_PROD_CD AS SRC_ACCT_TYPE,        D.INSTR_PROD_CD_DESC AS PRODUCT_DESCRIPTION,        E.FIXED_VARIABLE_RATE_IND,        A.DIM_INSTRUMENT_KEY,        C.COST_CENTER   FROM EDR.FACT_DEPOSITS_MONTHLY A   LEFT OUTER JOIN EDR.DIM_LG_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LG_GL_MAP_KEY    AND B.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_ORG_HIERARCHY C     ON A.DIM_ORG_KEY = C.DIM_ORG_KEY    AND C.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND D.HIERARCHY_TYPE = 'TCMO'    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_DEPOSIT_PROFILE E     ON A.DIM_INSTRUMENT_KEY = E.DIM_INSTRUMENT_KEY    AND E.CURRENT_FLG = 'Y'   LEFT JOIN EDR.CUST_INSTR_RELATION H     ON A.SRC_INSTR_ID = H.ACCT_NUM_DERIVED    AND A.SRC_SYS_CD = H.SRC_SYS_CD    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= H.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(H.END_EFF_DTTM, '31-DEC-9999')    AND H.DELETE_INDICATOR = 'N'    LEFT OUTER JOIN EDR.DIM_CUSTOMER F     ON F.CUST_NUM = H.CUST_NUM    AND F.SRC_SYS_CD = H.SRC_SYS_CD    AND F.CURRENT_FLAG = 'Y'    AND F.DELETE_INDICATOR = 'N'   LEFT OUTER JOIN EDR.EDR_CODE_MAP MAP     ON MAP.SRC_SYS_CD = F.SRC_SYS_CD    AND MAP.CODE_DEFINITION_NAME = 'LG:CUSTOMER TYPE DESC'    AND MAP.ENTITY_COLUMN_NAME = 'CUSTOMER_TYPE'    AND MAP.SOURCE_CODE = F.SRC_CUST_TYP_CD    AND MAP.ENTITY_NAME = 'DIM_DEPOSIT_PROFILE'    AND MAP.CURRENT_FLG = 'Y'  WHERE A.SRC_SYS_CD = 'LG'    AND A.YEAR_MONTH = :dt1_year_month    AND C.COMPANY = '0002'    AND (A.PRINCIPAL_ENDING_BAL <> 0 OR A.INTEREST_EXPENSE <> 0)    AND A.SHADOW_DEPOSIT_FLG = 'N'  UNION ALL   SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        C.COMPANY,        X.SRC_INSTR_ID,        F.GU_FULL_NAME AS CUSTOMER_NAME,        A.INTEREST_RATE,        A.MTD_INTEREST_AMT AS INTEREST_EXPENSE,        A.PRINCIPAL_AMT AS ENDING_BAL,        B.GL_PRINCIPAL_ACCT_NUM AS GL_ACCT_NUM,        A.MTD_AVERAGE_BAL AS AVERAGE_BAL,        D.INSTR_PROD_CD AS SRC_ACCT_TYPE,        D.INSTR_PROD_CD_DESC AS PRODUCT_DESCRIPTION,        E.FIXED_VARIABLE_RATE_IND,        A.DIM_INSTRUMENT_KEY,        C.COST_CENTER   FROM EDR.FACT_MISC_LIABILITY_MONTHLY  A   LEFT JOIN EDR.DIM_INSTRUMENT X   ON A.DIM_INSTRUMENT_KEY = X.DIM_INSTRUMENT_KEY         LEFT OUTER JOIN EDR.DIM_LG_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_LG_GL_MAP_KEY    AND B.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_ORG_HIERARCHY C     ON A.DIM_ORG_KEY = C.DIM_ORG_KEY    AND C.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND D.HIERARCHY_TYPE = 'TCMO'    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_DEPOSIT_PROFILE E     ON A.DIM_INSTRUMENT_KEY = E.DIM_INSTRUMENT_KEY    AND E.CURRENT_FLG = 'Y'   LEFT JOIN EDR.CUST_INSTR_RELATION H     ON X.SRC_INSTR_ID = H.ACCT_NUM_DERIVED    AND A.SRC_SYS_CD = H.SRC_SYS_CD    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= H.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(H.END_EFF_DTTM, '31-DEC-9999')    AND H.DELETE_INDICATOR = 'N'    LEFT OUTER JOIN EDR.DIM_CUSTOMER F     ON F.CUST_NUM = H.CUST_NUM    AND F.SRC_SYS_CD = H.SRC_SYS_CD    AND F.CURRENT_FLAG = 'Y'    AND F.DELETE_INDICATOR = 'N'   LEFT OUTER JOIN EDR.EDR_CODE_MAP MAP     ON MAP.SRC_SYS_CD = F.SRC_SYS_CD    AND MAP.CODE_DEFINITION_NAME = 'LG:CUSTOMER TYPE DESC'    AND MAP.ENTITY_COLUMN_NAME = 'CUSTOMER_TYPE'    AND MAP.SOURCE_CODE = F.SRC_CUST_TYP_CD    AND MAP.ENTITY_NAME = 'DIM_DEPOSIT_PROFILE'    AND MAP.CURRENT_FLG = 'Y'  WHERE A.SRC_SYS_CD = 'LG'    AND A.YEAR_MONTH = :dt1_year_month    AND C.COMPANY = '0002'    AND (A.PRINCIPAL_AMT <> 0 OR A.MTD_INTEREST_AMT <> 0)    AND B.SHADOW_DEPOSIT_FLG = 'N'    UNION ALL   SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        C.COMPANY,        A.SRC_INSTR_ID,        F.GU_FULL_NAME AS CUSTOMER_NAME,        A.INTEREST_RATE,        A.INTEREST_EXPENSE,        A.PRINCIPAL_ENDING_BAL AS ENDING_BAL,        B.GL_PRINCIPAL_ACCT_NUM AS GL_ACCT_NUM,        A.MTD_AVG_BAL AS AVERAGE_BAL,        D.INSTR_PROD_CD AS SRC_ACCT_TYPE,        D.INSTR_PROD_CD_DESC AS PRODUCT_DESCRIPTION,        E.FIXED_VARIABLE_RATE_IND,        A.DIM_INSTRUMENT_KEY,        C.COST_CENTER   FROM EDR.FACT_DEPOSITS_MONTHLY A   LEFT OUTER JOIN EDR.DIM_PM_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_PM_GL_MAP_KEY    AND B.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_ORG_HIERARCHY C     ON A.DIM_ORG_KEY = C.DIM_ORG_KEY    AND C.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND D.HIERARCHY_TYPE = 'TCMO'    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_DEPOSIT_PROFILE E     ON A.DIM_INSTRUMENT_KEY = E.DIM_INSTRUMENT_KEY    AND E.CURRENT_FLG = 'Y'   LEFT JOIN EDR.CUST_INSTR_RELATION H     ON A.SRC_INSTR_ID = H.ACCT_NUM_DERIVED    AND A.SRC_SYS_CD = H.SRC_SYS_CD    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= H.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(H.END_EFF_DTTM, '31-DEC-9999')    AND H.DELETE_INDICATOR = 'N'    LEFT OUTER JOIN EDR.DIM_CUSTOMER F     ON F.CUST_NUM = H.CUST_NUM    AND F.SRC_SYS_CD = H.SRC_SYS_CD    AND F.CURRENT_FLAG = 'Y'    AND F.DELETE_INDICATOR = 'N'   LEFT OUTER JOIN EDR.EDR_CODE_MAP MAP     ON MAP.SRC_SYS_CD = F.SRC_SYS_CD    AND MAP.CODE_DEFINITION_NAME = 'PM:CUSTOMER_TYPE'    AND MAP.ENTITY_COLUMN_NAME = 'CUSTOMER_TYPE'    AND MAP.SOURCE_CODE = F.SRC_CUST_TYP_CD    AND MAP.ENTITY_NAME = 'DIM_DEPOSIT_PROFILE'    AND MAP.CURRENT_FLG = 'Y'  WHERE A.SRC_SYS_CD = 'PM'    AND A.YEAR_MONTH = :dt1_year_month    AND C.COMPANY = '0002'    AND (A.PRINCIPAL_ENDING_BAL <> 0 OR A.INTEREST_EXPENSE <> 0)    AND A.SHADOW_DEPOSIT_FLG = 'N'  UNION ALL SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        C.COMPANY,        X.SRC_INSTR_ID,        F.GU_FULL_NAME AS CUSTOMER_NAME,        A.INTEREST_RATE,        A.MTD_INTEREST_AMT AS INTEREST_EXPENSE,        A.PRINCIPAL_AMT AS ENDING_BAL,        B.GL_PRINCIPAL_ACCT_NUM AS GL_ACCT_NUM,        A.MTD_AVERAGE_BAL AS AVERAGE_BAL,        D.INSTR_PROD_CD AS SRC_ACCT_TYPE,        D.INSTR_PROD_CD_DESC AS PRODUCT_DESCRIPTION,        E.FIXED_VARIABLE_RATE_IND,        A.DIM_INSTRUMENT_KEY,        C.COST_CENTER   FROM EDR.FACT_MISC_LIABILITY_MONTHLY  A   LEFT JOIN EDR.DIM_INSTRUMENT X   ON A.DIM_INSTRUMENT_KEY = X.DIM_INSTRUMENT_KEY         LEFT OUTER JOIN EDR.DIM_PM_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_PM_GL_MAP_KEY    AND B.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_ORG_HIERARCHY C     ON A.DIM_ORG_KEY = C.DIM_ORG_KEY    AND C.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND D.HIERARCHY_TYPE = 'TCMO'    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_DEPOSIT_PROFILE E     ON A.DIM_INSTRUMENT_KEY = E.DIM_INSTRUMENT_KEY    AND E.CURRENT_FLG = 'Y'   LEFT JOIN EDR.CUST_INSTR_RELATION H     ON X.SRC_INSTR_ID = H.ACCT_NUM_DERIVED    AND A.SRC_SYS_CD = H.SRC_SYS_CD    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= H.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(H.END_EFF_DTTM, '31-DEC-9999')    AND H.DELETE_INDICATOR = 'N'    LEFT OUTER JOIN EDR.DIM_CUSTOMER F     ON F.CUST_NUM = H.CUST_NUM    AND F.SRC_SYS_CD = H.SRC_SYS_CD    AND F.CURRENT_FLAG = 'Y'    AND F.DELETE_INDICATOR = 'N'   LEFT OUTER JOIN EDR.EDR_CODE_MAP MAP     ON MAP.SRC_SYS_CD = F.SRC_SYS_CD    AND MAP.CODE_DEFINITION_NAME = 'PM:CUSTOMER_TYPE'    AND MAP.ENTITY_COLUMN_NAME = 'CUSTOMER_TYPE'    AND MAP.SOURCE_CODE = F.SRC_CUST_TYP_CD    AND MAP.ENTITY_NAME = 'DIM_DEPOSIT_PROFILE'    AND MAP.CURRENT_FLG = 'Y'  WHERE A.SRC_SYS_CD = 'PM'    AND A.YEAR_MONTH = :dt1_year_month    AND C.COMPANY = '0002'    AND (A.PRINCIPAL_AMT <> 0 OR A.MTD_INTEREST_AMT <> 0)    AND B.SHADOW_DEPOSIT_FLG = 'N' ), LG_PM_GL_INT_EXPENSE AS (  SELECT DISTINCT        A.CYCLE_DT_NUM,        A.JOURNAL_CODE AS JOURNAL_CD,        A.DESCR,        A.ACCTNO AS GL_INTEREST_EXPENSE   FROM DATASTORE_LG.MSUMMARY A  WHERE DESCR LIKE '%INTEREST EXPENSE%'    AND CYCLE_DT_NUM = :dt1_cycle_dt_num ) SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        A.COMPANY,        A.SRC_INSTR_ID,        A.CUSTOMER_NAME,        A.INTEREST_RATE,        A.INTEREST_EXPENSE AS TOTAL_INT_INCOME,        TO_NUMBER (NVL (B.GL_INTEREST_EXPENSE, -1)) AS GL_INTEREST_INCOME,        A.ENDING_BAL AS PRINCIPAL_ENDING_BAL,        A.GL_ACCT_NUM,        A.AVERAGE_BAL,        A.SRC_ACCT_TYPE,        A.PRODUCT_DESCRIPTION,        A.FIXED_VARIABLE_RATE_IND,        A.COST_CENTER   FROM LG_PM_INSTR A   LEFT JOIN LG_PM_GL_INT_EXPENSE B     ON A.SRC_ACCT_TYPE = B.JOURNAL_CD  WHERE A.GL_ACCT_NUM NOT IN ('229161', '229153')  ORDER BY A.SRC_INSTR_ID"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                 
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM) 
                self.Excel(SRC,YM1) 
                
            if self.checkBox_16.isChecked() == True:    
                SRC='DepoFee550139'                    
                
                SQL="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,           550139 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('3005', '3006')     AND D.INSTR_TYPE_CD IN ('802')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     550139 order by 1"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                GL_SUMM=GL_SUMM_fee
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                 
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                 
                
            if self.checkBox_17.isChecked() == True:           
                SRC='DepoFee550141'                
                GL_SUMM=GL_SUMM_fee
                SQL="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,           550141 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('3007', '3008')     AND D.INSTR_TYPE_CD IN ('802')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     550141 order by 1"     
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                 
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                  

            if self.checkBox_18.isChecked() == True:              
                SRC='DepoFee550557'                
                GL_SUMM=GL_SUMM_fee
                SQL="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,           550557 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('3007', '3008')     AND D.INSTR_TYPE_CD IN ('890') 	GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     550557 order by 1"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                
 
            if self.checkBox_19.isChecked() == True:            
                SRC='DepoFee550569'                
                GL_SUMM=GL_SUMM_fee
                SQL="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,           550569 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('3005', '3006')     AND D.INSTR_TYPE_CD IN ('890') 	GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     550569 order by 1"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)          
                
            if self.checkBox_20.isChecked() == True:
                SRC='DepoFee552872'                
                GL_SUMM=GL_SUMM_fee
                SQL="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,        C.SRC_INSTR_ID,        D.INSTR_TYPE_CD,        B.TRANSACTION_CD,        552872 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */     X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('0096')     AND D.INSTR_TYPE_CD IN ('087')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     C.SRC_INSTR_ID,     D.INSTR_TYPE_CD,     B.TRANSACTION_CD,     552872 order by 1"     
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                 
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM) 


            if self.checkBox_21.isChecked() == True:            
                SRC='DepoFee552877'                
                GL_SUMM=GL_SUMM_fee
                SQL="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,           552877 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN  ('0096')     AND D.INSTR_TYPE_CD IN ('010', '015', '016', '017', '019', '020', '053', '060', '061', '142')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     552877 order by 1"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                 
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)               

            if self.checkBox_22.isChecked() == True:           
                SRC='DF550129and550710'             
                SQL="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,           550710550129 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('0068')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     550710550129"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1)  
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                 
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)   
#                self.Excel(SRC,YM1)  
                
            if self.checkBox_23.isChecked() == True:             
                SRC='TSYS'          
                
                """TSYS Query"""
                SQL="SELECT CYCLE_DT_NUM,        to_number(SUBSTR (CYCLE_DT_NUM, 1, 6)) AS YEAR_MONTH,        ENTITY_ID,        ENTITY_DESC,        SRC_INSTR_ID,        SRC_SYS_CD,        COMPANY_NUM,        COST_CTR_NUM,        to_number(GL_ACCT_NUM) as GL_ACCT_NUM,        PRODUCT_CLASSIFICATION,        LOAN_COUNTERPARTY_CD,        AMORTIZATION_CD,        NON_ACCRUAL_CD,        PAYMENT_FREQ_CD,        INTEREST_RATE,        INTEREST_RATE_TYPE_CD,        INTEREST_INDEX_CD,        INTEREST_MARGIN_PCT,        ORIGINATION_DT,        MATURITY_DT,        ACCRUAL_METHOD_CD,        COLLATERAL_DESC,        NON_CONFORMING_LOAN,        AF_FED_CLASS,        AF_PROCESS_TYPE,        AF_OBLIGATION_TYPE,        PAYMENT_AMT,        PRINCIPAL_ENDING_BAL,        RECORD_NUM,        BALLOON_AMT,        NEXT_REPRICE_DT,        EU_COUNTERPARTY_DESC,        EU_CPTY_CLASS_CD,        US_CPTY_CLASS_CD,        FIVE_G_CPTY_CLASS_CD,        CUST_MSTR_ID,        DG_SRC_SYS_CD,        CUST_RMPM_ID,        HELD_FOR_SALE_IND,        ASSET_CATEGORY_CD,        ASSET_CATEGORY_DESC,        LENDABLE_VALUE,        MARKET_VALUE,        COLLATERAL_VALUE,        PLEDGE_COLLATERAL_TYPE,        PLEDGE_IND,        PLEDGE_INSTITUTION,        REVOLVING_LINE_IND,        SECURED_LOAN_IND,        SECURITIZATION_IND,        SECURITIZATION_TYPE,        SYNDICATION_IND,        PAYMENT_DT   FROM FARM.FACT_BHC_INSTR_LOANS_CF  WHERE CYCLE_DT_NUM = :dt1_cycle_dt_num AND src_sys_cd = 'TSYS_COMM'"
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)                  
                
                """TSYS Int. Income Query"""
                SQL="SELECT A.COMPANY_NUM,        to_number(A.GL_ACCT_NUM) as GL_INTEREST_INCOME,        A.COST_CTR_NUM,        A.EFFECTIVE_DT_NUM,        A.POST_DT_NUM,        A.TRAN_DT_NUM,        A.TRAN_AMT as TOTAL_INT_INCOME,        A.TRAN_CD,        A.TRAN_DESC,        A.DOC_NBR,        A.GEN_DOC_NBR,        A.UNIQUE_SEQ_NBR,        A.GL_BATCH_NUM,        A.USER_BATCH_NUM,        A.PREPARER_ID,        A.APPROVER_ID   FROM GLM.FACT_GL_TRANSACTION A  WHERE A.EFFECTIVE_DT_NUM = (select min(CYCLE_DT_NUM2) as EFFECTIVE_DT_NUM from EDR.period where YEAR_MONTH2=(:dt1_year_month)+1                               and CYCLE_DT_NUM2 is not null and weekday_FLG='Y')    AND A.POST_DT_NUM = (  select min(CYCLE_DT_NUM2) as EFFECTIVE_DT_NUM from EDR.period where YEAR_MONTH2=(:dt1_year_month)+1                             and CYCLE_DT_NUM2 is not null and weekday_FLG='Y')    AND TRIM(UPPER(A.TRAN_DESC)) LIKE '%SETTLEMENT%'    AND A.GL_ACCT_NUM = '126120'"
                Data_INT=self.execution(dt1,YM1,SRC,SQL)  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)               
                
            if self.checkBox_24.isChecked() == True:             
                SRC='LoanFee-533901'                
                GL_SUMM=GL_SUMM_fee
                
                SQL="WITH AF_MISC_FEE_INCOME_INSTR AS ( SELECT SUBSTR(TO_CHAR(CYCLE_DT_NUM),1,6) AS YEAR_MONTH,        'AF' AS SRC_SYS_CD,        LPAD(OBLIGOR,10,'0') || '-' || LPAD(OBLIGATION,5,'0') AS SRC_INSTR_ID,        OBLIGOR,        OBLIGATION,        ASSIGN_UNIT,        GL_CODE,        CHARGE_CODE,        PROCESS_TYPE,        533901 AS GL_MISC_FEE_INCOME,        EARNED_ITD,        EARNED_MTD,        CASE WHEN TO_NUMBER(SUBSTR(PROCESS_TYPE,1,1)) <= 3              AND TO_NUMBER(SUBSTR(PROCESS_TYPE,2,1)) IN (5,8)              AND GL_CODE = '54'              AND TO_NUMBER(CHARGE_CODE) BETWEEN 530 AND 536             THEN 1             ELSE 0        END AS SYNDICATION_FEES_IND,        'NORMAL LOANS' AS SOURCE   FROM DATASTORE_AFS.FEE_CST  WHERE CYCLE_DT_NUM = :dt1_cycle_dt_num    AND TO_NUMBER(GL_CODE) IN (1,2,3,4,5,6,7,23,25,26,27,28,29,41,43,44,47,49,54)    AND (        (TO_NUMBER(CHARGE_CODE) BETWEEN 500 AND 510) OR        (TO_NUMBER(CHARGE_CODE) BETWEEN 513 AND 527) OR        (TO_NUMBER(CHARGE_CODE) BETWEEN 540 AND 549) OR         CHARGE_CODE LIKE '4%'        )  ),  AF_MISC_FEE_INCOME_INSTR_PREV AS ( SELECT SUBSTR(TO_CHAR(A.CYCLE_DT_NUM),1,6) AS YEAR_MONTH,        'AF' AS SRC_SYS_CD,        LPAD(OBLIGOR,10,'0') || '-' || LPAD(OBLIGATION,5,'0') AS SRC_INSTR_ID,        OBLIGOR,        OBLIGATION,        ASSIGN_UNIT,        GL_CODE,        CHARGE_CODE,        PROCESS_TYPE,        533901 AS GL_MISC_FEE_INCOME,        EARNED_ITD,        EARNED_MTD,        'NORMAL LOANS' AS SOURCE   FROM DATASTORE_AFS.FEE_CST A INNER JOIN             (SELECT max(B.CYCLE_DT_NUM2) as CYCLE_DT_NUM                           FROM (SELECT YEAR_MONTH2                               FROM edr.period                              WHERE     year_month2 = :dt1_year_month                                   ) A, EDR.PERIOD B                      WHERE A.YEAR_MONTH2-1=B.YEAR_MONTH2                            and WEEKDAY_FLG = 'Y') B ON(A.CYCLE_DT_NUM=B.CYCLE_DT_NUM)    AND TO_NUMBER(GL_CODE) IN (1,2,3,4,5,6,7,23,25,26,27,28,29,41,43,44,47,49,54)    AND (        (TO_NUMBER(CHARGE_CODE) BETWEEN 500 AND 510) OR        (TO_NUMBER(CHARGE_CODE) BETWEEN 513 AND 527) OR        (TO_NUMBER(CHARGE_CODE) BETWEEN 540 AND 549) OR         CHARGE_CODE LIKE '4%'        )        ), PARTSOLD_TO_INSTR_RELTN AS (SELECT DISTINCT LPAD(LI.OBLIGOR,10,'0') || '-' || LPAD(LI.OBLIGATION,5,'0') AS BASE_INSTR_ID,        SR.SRC_INSTR_ID PARTSOLD_TO_INSTR_ID   FROM DATASTORE_AFS.FEE_CST LI,        EDR.FACT_PART_SYND_REFERENCE_DAILY SR  WHERE LI.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND SR.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND LPAD(LI.OBLIGOR,10,'0') || '-' || LPAD(LI.OBLIGATION,5,'0') = SR.REF_SRC_INSTR_ID    AND SUBSTR(SR.PROCESS_TYPE,2,1) = 3  ORDER BY LPAD(LI.OBLIGOR,10,'0') || '-' || LPAD(LI.OBLIGATION,5,'0') ), PARTSOLD_TO_INSTR AS ( SELECT SUBSTR(TO_CHAR(LI.CYCLE_DT_NUM),1,6) AS YEAR_MONTH,        'AF' AS SRC_SYS_CD,        LPAD(LI.OBLIGOR,10,'0') || '-' || LPAD(LI.OBLIGATION,5,'0') AS PARTSOLD_INSTR_ID,             LI.CHARGE_CODE,        LI.PROCESS_TYPE,        LI.GL_CODE   FROM DATASTORE_AFS.FEE_CST LI  WHERE LI.CYCLE_DT_NUM = :dt1_cycle_dt_num    AND LPAD(LI.OBLIGOR,10,'0') || '-' || LPAD(LI.OBLIGATION,5,'0')     IN (SELECT PARTSOLD_TO_INSTR_ID           FROM PARTSOLD_TO_INSTR_RELTN        )     ) SELECT to_number(A.YEAR_MONTH) as YEAR_MONTH,        A.SRC_SYS_CD,        A.SRC_INSTR_ID,        A.OBLIGOR,        A.OBLIGATION,        A.ASSIGN_UNIT AS COST_CENTER,        A.GL_CODE,        A.CHARGE_CODE,        A.PROCESS_TYPE,        A.GL_MISC_FEE_INCOME as GL_ACCT_NUM,        A.EARNED_ITD,        A.EARNED_MTD,        NVL(C.EARNED_ITD,0) AS PREV_EARNED_ITD,        CASE WHEN B.PARTSOLD_INSTR_ID IS NOT NULL             THEN (A.EARNED_ITD - NVL(C.EARNED_ITD,0)) * -1             ELSE (A.EARNED_ITD - NVL(C.EARNED_ITD,0))        END as TOTAL_ENDING_BAL,        CASE WHEN B.PARTSOLD_INSTR_ID IS NOT NULL             THEN 1             ELSE 0        END AS PARTSOLD_IND,        A.SOURCE   FROM AF_MISC_FEE_INCOME_INSTR A   LEFT JOIN PARTSOLD_TO_INSTR B     ON A.SRC_INSTR_ID = B.PARTSOLD_INSTR_ID    AND A.GL_CODE = B.GL_CODE    AND A.CHARGE_CODE = B.CHARGE_CODE    AND A.PROCESS_TYPE = B.PROCESS_TYPE   LEFT JOIN AF_MISC_FEE_INCOME_INSTR_PREV C     ON A.SRC_INSTR_ID = C.SRC_INSTR_ID    AND A.CHARGE_CODE = C.CHARGE_CODE  WHERE (A.EARNED_ITD - NVL(C.EARNED_ITD,0)) <> 0    AND A.SYNDICATION_FEES_IND = 0 "
                print('STEP2: ')
                print(SRC +' Query execution is in progress for ' +  dt1) 
                Data_BAL=self.execution(dt1,YM1,SRC,SQL)               
                Data_INT=1 # Dummy value  
                Data_PAR=1 #dummy value
                self.GL_Recon(SRC,dt1,YM1,Data_BAL,Data_INT,Data_PAR,GL_SUMM)                 
                self.Excel(SRC,YM1)        
                
                
                
                
            #end of first cycle date loop    
            print(SRC+ ": Execution completed for " + str(dt1))
            print("-------------------------------------------")
            i += 1 
        
        self.Email()             
        QtGui.QMessageBox.about(self, 'Sucess!',"Execution completed")
        


      
            
    def GL_Recon(self,SRC,dt1,YM1,Data_BAL,Data_INT,DATA_PAR,GL_SUMM)            :
        
#            print(Data_BAL)            
            print(SRC+" : GL Recon is in Progress")
            
            SRCINT=SRC+''+'INT'            
            dest_loc=self.lineEdit_2.text()
            
            #to match the file name with others following code is used. 
            if SRC=='DF550129and550710':
                SRC1='DepoFee550129and550710'
                dest_file=dest_loc+SRC1+'_'+''+YM1+'.xlsx'  
            else:    
                dest_file=dest_loc+SRC+'_'+''+YM1+'.xlsx'  
            
            #GL SUMM is different for the following source system hence we need a different calculation
            if SRC=='DF550129and550710':
               sql_gl="SELECT YEAR_MONTH,        COMPANY_NUM,        550710550129 AS GL_ACCT_NUM,        SUM (MTD_GL_VAL) AS MTD_GL_VAL   FROM (SELECT A.YEAR_MONTH,                A.COMPANY_NUM,                TO_NUMBER (A.GL_ACCT_NUM) AS GL_ACCT_NUM,                SUM (A.MTD_GL_VAL) AS MTD_GL_VAL           FROM GLM.FACT_GL_MONTHLY_SUMMARY A          WHERE     A.YEAR_MONTH IN (SELECT YEAR_MONTH2                               FROM edr.period                              WHERE     year_month2 <= :dt1_year_month                                    AND year_NM =                                           SUBSTR ( :dt1_year_month, 1, 4)                             GROUP BY year_month2)                AND A.GL_MEASURE_KEY = 100                AND A.SCENARIO_GROUP = 'ACTUAL_GL'                AND GL_ACCT_NUM IN ('550710', '550129')                AND A.COMPANY_NUM = 2         GROUP BY A.YEAR_MONTH, A.GL_ACCT_NUM, A.COMPANY_NUM) GROUP BY YEAR_MONTH, COMPANY_NUM, 550710550129"
               sql_gl=sql_gl.replace(':dt1_year_month',YM1)  
               GL_SUMM = pd.io.sql.read_sql(sql_gl, self.db)  
               cx_Oracle.connect.close
            
            

            """ Step3: Writing all Output to file"""         
            
            print("STEP3: Writing " + SRC + " all Elements data into Excel")
            
            writer=pd.ExcelWriter(dest_file, engine='xlsxwriter' )             
            
            #Following Code is used for larger data e.g. ALS, IM, ST
            
            LARGE=('ALS', 'ST', 'IM','DepoFee550139','DepoFee550141', 'DepoFee550557', 'DepoFee550569', 'DepoFee552872', 'DepoFee552877','DF550129and550710')
            
            if SRC in LARGE:
                
#            if (SRC =='ALS' or SRC=='ST' or SRC=='IM' or SRC=='DepoFee550139'):
#                print(SRC)
#                print('Im here')
                
                 
              
                
                if SRC=='ALS':
                    sql="SELECT distinct                 A.SRC_SYS_CD,                 A.YEAR_MONTH,                 C.CUST_NUM,                 C.GU_FULL_NAME AS CUSTOMER_NAME,                 A.SRC_INSTR_ID,                 A.EFFECTIVE_START_DT,                 A.CURR_MATURITY_DT,                 A.ORIGINAL_BAL,                 A.LAST_PAYMENT_DT,                 D.STATUS_DESC AS LOAN_STATUS,                 D.ACCRUAL_STATUS_DESC AS LOAN_ACCRUAL_STATUS,                 E.INSTR_INT_TYPE_CD_DESC,                 A.ACCRUAL_METHOD,                 A.ORIGINAL_INTEREST_RATE,                 A.CURR_INT_RATE,                 A.PREV_PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_ENDING_BAL,                 A.PRINCIPAL_UNPAID_BAL,                 A.INTEREST_INCOME,                 A.ACCR_INTEREST_RECEIVABLE,                 A.LAST_PAYMENT_AMT,                 E.INSTR_TYPE_CD_DESC,                 A.PRINCIPAL_BAL_GL_ACCT_NUM,                 F.GL_INTEREST_INCOME,                 F.GL_ACCR_INTEREST_RECEIVABLE,                 E.INSTR_PROD_CD_DESC,                 A.GL_NA_CD,                 A.TOTAL_PAST_DUE_DAYS            FROM EDR.FACT_LOAN_INSTR_MONTHLY A                 LEFT JOIN EDR.CUST_INSTR_RELATION B                    ON     A.SRC_INSTR_ID = B.ACCT_NUM_DERIVED                       AND B.SRC_SYS_CD = 'RM'                       AND A.SRC_SYS_CD = B.APPLCN_CD_DERIVED                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                              B.BEGIN_EFF_DTTM                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                              NVL (B.END_EFF_DTTM, '31-DEC-9999')                       AND A.RELATIONSHIP_MANAGER_NUM = B.CUST_NUM                       AND B.DELETE_INDICATOR = 'N'                 LEFT JOIN EDR.DIM_CUSTOMER C                    ON     B.CUST_NUM = C.CUST_NUM                       AND C.SRC_SYS_CD = 'RM'                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >=                              C.BEGIN_EFF_DTTM                       AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                              NVL (C.END_EFF_DTTM, '31-DEC-9999')                       AND C.DELETE_INDICATOR = 'N'                       AND A.RELATIONSHIP_MANAGER_NUM = C.CUST_NUM                 LEFT OUTER JOIN EDR.DIM_STATUS D                    ON     A.DIM_STATUS_KEY = D.DIM_STATUS_KEY                       AND A.SRC_SYS_CD = D.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY E                    ON     A.DIM_PRODUCT_KEY = E.DIM_PRODUCT_KEY                       AND A.SRC_SYS_CD = E.SRC_SYS_CD                 LEFT OUTER JOIN EDR.DIM_ALS_GL_MAP F                    ON     A.GL_MAP_KEY = F.DIM_ALS_GL_MAP_KEY                       AND A.PRINCIPAL_BAL_GL_ACCT_NUM = F.GL_PRINCIPAL                       AND A.GL_INVESTOR_CD = F.GL_INVESTOR_CD                       AND A.GL_NA_CD = F.GL_NA_CD           WHERE     A.YEAR_MONTH = :dt1_year_month                 AND A.SRC_SYS_CD = 'ALS'                 AND E.CURRENT_FLG = 'Y'                 AND E.LEVEL_NUM IS NULL                 AND D.CURRENT_FLG = 'Y'                 AND F.CURRENT_FLG = 'Y'                 AND (A.INTEREST_INCOME <> 0 OR A.PRINCIPAL_UNPAID_BAL <> 0)  "
                elif SRC=='ST':
                    sql="SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        C.COMPANY,        A.SRC_INSTR_ID,        F.GU_FULL_NAME AS CUSTOMER_NAME,        A.INTEREST_RATE,        A.INTEREST_EXPENSE,        B.GL_INTEREST_EXPENSE,        A.PRINCIPAL_ENDING_BAL,        B.GL_PRINCIPAL AS GL_ACCT_NUM,        A.MONTH_END_CAL_AVG_PRIN_END_BAL AS AVERAGE_BAL,        D.INSTR_TYPE_CD AS SRC_ACCT_TYPE,        D.INSTR_TYPE_CD_DESC AS PRODUCT_DESCRIPTION,        E.FIXED_VARIABLE_RATE_IND   FROM EDR.FACT_DEPOSITS_MONTHLY A        LEFT OUTER JOIN EDR.DIM_ST_GL_MAP B           ON A.DIM_GL_MAP_KEY = B.DIM_ST_GL_MAP_KEY AND B.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_ORG_HIERARCHY C           ON A.DIM_ORG_KEY = C.DIM_ORG_KEY AND C.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D           ON     A.SRC_SYS_CD = D.SRC_SYS_CD              AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY              AND D.HIERARCHY_TYPE = 'RBG_DDR'              AND D.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_DEPOSIT_PROFILE E           ON     A.DIM_INSTRUMENT_KEY = E.DIM_INSTRUMENT_KEY              AND E.CURRENT_FLG = 'Y'        LEFT OUTER JOIN EDR.DIM_CUSTOMER F           ON     A.SRC_SYS_CD = F.SRC_SYS_CD              AND A.SRC_INSTR_ID =                     LTRIM (SUBSTR (F.CUST_NUM, 1, LENGTH (F.CUST_NUM) - 2),                            '0')              AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') >= F.BEGIN_EFF_DTTM              AND TO_DATE (A.CYCLE_DT_NUM, 'YYYYMMDD') <                     NVL (F.END_EFF_DTTM, '31-DEC-9999')              AND F.DELETE_INDICATOR = 'N'              AND SUBSTR (F.CUST_NUM, LENGTH (F.CUST_NUM), 1) = '1'  WHERE     A.SRC_SYS_CD = 'ST'        AND A.YEAR_MONTH = :dt1_year_month        AND C.COMPANY = '0002'        AND (A.PRINCIPAL_ENDING_BAL <> 0 OR A.INTEREST_EXPENSE <> 0)"
                elif SRC=='IM':
#                    sql_dummy="select src_sys_cd, year_month, DIM_INSTRUMENT_KEY  FROM EDR.FACT_DEPOSITS_MONTHLY A where DIM_INSTRUMENT_KEY ='1338605131'"
                    sql="SELECT A.SRC_SYS_CD,        A.YEAR_MONTH,        C.COMPANY,        A.SRC_INSTR_ID,        F.GU_FULL_NAME AS CUSTOMER_NAME,        A.INTEREST_RATE,        A.INTEREST_EXPENSE,        B.IM_GL_INTEREST_EXPENSE,        A.PRINCIPAL_ENDING_BAL,        CASE WHEN A.PRINCIPAL_ENDING_BAL >= 0             THEN B.IM_GL_PRINCIPAL             ELSE B.IM_GL_OVERDRAFT        END AS GL_ACCT_NUM,        A.MONTH_END_CAL_AVG_PRIN_END_BAL AS AVERAGE_BAL,        D.INSTR_TYPE_CD AS SRC_ACCT_TYPE,        D.INSTR_TYPE_CD_DESC AS PRODUCT_DESCRIPTION,        E.FIXED_VARIABLE_RATE_IND   FROM EDR.FACT_DEPOSITS_MONTHLY A   LEFT OUTER JOIN EDR.DIM_IM_GL_MAP B     ON A.DIM_GL_MAP_KEY = B.DIM_IM_GL_MAP_KEY    AND B.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_ORG_HIERARCHY C     ON A.DIM_ORG_KEY = C.DIM_ORG_KEY    AND C.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY D     ON A.SRC_SYS_CD = D.SRC_SYS_CD    AND A.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY    AND D.HIERARCHY_TYPE = 'RBG_DDR'    AND D.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_DEPOSIT_PROFILE E     ON A.DIM_INSTRUMENT_KEY = E.DIM_INSTRUMENT_KEY    AND E.CURRENT_FLG = 'Y'   LEFT OUTER JOIN EDR.DIM_CUSTOMER F     ON A.SRC_SYS_CD = F.SRC_SYS_CD    AND A.SRC_INSTR_ID = LTRIM(SUBSTR(F.CUST_NUM,1,LENGTH(F.CUST_NUM)-2),'0')    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= F.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(F.END_EFF_DTTM, '31-DEC-9999')    AND F.DELETE_INDICATOR = 'N'    AND SUBSTR(F.CUST_NUM,LENGTH(F.CUST_NUM),1) = '1'  WHERE A.SRC_SYS_CD = 'IM'    AND A.YEAR_MONTH = :dt1_year_month    AND C.COMPANY = '0002'    AND (A.PRINCIPAL_ENDING_BAL <> 0 OR A.INTEREST_EXPENSE <> 0) "
                elif SRC=='DepoFee550139':      
                    sql="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,        C.SRC_INSTR_ID,        D.INSTR_TYPE_CD,        B.TRANSACTION_CD,        550139 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num       AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('3005', '3006')     AND D.INSTR_TYPE_CD IN ('802')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     C.SRC_INSTR_ID,     D.INSTR_TYPE_CD,     B.TRANSACTION_CD,     550139 order by 1"
                elif SRC=='DepoFee550141':  
                     sql="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,        C.SRC_INSTR_ID,        D.INSTR_TYPE_CD,        B.TRANSACTION_CD,        550141 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */     X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('3007', '3008')     AND D.INSTR_TYPE_CD IN ('802')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     C.SRC_INSTR_ID,     D.INSTR_TYPE_CD,     B.TRANSACTION_CD,     550141 order by 1"                  
                elif SRC=='DepoFee550557':  
                      sql="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,        C.SRC_INSTR_ID,        D.INSTR_TYPE_CD,        B.TRANSACTION_CD,        550557 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */     X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('3007', '3008')     AND D.INSTR_TYPE_CD IN ('890')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     C.SRC_INSTR_ID,     D.INSTR_TYPE_CD,     B.TRANSACTION_CD,     550557 order by 1"
                elif SRC=='DepoFee550569':  
                     sql="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,        C.SRC_INSTR_ID,        D.INSTR_TYPE_CD,        B.TRANSACTION_CD,        550569 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('3005', '3006')     AND D.INSTR_TYPE_CD IN ('890')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     C.SRC_INSTR_ID,     D.INSTR_TYPE_CD,     B.TRANSACTION_CD,     550569 order by 1"                
                elif SRC=='DepoFee552872':  
                     sql="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,        C.SRC_INSTR_ID,        D.INSTR_TYPE_CD,        B.TRANSACTION_CD,        552872 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */     X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN ('0096')     AND D.INSTR_TYPE_CD IN ('087')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     C.SRC_INSTR_ID,     D.INSTR_TYPE_CD,     B.TRANSACTION_CD,     552872 order by 1"                     
                elif SRC=='DepoFee552877':      
                     sql="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,        C.SRC_INSTR_ID,        D.INSTR_TYPE_CD,        B.TRANSACTION_CD,        552877 AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   FROM EDR.FACT_DEPOSITS_MONTHLY C   LEFT JOIN   (SELECT /*+ parallel 31 */    X.SRC_SYS_CD,X.TRAN_AMT,X.DIM_INSTRUMENT_KEY,X.DIM_TRANSACTION_KEY,     SUBSTR(X.CYCLE_DT_NUM,1,6) AS YEAR_MONTH FROM EDR.FACT_TRANSACTIONS_DETAIL X         WHERE X.CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD')) AND :dt1_cycle_dt_num         AND X.SRC_SYS_CD = 'IM' AND X.TRAN_AMT <> 0         ) A        ON A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY           AND C.YEAR_MONTH = A.YEAR_MONTH           AND C.SRC_SYS_CD = A.SRC_SYS_CD        LEFT JOIN  EDR.DIM_TRANSACTION B        ON A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY        LEFT JOIN EDR.DIM_PRODUCT D        ON C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY WHERE      A.SRC_SYS_CD = 'IM'     AND B.TRANSACTION_CD IN  ('0096')     AND D.INSTR_TYPE_CD IN ('010', '015', '016', '017', '019', '020', '053', '060', '061', '142')     GROUP BY C.YEAR_MONTH,     A.SRC_SYS_CD,     C.SRC_INSTR_ID,     D.INSTR_TYPE_CD,     B.TRANSACTION_CD,     552877 order by 1"                     
                elif SRC=='DF550129and550710':
                     sql="SELECT C.YEAR_MONTH,        A.SRC_SYS_CD,        C.SRC_INSTR_ID,        D.INSTR_TYPE_CD,        B.TRANSACTION_CD,        '550129 AND 550710' AS GL_ACCT_NUM,        SUM (A.TRAN_AMT) AS TOTAL_ENDING_BAL   From      (select /*+ parallel 31 */         SRC_SYS_CD,TRAN_AMT, DIM_TRANSACTION_KEY,A.DIM_INSTRUMENT_KEY,         SUBSTR(CYCLE_DT_NUM,1,6) AS YEAR_MONTH       FROM EDR.FACT_TRANSACTIONS_DETAIL A       WHERE             CYCLE_DT_NUM BETWEEN TO_NUMBER(TO_CHAR(TRUNC(TO_DATE(:dt1_cycle_dt_num,'YYYYMMDD'),'Y'),'YYYYMMDD'))                                     AND :dt1_cycle_dt_num              AND SRC_SYS_CD = 'IM'                 AND TRAN_AMT <> 0        ) A ,        EDR.DIM_TRANSACTION B,        EDR.FACT_DEPOSITS_MONTHLY C,        EDR.DIM_PRODUCT D  WHERE  A.DIM_TRANSACTION_KEY = B.DIM_TRANSACTION_KEY              AND B.TRANSACTION_CD IN ('0068')              AND A.DIM_INSTRUMENT_KEY = C.DIM_INSTRUMENT_KEY        AND C.YEAR_MONTH = A.YEAR_MONTH        AND C.SRC_SYS_CD = A.SRC_SYS_CD        AND C.DIM_PRODUCT_KEY = D.DIM_PRODUCT_KEY GROUP BY C.YEAR_MONTH,          A.SRC_SYS_CD,          C.SRC_INSTR_ID,          D.INSTR_TYPE_CD,          B.TRANSACTION_CD,          '550129 AND 550710' Order by 1 "
                else:
                    pass

               
                sql=sql.replace(':dt1_year_month',YM1)
                sql=sql.replace(':dt1_cycle_dt_num',dt1)
                
#                connection = cx_Oracle.connect('abdulr06', 'Sep@2017', 'PEDW')
#                cursor = cx_Oracle.Cursor(connection)
                try:             
                  self.cursor.execute(sql)                   
                  
                except cx_Oracle.DatabaseError as e:   
                 self.Email_Error(str(e),SRC)              
                 writer.save()  
                 return
                

                
#                dest_file=dest_loc+SRC+'_'+''+YM1+'.xlsx'  
                writer = pd.ExcelWriter(dest_file, engine='xlsxwriter')
                
                workbook  = writer.book
                worksheet1=workbook.add_worksheet(SRC+''+' All Elements')
                
                cell_format = workbook.add_format({'bold': True})
                cell_format_date = workbook.add_format({'num_format': '[$-409]m/d/yy h:mm AM/PM;@'})    
                
                #writing into Excel
                
                if SRC=='IM':
                    worksheet2=workbook.add_worksheet(SRC+''+' All Elements cont..')
                
                    for r, row in enumerate(self.cursor):
                        for c, col in enumerate(row): 
                            if r<=1048575:
                                worksheet1.write(r+1,c, row[c]) 
                            else:
                                worksheet2.write((r+1)-1048576,c, row[c]) 
                else:                
                    for r, row in enumerate(self.cursor):
                        for c, col in enumerate(row):
                            worksheet1.write(r+1,c, row[c]) 
                
                
                
                
                 
    
                #Getting columns headers   & writing      
                col_names = []
                for i in range(0, len(self.cursor.description)):
                    col_names.append(self.cursor.description[i][0])   
                    if SRC=='IM':
                        worksheet1.write(0,i,col_names[i],cell_format)  
                        worksheet2.write(0,i,col_names[i],cell_format)  
                    else:
                        worksheet1.write(0,i,col_names[i],cell_format)  
               
                
               #Formatting Date Columns  
                if SRC=='ALS':                   
                    worksheet1.set_column('F:G', 18, cell_format_date)
                    worksheet1.set_column('I:I', 18, cell_format_date)
                else: 
                   pass
                   
               #Column Width 
                if SRC=='DF550129and550710':
                    worksheet1.set_column('A:G',28)
                else: 
                   pass 
             
#                self.cursor.close()
#                self.oracle.disconnect()
#                connection.close()  
                
                
                
                
                
            
            
            """Step2: Following Operations are common for rest of the source sytems"""
           
            #Writing the All Elements value 
            Data_BAL_Orginal=Data_BAL
            Data_INT_Orginal=Data_INT
            
            #Retaining orginal colum names 
            if SRC=='AFS':
                Data_BAL_Orginal=Data_BAL_Orginal.rename(columns={'TOTAL_ENDING_BAL':'PRINCIPAL_ENDING_BAL'})
            elif (SRC=='ML' or SRC=='MO'):
                Data_BAL_Orginal=Data_BAL_Orginal.rename(columns={'TOTAL_ENDING_BAL':'PRINCIPAL_UNPAID_BAL'})
                Data_INT_Orginal=Data_INT_Orginal.rename(columns={'TOTAL_INT_INCOME':'INTEREST_INCOME'})
            elif SRC=='MS':
                Data_INT_Orginal=Data_INT_Orginal.rename(columns={'TOTAL_INT_INCOME':'INTEREST_INCOME'})
            elif (SRC=='LV'):
                Data_BAL_Orginal=Data_BAL_Orginal.rename(columns={'TOTAL_ENDING_BAL':'BANK_BAL'})       
            elif (SRC=='LoanFee-533975'):  
                Data_BAL_Orginal=Data_BAL_Orginal.rename(columns={'TOTAL_ENDING_BAL':'AF_SYNDICATION_FEE_2','GL_ACCT_NUM':'GL_SYNDICATION_FEE_2'})  
            elif (SRC=='LoanFee-534070'):
                Data_BAL_Orginal=Data_BAL_Orginal.rename(columns={'TOTAL_ENDING_BAL':'AF_SYNDICATION_FEE_3','GL_ACCT_NUM':'GL_SYNDICATION_FEE_3'})
            elif (SRC=='LoanFee-533901')    :
                Data_BAL_Orginal=Data_BAL_Orginal.rename(columns={'TOTAL_ENDING_BAL':'GL_MISC_FEE_INCOME','GL_ACCT_NUM':'MISC_FEE_INCOME'})                 
            elif SRC=='LG':
                Data_BAL_Orginal=Data_BAL_Orginal.rename(columns={'PRINCIPAL_ENDING_BAL':'ENDING_BAL','GL_INTEREST_INCOME':'GL_INTEREST_EXPENSE','TOTAL_INT_INCOME':'INTEREST_EXPENSE'})            
            else:
                pass
            
            #Wirintg All Elements records expcet the following systems. ALS/ST/IM & Deposit fee extraction are huge. 
            if SRC not in LARGE:
#            if SRC!='ALS' and SRC!='ST' and SRC!='IM' and SRC!='DF550139':
#                print("BINGOOOOOOOOO")
                Data_BAL_Orginal.to_excel(writer, sheet_name= SRC+''+' All Elements', index=False) 
            else:
                pass
            
            #Following condition needed as becuase we are derving Intrest income from balance query.
            if (SRC != 'IMOD' and SRC != 'LV'and SRC != 'ST' and SRC!='IM' and SRC!='LG' and SRC!='ALS' and SRC!='AFS' and SRC!='LoanFee-533975' and SRC !='LoanFee-534070' and SRC!='TSYS' and SRC!='LoanFee-533901' and SRC!='DepoFee550139' and SRC!='DepoFee550141' and SRC!='DepoFee550557' and SRC!='DepoFee550569' and SRC!='DepoFee552872' and SRC!='DepoFee552877' and SRC!='DF550129and550710'):
                Data_INT_Orginal.to_excel(writer, sheet_name= SRCINT+''+' All Elements', index=False) 
            elif SRC=='AFS':
                Data_INT_Orginal.to_excel(writer, sheet_name= SRCINT+''+' All Elements', index=False)
                DATA_PAR.to_excel(writer, sheet_name= 'AFS Partsold'+' All Elements', index=False) 
                Data_INT_V2=DATA_PAR.rename(columns={'PARTSOLD_TO_GL_INT_EXP_PART':'GL_AUTO_LEASE_INCOME','PARTSOLD_TO_INTEREST_INCOME':'INCOME_EARNED'})
            else:
                pass
            
          
            
           #Writing customer details & deriving int. income from Data_BAL query
            if SRC=='MS':            
                MS_CUS=Data_BAL[['SRC_SYS_CD','SRC_INSTR_ID','CUST_NUM','CUSTOMER_NAME']] 
                MS_CUS.to_excel(writer, sheet_name='Customer Info', index=False)
                
            elif SRC=='FD':
                SQL="WITH FD_INSTR AS ( SELECT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.COMPANY_NUM,        A.SRC_INSTR_ID,        A.INSTR_OPEN_DT,         B.STATUS_DESC AS LOAN_STATUS,        B.ACCRUAL_STATUS_DESC AS LOAN_ACCRUAL_STATUS,        C.INSTR_INT_TYPE_CD,        A.PRI_CASH_INT_RT,        A.PRI_PUR_INT_RT,        A.PREV_PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_ENDING_BAL,        A.PRINCIPAL_UNPAID_BAL,        A.LAST_BILLING_CYCLE_DT AS LAST_STMT_DT,        E.PRINCIPAL_ENDING_BAL AS STMT_PRINCIPAL_ENDING_BAL,        E.PRINCIPAL_UNPAID_BAL AS STMT_PRINCIPAL_UNPAID_BAL,        E.PREV_PRINCIPAL_ENDING_BAL AS STMT_PREV_PRINCIPAL_ENDING_BAL,        A.MTD_INTEREST_INCOME,        A.ACCRUED_INTEREST_RECEIVABLE,        A.LAST_PAYMENT_AMT,        C.INSTR_TYPE_CD_DESC,        A.PRINCIPAL_BAL_GL_ACCT_NUM,        D.GL_INTEREST_INCOME,        D.GL_ACCR_INTEREST_RECEIVABLE,        C.INSTR_PROD_CD_DESC,        A.CUSTOMER_GROUP_DESC,        A.ACTUAL_PAST_DUE_DAYS   FROM EDR.FACT_CREDITCARD_INSTR_MONTHLY A   LEFT OUTER JOIN EDR.DIM_STATUS B     ON A.DIM_STATUS_KEY = B.DIM_STATUS_KEY    AND A.SRC_SYS_CD = B.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_PRODUCT_HIERARCHY C     ON A.DIM_PRODUCT_KEY = C.DIM_PRODUCT_KEY    AND A.SRC_SYS_CD = C.SRC_SYS_CD   LEFT OUTER JOIN EDR.DIM_FD_GL_MAP D     ON A.DIM_GL_MAP_KEY = D.DIM_FD_GL_MAP_KEY    AND A.PRINCIPAL_BAL_GL_ACCT_NUM = D.GL_PRINCIPAL_BALANCE   LEFT OUTER JOIN EDR.FACT_CREDITCARD_INSTR_CYCLE E     ON A.SRC_INSTR_ID = E.SRC_INSTR_ID    AND A.LAST_BILLING_CYCLE_DT = E.LAST_BILLING_CYCLE_DT  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'FD'    AND A.PRINCIPAL_UNPAID_BAL <> 0    AND B.CURRENT_FLG = 'Y'    AND C.LEVEL_NUM IS NULL    AND C.CURRENT_FLG = 'Y'    AND D.CURRENT_FLG = 'Y'  ORDER BY A.SRC_INSTR_ID ), FD_CUST AS ( SELECT B.APPLCN_CD_DERIVED AS SRC_SYS_CD,        B.BEGIN_EFF_DTTM,        B.END_EFF_DTTM,        B.ACCT_NUM_DERIVED AS SRC_INSTR_ID,        C.CUST_NUM,        C.GU_FULL_NAME AS CUSTOMER_NAME   FROM EDR.CUST_INSTR_RELATION B   LEFT OUTER JOIN EDR.DIM_CUSTOMER C    ON B.CUST_NUM = C.CUST_NUM    AND C.SRC_SYS_CD = 'RM'    AND TO_DATE(':dt1_cycle_dt_num','YYYYMMDD') >= C.BEGIN_EFF_DTTM    AND TO_DATE(':dt1_cycle_dt_num','YYYYMMDD') < NVL(C.END_EFF_DTTM, '31-DEC-9999')    AND C.DELETE_INDICATOR = 'N' WHERE B.APPLCN_CD_DERIVED = 'FD'    AND B.SRC_SYS_CD = 'RM'    AND TO_DATE(':dt1_cycle_dt_num','YYYYMMDD') >= B.BEGIN_EFF_DTTM    AND TO_DATE(':dt1_cycle_dt_num','YYYYMMDD') < NVL(B.END_EFF_DTTM, '31-DEC-9999')    AND B.DELETE_INDICATOR = 'N'  ORDER BY B.ACCT_NUM_DERIVED,           C.CUST_NUM,           C.GU_FULL_NAME)          SELECT DISTINCT        A.YEAR_MONTH,        A.SRC_SYS_CD,        B.CUST_NUM,        B.CUSTOMER_NAME,        A.SRC_INSTR_ID   FROM FD_INSTR A   LEFT OUTER JOIN FD_CUST B     ON A.SRC_INSTR_ID = B.SRC_INSTR_ID  WHERE B.CUST_NUM IS NOT NULL  ORDER BY A.SRC_INSTR_ID,           B.CUST_NUM"
                df_FDCUST=self.execution(dt1,YM1,SRC,SQL)
                df_FDCUST.to_excel(writer, sheet_name='Customer Info', index=False)

            elif SRC=='IMOD':                
                Data_INT=Data_BAL[['YEAR_MONTH','GL_INTEREST_INCOME','TOTAL_INT_INCOME']] 
                SQL="SELECT DISTINCT A.YEAR_MONTH,        A.SRC_SYS_CD,        A.SRC_INSTR_ID,        C.GU_FULL_NAME AS CUSTOMER_NAME   FROM EDR.FACT_OVERDRAFT_MONTHLY A   LEFT JOIN EDR.DIM_CUSTOMER C     ON C.SRC_SYS_CD = 'IM'    AND A.SRC_INSTR_ID = LTRIM(SUBSTR(C.CUST_NUM,1,LENGTH(C.CUST_NUM)-2),'0')    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') >= C.BEGIN_EFF_DTTM    AND TO_DATE(A.CYCLE_DT_NUM,'YYYYMMDD') < NVL(C.END_EFF_DTTM, '31-DEC-9999')    AND C.DELETE_INDICATOR = 'N'  WHERE A.YEAR_MONTH = :dt1_year_month    AND A.SRC_SYS_CD = 'IMOD'    AND (A.PRINCIPAL_ENDING_BAL <> 0 OR A.INTEREST_BAL <> 0)  ORDER BY A.SRC_INSTR_ID "
                df_IMODCUST=self.execution(dt1,YM1,SRC,SQL)
                df_IMODCUST.to_excel(writer, sheet_name='Customer Info', index=False)
            elif SRC=='LV':
                Data_INT=Data_BAL[['YEAR_MONTH','GL_UNEARNED_INCOME','RECEIVABLES_UNEARNED']] 
                Data_INT=Data_INT.rename(columns={'GL_UNEARNED_INCOME':'GL_INTEREST_INCOME','RECEIVABLES_UNEARNED':'TOTAL_INT_INCOME'})
                Data_INT_V2=Data_BAL[['YEAR_MONTH','GL_AUTO_LEASE_INCOME','INCOME_EARNED']]
            elif SRC=='ALS' :
                Data_INT=Data_BAL[['YEAR_MONTH','GL_INTEREST_INCOME','TOTAL_INT_INCOME']] 
            elif (SRC=='LoanFee-533975' or SRC=='LoanFee-534070' or SRC=='LoanFee-533901' or SRC=='DepoFee550139' or SRC=='DepoFee550141' or SRC=='DepoFee550557' or SRC=='DepoFee550569' or SRC=='DepoFee552872' or SRC=='DepoFee552877' or SRC=='DF550129and550710'):
                Data_INT=pd.DataFrame()
                Data_INT_GL=Data_INT
            elif SRC=='ST' or SRC=='IM' or SRC=='LG':
                Data_INT=Data_BAL[['YEAR_MONTH','GL_INTEREST_INCOME','TOTAL_INT_INCOME']]             
                
            else:
                pass

            
            print(SRC+ " all elements data saved")
            
            
            """Step2: Renaming the columns"""
            Data_BAL=Data_BAL.rename(columns={'PRINCIPAL_ENDING_BAL':'TOTAL_ENDING_BAL','GL_PRINCIPAL':'GL_ACCT_NUM','PRINCIPAL_BAL_GL_ACCT_NUM':'GL_ACCT_NUM'}) #Bal & Amount
#            print(Data_BAL)
            GL_SUMM_INT=GL_SUMM.rename(columns={'GL_ACCT_NUM':'GL_INTEREST_INCOME'})
#            print(GL_SUMM_INT.head(2))
            print("Renamed sucess")
            
            
            
#            print(Data_BAL)
            """Step3: Aggregate Balance & Int.Income by GL"""
              
            Data_BAL_GL=Data_BAL.groupby(['YEAR_MONTH','GL_ACCT_NUM']).agg({'TOTAL_ENDING_BAL':np.sum})
            Data_BAL_GL=Data_BAL_GL.reset_index()
            
          
            #For ALL LOANS Fee & Deposits Fee the following condition is not applicable 
            if (SRC!='LoanFee-533975' and SRC!='LoanFee-534070' and SRC!='TSYS' and SRC!='LoanFee-533901' and SRC!='DepoFee550139' and SRC!='DepoFee550141' and SRC!='DepoFee550557' and SRC!='DepoFee550569' and SRC!='DepoFee552872' and SRC!='DepoFee552877' and SRC!='DF550129and550710'):
                Data_INT_GL=Data_INT.groupby(['YEAR_MONTH','GL_INTEREST_INCOME']).agg({'TOTAL_INT_INCOME':np.sum})
                Data_INT_GL=Data_INT_GL.reset_index() 
            elif SRC=='TSYS':
                Data_INT_GL=Data_INT.groupby(['GL_INTEREST_INCOME']).agg({'TOTAL_INT_INCOME':np.sum})
                Data_INT_GL=Data_INT_GL.reset_index()    
            else:
                pass
                
            #Third caluclations need only for the following two source system and for rest we are setting null. 
            if (SRC=='LV' or SRC=='AFS') :
                Data_INT_V2_GL=Data_INT_V2.groupby(['YEAR_MONTH','GL_AUTO_LEASE_INCOME']).agg({'INCOME_EARNED':np.sum})
                Data_INT_V2_GL=Data_INT_V2_GL.reset_index()  
            else:
                Data_INT_V2_GL=pd.DataFrame()        
        
 
            print("Group by performed")    
            
#            print(Data_BAL_GL)
            
            #filtering the company name  & assigning GL SUMM BALANCE & Int. Income
            
            if (SRC=='AFS' or SRC=='DF' or SRC=='MS' or SRC=='LoanFee-533901' or SRC=='DF550129and550710' or SRC=='ST' or SRC=='IM' or SRC=='LG'):
                GL_SUMM_BAL=GL_SUMM[GL_SUMM['COMPANY_NUM'] == 2]
                GL_SUMM_INT=GL_SUMM_INT[GL_SUMM_INT['COMPANY_NUM'] == 2]
            
            elif SRC=='LS':    
                GL_SUMM_BAL=GL_SUMM.groupby(['YEAR_MONTH','GL_ACCT_NUM']).agg({'MTD_GL_VAL':np.sum})
                GL_SUMM_BAL=GL_SUMM_BAL.reset_index()
                GL_SUMM_INT=GL_SUMM_INT[GL_SUMM_INT['COMPANY_NUM'] == 2]
            elif SRC=='ALS':
                GL_SUMM_BAL=GL_SUMM.groupby(['YEAR_MONTH','GL_ACCT_NUM']).agg({'MTD_GL_VAL':np.sum})
                GL_SUMM_BAL=GL_SUMM_BAL.reset_index()  
                GL_SUMM_INT=GL_SUMM_INT.groupby(['YEAR_MONTH','GL_INTEREST_INCOME']).agg({'MTD_GL_VAL':np.sum})
                GL_SUMM_INT=GL_SUMM_INT.reset_index()  
            else:
                GL_SUMM_BAL=GL_SUMM            

#            print(GL_SUMM_BAL)
#            print('--------------')
#            print('Data Bal Value\n', Data_BAL_GL)
#            
            """Step4: Join with GL Summary Table """
            
            if not Data_BAL_GL.empty:   
#                print("Hey Im inside not empty loop")
                COMP_BAL=pd.merge(Data_BAL_GL, GL_SUMM_BAL, on=['GL_ACCT_NUM','YEAR_MONTH'], how='outer', suffixes=['','_GL_SUMMARY'], indicator=True)
                
                #if it is AFS then pick up both matching and unmatching record from DATA BAL
                if SRC=='AFS' or SRC=='IM':
                    COMP_BAL1=COMP_BAL.loc[COMP_BAL._merge.eq('left_only')][['YEAR_MONTH','GL_ACCT_NUM','TOTAL_ENDING_BAL','MTD_GL_VAL']]
                    COMP_BAL1.fillna(0, inplace=True)
#                    print(COMP_BAL1)
                    COMP_BAL2=COMP_BAL.loc[COMP_BAL._merge.eq('both')][['YEAR_MONTH','GL_ACCT_NUM','TOTAL_ENDING_BAL','MTD_GL_VAL']]
                    COMP_BAL=pd.concat([COMP_BAL1,COMP_BAL2])
#                    print(COMP_BAL)
                else:        
                    COMP_BAL=COMP_BAL.loc[COMP_BAL._merge.eq('both')][['YEAR_MONTH','GL_ACCT_NUM','TOTAL_ENDING_BAL','MTD_GL_VAL']]
                   
#                    print('After join\n',COMP_BAL)   
                COMP_BAL['VARIANCE']=abs(COMP_BAL['TOTAL_ENDING_BAL'])-abs(COMP_BAL['MTD_GL_VAL'])
                COMP_BAL['VARIANCE %']=((COMP_BAL['VARIANCE'])/(COMP_BAL['MTD_GL_VAL']))   
                
                if (SRC=='MO' or SRC=='ML' or SRC=='ALS'):  
                    COMP_BAL=COMP_BAL.rename(columns={'TOTAL_ENDING_BAL':'PRINCIPAL_UNPAID_BAL'})   
                elif SRC=='LoanFee-533975':    
                    COMP_BAL=COMP_BAL.rename(columns={'TOTAL_ENDING_BAL':'AF_SYNDICATION_FEE_2','GL_ACCT_NUM':'GL_SYNDICATION_FEE_2'})   
                elif SRC=='LoanFee-534070':
                    COMP_BAL=COMP_BAL.rename(columns={'TOTAL_ENDING_BAL':'AF_SYNDICATION_FEE_3','GL_ACCT_NUM':'GL_SYNDICATION_FEE_3'})
                elif (SRC=='LoanFee-533901')    :
                    COMP_BAL=COMP_BAL.rename(columns={'TOTAL_ENDING_BAL':'GL_MISC_FEE_INCOME','GL_ACCT_NUM':'MISC_FEE_INCOME'})      
                elif (SRC=='DepoFee550139' or SRC=='DepoFee550141' or SRC=='DepoFee550557' or SRC=='DepoFee550569' or SRC=='DepoFee552872' or SRC=='DepoFee552877' or SRC=='DF550129and550710')    :
                    COMP_BAL=COMP_BAL.rename(columns={'TOTAL_ENDING_BAL':'TRAN_AMT'})    
                else: 
                    COMP_BAL=COMP_BAL.rename(columns={'TOTAL_ENDING_BAL':'PRINCIPAL_ENDING_BAL'}) 
#                print("inside non empty loop")    
#                print(COMP_BAL)     
            if not Data_INT_GL.empty:       
                COMP_INT=pd.merge(Data_INT_GL, GL_SUMM_INT, on=['GL_INTEREST_INCOME'], how='outer', suffixes=['','_GL_SUMMARY'], indicator=True)
                
                if SRC=='IM' or SRC=='LG':
                    COMP_INT1=COMP_INT.loc[COMP_INT._merge.eq('left_only')][['YEAR_MONTH','GL_INTEREST_INCOME','TOTAL_INT_INCOME','MTD_GL_VAL']]    
                    COMP_INT1.fillna(0, inplace=True)
                    COMP_INT2=COMP_INT.loc[COMP_INT._merge.eq('both')][['YEAR_MONTH','GL_INTEREST_INCOME','TOTAL_INT_INCOME','MTD_GL_VAL']]
                    COMP_INT=pd.concat([COMP_INT1,COMP_INT2])
                else:
                    COMP_INT=COMP_INT.loc[COMP_INT._merge.eq('both')][['YEAR_MONTH','GL_INTEREST_INCOME','TOTAL_INT_INCOME','MTD_GL_VAL']]
                    
                COMP_INT['VARIANCE']=abs(COMP_INT['TOTAL_INT_INCOME'])-abs(COMP_INT['MTD_GL_VAL'])
                COMP_INT['VARIANCE %']=((COMP_INT['VARIANCE'])/(COMP_INT['MTD_GL_VAL'])) 
                if SRC=='LV':
                    COMP_INT=COMP_INT.rename(columns={'GL_INTEREST_INCOME':'GL_UNEARNED_INCOME','TOTAL_INT_INCOME':'RECEIVABLES_UNEARNED'})  
                elif SRC=='ST' or SRC=='IM' or SRC=='LG':
                    COMP_INT=COMP_INT.rename(columns={'GL_INTEREST_INCOME':'GL_INTEREST_EXPENSE','TOTAL_INT_INCOME':'INTEREST_EXPENSE'})    
                elif SRC=='MO' or SRC=='MS' or SRC=='ML':
                    COMP_INT=COMP_INT.rename(columns={'TOTAL_INT_INCOME':'INTEREST_INCOME'})     
                       
                else: 
                    pass
           
            if not Data_INT_V2_GL.empty:            
                if  SRC=='LV' :
                    GL_SUMM_INT_V2=GL_SUMM_BAL.rename(columns={'GL_ACCT_NUM':'GL_AUTO_LEASE_INCOME'})                
                    COMP_INT_V2=pd.merge(Data_INT_V2_GL, GL_SUMM_INT_V2, on=['GL_AUTO_LEASE_INCOME'], how='outer', suffixes=['','_GL_SUMMARY'], indicator=True)
                    COMP_INT_V2=COMP_INT_V2.loc[COMP_INT_V2._merge.eq('both')][['YEAR_MONTH','GL_AUTO_LEASE_INCOME','INCOME_EARNED','MTD_GL_VAL']]
                    COMP_INT_V2['VARIANCE']=abs(COMP_INT_V2['INCOME_EARNED'])-abs(COMP_INT_V2['MTD_GL_VAL'])
                    COMP_INT_V2['VARIANCE %']=((COMP_INT_V2['VARIANCE'])/(COMP_INT_V2['MTD_GL_VAL']))  
                elif SRC=='AFS':
                    GL_SUMM_INT_V2=GL_SUMM_BAL.rename(columns={'GL_ACCT_NUM':'GL_AUTO_LEASE_INCOME'})                
                    COMP_INT_V2=pd.merge(Data_INT_V2_GL, GL_SUMM_INT_V2, on=['GL_AUTO_LEASE_INCOME'], how='outer', suffixes=['','_GL_SUMMARY'], indicator=True)                    
                    COMP_INT_V2_A=COMP_INT_V2.loc[COMP_INT_V2._merge.eq('left_only')][['YEAR_MONTH','GL_AUTO_LEASE_INCOME','INCOME_EARNED','MTD_GL_VAL']]
                    COMP_INT_V2_A.fillna(0, inplace=True)
                    COMP_INT_V2_B=COMP_INT_V2.loc[COMP_INT_V2._merge.eq('both')][['YEAR_MONTH','GL_AUTO_LEASE_INCOME','INCOME_EARNED','MTD_GL_VAL']]
                    COMP_INT_V2=pd.concat([COMP_INT_V2_A,COMP_INT_V2_B])
                    COMP_INT_V2['VARIANCE']=abs(COMP_INT_V2['INCOME_EARNED'])-abs(COMP_INT_V2['MTD_GL_VAL'])
                    COMP_INT_V2['VARIANCE %']=((COMP_INT_V2['VARIANCE'])/(COMP_INT_V2['MTD_GL_VAL']))                      
                    COMP_INT_V2=COMP_INT_V2.rename(columns={'GL_AUTO_LEASE_INCOME':'PARTSOLD_TO_GL_INT_EXP_PART','INCOME_EARNED':'PARTSOLD_TO_INTEREST_INCOME'})
                else:
                    pass
                    
                    
                    
            
            #consider modifying this code (when the data frame is empty, the following code will be trigerred)
            if  Data_BAL_GL.empty:   
                data=np.array([['','YEAR_MONTH','GL_ACCT_NUM','TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %'], ['1',YM1,000000,0,0,0,0]])
                COMP_BAL=pd.DataFrame(data=data[1:,1:], index=data[1:,0], columns=data[0,1:])
                COMP_BAL[['TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %']] = COMP_BAL[['TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %']].apply(pd.to_numeric)
            
                    
            if Data_INT_GL.empty: 
                
                data=np.array([['','YEAR_MONTH','GL_ACCT_NUM','TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %'], ['1',YM1,000000,0,0,0,0]])
                COMP_INT=pd.DataFrame(data=data[1:,1:], index=data[1:,0], columns=data[0,1:])
                COMP_INT[['TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %']] = COMP_INT[['TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %']].apply(pd.to_numeric)
                
            if Data_INT_V2_GL.empty:                
                data=np.array([['','YEAR_MONTH','GL_ACCT_NUM','TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %'], ['1',YM1,000000,0,0,0,0]])
                COMP_INT_V2=pd.DataFrame(data=data[1:,1:], index=data[1:,0], columns=data[0,1:])
                COMP_INT_V2[['TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %']] = COMP_INT_V2[['TOTAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','VARIANCE %']].apply(pd.to_numeric)
                
                

            
            
            """Step7: Writing Output to an Excel sheet"""      
            print("Writing GL Recon output to Excel Sheet")
#            print(COMP_BAL)
            if SRC=='LV':
                COMP_BAL.to_excel(writer, sheet_name=SRC+''+' GL-Recon', index=False)
                COMP_INT.to_excel(writer, sheet_name=SRC+''+' GL-Recon', startrow =5, startcol=0,  index=False)
                COMP_INT_V2.to_excel(writer, sheet_name=SRC+''+' GL-Recon', startrow =10, startcol=0,  index=False)
            
            elif SRC=='AFS':   
                COMP_BAL.to_excel(writer, sheet_name=SRC+''+' GL-Recon', index=False)
                COMP_INT.to_excel(writer, sheet_name=SRCINT+''+' GL-Recon', index=False)
                COMP_INT_V2.to_excel(writer, sheet_name='Partsold'+' GL-Recon', index=False)
                
            elif (SRC=='LoanFee-533975' or SRC=='LoanFee-534070' or SRC=='LoanFee-533901' or SRC=='DepoFee550139' or SRC=='DepoFee550141' or SRC=='DepoFee550557' or SRC=='DepoFee550569' or SRC=='DepoFee552872' or SRC=='DepoFee552877'):
                COMP_BAL.to_excel(writer, sheet_name=SRC+''+' GL-Recon', index=False)   
                
            elif SRC=='DF550129and550710'  :
                COMP_BAL['GL_ACCT_NUM']='550129 AND 550710'
                COMP_BAL.to_excel(writer, sheet_name=SRC+''+' GL-Recon', index=False)  
                
        
            elif SRC=='TSYS':                
               
#                SETTLEMENT_POSTED='SETTL_POST'+'_'+str(int(YM1)+1)+'01'
                COMP_INT=COMP_INT.rename(columns={'GL_INTEREST_INCOME':'GL_ACCT_NUM','TOTAL_INT_INCOME':'SETTLEMENT_POSTED'})    
                
                COMP_BI=pd.merge(COMP_BAL,COMP_INT, on=['GL_ACCT_NUM'], how='outer', suffixes=['','_GL_SUMMARY'], indicator=True) 
                COMP_BI=COMP_BI.loc[COMP_BI._merge.eq('both')][['YEAR_MONTH','GL_ACCT_NUM','PRINCIPAL_ENDING_BAL','MTD_GL_VAL','VARIANCE','SETTLEMENT_POSTED']]                
                COMP_BI['NET_VARIANCE']=abs(COMP_BI['VARIANCE'])-abs(COMP_INT['SETTLEMENT_POSTED'])                
                COMP_BI.to_excel(writer, sheet_name=SRC+''+' GL-Recon', index=False)
                
                #Get the sheetname to peform the calulcation and formatting
                workbook =writer.book       
                worksheetTSYS=writer.sheets[SRC+''+' GL-Recon']  
                
         
                
                fmtt = workbook.add_format({'bold':True,'font_color':'blue'})
                fmtt.set_bottom(6)
                worksheetTSYS.write(0,5,'SETTLEMENT POSTED'+'_'+str(int(YM1)+1)+'01',fmtt)
                worksheetTSYS.write(5,0, 'TSYS does not have a direct GL interface.  The accounting entries are posted the next day')
                worksheetTSYS.write(6,0, 'by users based on system settlement reports.  To reconcile the '+  str(int(YM1)-1) +' balances requires')
                worksheetTSYS.write(7,0, 'the inclusion of the settlement entries posted the next processing day.')
                
                fmt1 = workbook.add_format({'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'})
                worksheetTSYS.set_column('A:G',28)
                worksheetTSYS.set_column('C:G',None,fmt1)
                
                #Following condition formatting help to highight the variance
                fmtR = workbook.add_format({'font_color':'red'})
                worksheetTSYS.conditional_format('E2:E2', {'type':     'cell',
                                        'criteria': '>=',
                                        'value':    50,
                                        'format':   fmtR})
                        
            else:  
                COMP_BAL.to_excel(writer, sheet_name=SRC+''+' GL-Recon', index=False)
                COMP_INT.to_excel(writer, sheet_name=SRCINT+''+' GL-Recon', index=False)
                
                
            """Step:8 FORMATTING EXCEL WITHOUT LOADING - Following """
            if SRC in LARGE:
                    
#            if (SRC=='ALS'  or SRC=='ST' or SRC=='IM' ):
                
                print(SRC+''+"- Excel formatting in progress")
#                print(COMP_BAL)
                #Format for Comma value
                format1 = workbook.add_format({'num_format': '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'})                
                format2 = workbook.add_format({'num_format': '0.00%'})  
                fmtR = workbook.add_format({'font_color':'red'})
                 
                DepoFee=['DepoFee550139','DepoFee550141', 'DepoFee550557', 'DepoFee550569', 'DepoFee552872', 'DepoFee552877','DF550129and550710']
                
                #Handeling ALS Column name     
                if SRC=='ALS':
                    COMP_BAL=COMP_BAL.rename(columns={'PRINCIPAL_UNPAID_BAL':'PRINCIPAL_ENDING_BAL'}) 
                elif SRC in DepoFee:
                    COMP_BAL=COMP_BAL.rename(columns={'TRAN_AMT':'PRINCIPAL_ENDING_BAL'}) 
                else:    
                    pass
                     
                #BALANCE
                EDR_TOTAL = pd.Series(COMP_BAL['PRINCIPAL_ENDING_BAL'].sum())   
                MTD_TOTAL = pd.Series(COMP_BAL['MTD_GL_VAL'].sum())   
                VAR_TOTAL=abs(MTD_TOTAL)-abs(EDR_TOTAL)
                VAR_PER=VAR_TOTAL/MTD_TOTAL      
                
                rw=len(COMP_BAL)+3
                
                #Get the sheetname to peform the calulcation and formatting
                worksheetDF=writer.sheets[SRC+''+' GL-Recon']                

                worksheetDF.write(rw,1, 'GRAND TOTAL')
                worksheetDF.write(rw,2, EDR_TOTAL, format1)
                worksheetDF.write(rw,3, MTD_TOTAL, format1)
                worksheetDF.write(rw,4, VAR_TOTAL, format1)
                worksheetDF.write(rw,5, VAR_PER, format2)
                worksheetDF.set_column('C:E',25,format1)
                worksheetDF.set_column('A:B',25)
                worksheetDF.set_column('F:F',25,format2)
             
                #Formatting the variance 
                rang_1='E2:E'+ str(len(COMP_BAL)+1)   
                worksheetDF.conditional_format(rang_1, {'type':     'cell',
                                        'criteria': 'not between',
                                        'minimum':  -50,
                                        'maximum':   50,
                                        'format':   fmtR})
            
            
            
                    
                #Int. Income
                if SRC=='ALS'     :                    
                    EDRINT_TOTAL = pd.Series(COMP_INT['TOTAL_INT_INCOME'].sum())   
                elif SRC=='ST' or SRC=='IM'  :
                    EDRINT_TOTAL = pd.Series(COMP_INT['INTEREST_EXPENSE'].sum())
                else:
                    pass
                
                #Following coniditon need only for ALS, ST & IM. Other source systems doesn have Int. Income calculations
                if SRC=='ALS' or SRC=='ST' or SRC=='IM':
                    MTDINT_TOTAL = pd.Series(COMP_INT['MTD_GL_VAL'].sum())   
                    VARINT_TOTAL=abs(MTDINT_TOTAL)-abs(EDRINT_TOTAL)
                    VARINT_PER=VARINT_TOTAL/MTDINT_TOTAL   
                    
                    rw1=len(COMP_INT)+3
                    worksheetDF=writer.sheets[SRCINT+''+' GL-Recon']
    #                worksheetDF.set_column('C:E',format1)
    #                worksheetDF.set_column('F:F',format2)
                    worksheetDF.write(rw1,1, 'GRAND TOTAL')
                    worksheetDF.write(rw1,2, EDRINT_TOTAL, format1)
                    worksheetDF.write(rw1,3, MTDINT_TOTAL, format1)
                    worksheetDF.write(rw1,4, VARINT_TOTAL, format1)
                    worksheetDF.write(rw1,5, VARINT_PER, format2)
                    worksheetDF.set_column('C:E',25,format1)
                    worksheetDF.set_column('A:B',25)
                    worksheetDF.set_column('F:F',25,format2)
                else:
                   pass
                
                

            else:
                pass
                
            writer.save()  
            print("File saved")
  
          
        
            
    def Excel(self,SRC,YM1)      :
        
        #    """ LOAD THE VALUES FROM EXCEL """
        
          print("Excel Alignment is in progress for "+SRC)        
          
          SRCINT=SRC+''+'INT'         
          dest_loc=self.lineEdit_2.text()
          
          if SRC=='DF550129and550710':
                SRC1='DepoFee550129and550710'
                dest_file=dest_loc+SRC1+'_'+''+YM1+'.xlsx'  
          else:    
                dest_file=dest_loc+SRC+'_'+''+YM1+'.xlsx' 
          
          
          
          #Formatting Excel sheet
          fmtB=Font(color=colors.BLUE)
          fmtR=Font(color=colors.RED)
          
          #assigning sheet names as per the source systems
          if SRC=='AFS':
              
              sheetname = [SRC+''+' GL-Recon', SRCINT+''+' GL-Recon','Partsold'+' GL-Recon']   
              
          elif SRC =='LV':
              sheetname = SRC+''+' GL-Recon'  
              wb=load_workbook(dest_file)
              sheet1=wb.get_sheet_by_name(sheetname) 
              
              #Highlighiting the differece   for LV           
              last_record=sheet1.max_row+1            
              for m in range(2,last_record,5):
#                try:  
                if -30 <= sheet1.cell(row=m,column=5).value <=30:
#                except TypeError:    
                    ft=sheet1.cell(row=m,column=5)  
                    ft.font=fmtB
                    ft.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                    ft1=sheet1.cell(row=m,column=6)
                    ft1.number_format = '0.00%'
                else:
                    ft=sheet1.cell(row=m,column=5)  

                    ft.font=fmtR  
                    ft.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                    ft1=sheet1.cell(row=m,column=6)
                    ft1.number_format = '0.00%'                  
               
              #Alignment for LV                    
              colC=['A','B','C','D','E','F'] 
              for x in range(len(colC)):                  
                  sheet1.column_dimensions[colC[x]].width=24 
                  
              for y in range(2,last_record+7):
                fmt=sheet1.cell(row=y,column=3)      
                fmt.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'  
                
              for y in range(2,last_record+7):
                fmt=sheet1.cell(row=y,column=4)  
                fmt.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'      
                
              for y in range(2,last_record+7):  
                fmt=sheet1.cell(row=y,column=5)
                fmt.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'           
                  
                  
                  
              wb.save(filename=dest_file)         
              return 
          elif (SRC =='LoanFee-533975' or SRC=='LoanFee-534070' or SRC=='TSYS' or SRC=='LoanFee-533901' or SRC=='DepoFee550139' or SRC=='DepoFee550141' or SRC=='DepoFee550557' or SRC=='DepoFee550569' or SRC=='DepoFee552872' or SRC=='DepoFee552877' or SRC=='DF550129and550710'):
              sheetname = [SRC+''+' GL-Recon'] 
          else:
               sheetname = [SRC+''+' GL-Recon', SRCINT+''+' GL-Recon']   
          
          #End of LV code 
          
          

          
          """Following code is common for all the other source systems"""
         
          wb=load_workbook(dest_file)
          
          if (SRC=='MS' or SRC=='FD' or SRC=='IMOD'):          
              sheet2=wb.get_sheet_by_name('Customer Info') 
              colC=['A','B','C','D','E']                  
              for x in range(len(colC)):                  
                sheet2.column_dimensions[colC[x]].width=26  
                
          elif (SRC =='LoanFee-533975' or SRC =='LoanFee-534070'):
              sheet2=wb.get_sheet_by_name(SRC+''+' All Elements') 
              colC=['A','B','C','D','E']    
              for x in range(len(colC)):                  
                sheet2.column_dimensions[colC[x]].width=26               
          else:
              pass     
          
          
          
          for i in range(len(sheetname)):             

              sheet1=wb.get_sheet_by_name(sheetname[i])             
              print(sheetname[i])  
            
              last_record=sheet1.max_row+1 #note removed - 1               
       
              
              for m in range(2,last_record):
        
                try:    
                    if -30 <= sheet1.cell(row=m,column=5).value <=30:  
                        ft=sheet1.cell(row=m,column=5)
                        ft.font=fmtB
                        ft.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                        ft1=sheet1.cell(row=m,column=6)
                        ft1.number_format = '0.00%'
                    else:
                        ft=sheet1.cell(row=m,column=5)
                        ft.font=fmtR  
                        ft.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                        ft1=sheet1.cell(row=m,column=6)
                        ft1.number_format = '0.00%'
                except:                               
                    print("Found Null value")    
               
              #Calculating Grand Total 
              sheet1.cell(row=last_record+3,column=2).value = 'GRAND TOTAL'                
              Bal_total=[]              
              for cl in range(3,6):
                  for rw in range(2,sheet1.max_row-1) :                   
                      a = sheet1.cell(row=rw,column=cl).value
#                      print(rw,cl)
#                      print(sheet1.cell(row=rw,column=cl).value)
                      Bal_total.append(a)
#                      print(Bal_total) 
                  
                  #Code to remove the Null & None from the list
                  Bal_total=[x for x in Bal_total if x is not '']
                  Bal_total=[x for x in Bal_total if x is not None]                  
                  Total=sum(Bal_total)        
                  sheet1.cell(row=last_record+3,column=cl).value = Total
                  
                  #trying to print total percentage
                  if cl==4:
                      MTD=Total
                  elif cl==5:
                      Var=Total
                  else:
                      pass
                  
                  if cl==5:
                      try:
                          perc=Var/MTD    
                          sheet1.cell(row=last_record+3,column=6).value = perc
                      except ZeroDivisionError:
                          perc=0
                  else:
                      pass
                  
                 #formatting the grand total columns
                  ft3=sheet1.cell(row=last_record+3,column=6)
                  ft3.number_format = '0.00%'
                  Bal_total=[] 
                  
                  
              #Column alignment 
              col=['A','B','C','D','E','F']     
              
              for n in range(len(col)):                  
                sheet1.column_dimensions[col[n]].width=25 
              
              #formatting columns
              for y in range(2,last_record+7):
                fmt=sheet1.cell(row=y,column=3)      
                fmt.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'  
                
              for y in range(2,last_record+7):
                fmt=sheet1.cell(row=y,column=4)  
                fmt.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'      
                
              for y in range(2,last_record+7):  
                fmt=sheet1.cell(row=y,column=5)
                fmt.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'        
                
          if SRC=='DF550129and550710':
             sheet1.cell(row=2,column=2).value='550710 and 550129'
             currentCell = sheet1.cell('B2') 
             currentCell.alignment = Alignment(horizontal='right') 
          print(SRC + " Excel file is Saving")      
          wb.save(filename=dest_file) 







if __name__ == '__main__':
    app = QtGui.QApplication(sys.argv)
    GUI=Ui_MainWindow()
    GUI.show()    
    sys.exit(app.exec_())




 

       
